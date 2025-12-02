from logging import config
import re
import MySQLdb
from flask import Flask, Blueprint, jsonify, render_template, request, redirect, url_for, flash, session, current_app
from flask_mysqldb import MySQL
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
from werkzeug.utils import secure_filename
import os 
import uuid


app = Flask(__name__)

# Configure MySQL
mysql = MySQL(app)

from flask import Blueprint, render_template, request, redirect, url_for, flash, session
from werkzeug.security import check_password_hash

auth = Blueprint('auth', __name__)

from flask import request, render_template, redirect, url_for, flash, session
from werkzeug.security import generate_password_hash
from . import mail, create_app
from flask_mail import Message

import random

def generate_otp():
    """Generate a 6-digit OTP code."""
    return str(random.randint(100000, 999999))

import pandas as pd
import secrets
import string
from werkzeug.utils import secure_filename

ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

from flask import request, jsonify, session, current_app, render_template, flash, redirect, url_for
from flask_mail import Message
from threading import Thread
import psycopg2
import psycopg2.extras
import pandas as pd
import secrets
import string
import os
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash


# ========================================================
# ADD ACCOUNT (Single & Bulk)
# ========================================================
@auth.route('/add_account', methods=['GET', 'POST'])
def signup():
    # Role Check: Only Admin (1)
    if 'user_id' in session:
        if session.get('role_id') != 1:
            flash('Access denied. Only admins can create accounts.', 'danger')
            return redirect(url_for('auth.dashboard'))

    conn = current_app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # Fetch programs and year levels for dropdowns
    programs = []
    year_levels = []
    sections = []
    selected_program_id = request.args.get('program_id')

    try:
        cur.execute("SELECT program_id, program_code, program_name FROM programs ORDER BY program_name")
        programs = cur.fetchall()
        
        cur.execute("SELECT year_level_id, year_name FROM year_levels ORDER BY year_level_id")
        year_levels = cur.fetchall()

        if selected_program_id:
            cur.execute("""
                SELECT section_id, program_id, section_name 
                FROM sections 
                WHERE program_id = %s
                ORDER BY section_name
            """, (selected_program_id,))
            sections = cur.fetchall()

        # -------------------------
        # HANDLE POST REQUEST
        # -------------------------
        if request.method == 'POST':
            
            # A. BULK IMPORT
            # -------------------------
            file = request.files.get('bulk_file')
            if file and allowed_file(file.filename): # Ensure allowed_file is defined
                filename = secure_filename(file.filename)
                
                try:
                    df = pd.read_excel(file) if filename.endswith(('xlsx', 'xls')) else pd.read_csv(file)
                    
                    # Clean column names (strip spaces, lowercase)
                    df.columns = [c.strip().lower() for c in df.columns]
                    
                    # Map expected columns
                    required_cols = ['full_name', 'student_id', 'email', 'program_id', 'section_id', 'year_level_id']
                    if not all(col in df.columns for col in required_cols):
                        flash('File must contain columns: full_name, student_id, email, program_id, section_id, year_level_id', 'danger')
                        return redirect(url_for('auth.signup'))

                    new_accounts = [] # List to store info for emails

                    for _, row in df.iterrows():
                        s_id = str(row['student_id']).strip()
                        email = str(row['email']).strip()
                        
                        # Skip existing
                        cur.execute("SELECT 1 FROM users WHERE student_id = %s OR email = %s", (s_id, email))
                        if cur.fetchone():
                            continue

                        # Generate Password
                        characters = string.ascii_letters + string.digits
                        generated_password = ''.join(secrets.choice(characters) for _ in range(10))
                        password_hash = generate_password_hash(generated_password)

                        # Insert User
                        cur.execute("""
                            INSERT INTO users 
                            (full_name, student_id, email, password_hash, role_id, program_id, section_id, year_level_id, approved, email_verified, date_created)
                            VALUES (%s, %s, %s, %s, 3, %s, %s, %s, TRUE, TRUE, NOW())
                            RETURNING user_id
                        """, (
                            row['full_name'], s_id, email, password_hash,
                            row['program_id'], row['section_id'], row['year_level_id']
                        ))
                        user_id = cur.fetchone()['user_id']
                        
                        # Add to list for notification
                        new_accounts.append({
                            'user_id': user_id,
                            'full_name': row['full_name'],
                            'student_id': s_id,
                            'email': email,
                            'password': generated_password
                        })

                    conn.commit()

                    # Start Background Thread for Bulk Emails
                    if new_accounts:
                        app = current_app._get_current_object()
                        
                        def send_bulk_emails_thread(accounts):
                            with app.app_context():
                                thread_conn = app.get_db_connection()
                                thread_cur = thread_conn.cursor()
                                try:
                                    try:
                                        link_url = url_for('auth.login', _external=True)
                                    except:
                                        link_url = "#"

                                    for acc in accounts:
                                        # 1. DB Notification (Welcome) - course_id is None for system alerts
                                        thread_cur.execute("""
                                            INSERT INTO notifications (user_id, course_id, title, message, created_at, read)
                                            VALUES (%s, NULL, %s, %s, NOW(), FALSE)
                                        """, (acc['user_id'], "Welcome to LMS", f"Welcome {acc['full_name']}! Your account has been created."))

                                        # 2. HTML Email
                                        email_html = f"""
                                        <!DOCTYPE html>
                                        <html>
                                        <head>
                                            <style>
                                                body {{ font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif; background-color: #f4f6f8; margin: 0; padding: 0; }}
                                                .email-container {{ max-width: 600px; margin: 30px auto; background-color: #ffffff; border-radius: 8px; overflow: hidden; box-shadow: 0 4px 6px rgba(0,0,0,0.1); border: 1px solid #e0e0e0; }}
                                                .header {{ background-color: #2c3e50; color: #ffffff; padding: 25px; text-align: center; }}
                                                .header h1 {{ margin: 0; font-size: 24px; font-weight: 600; }}
                                                .content {{ padding: 30px; color: #333333; line-height: 1.6; }}
                                                .credentials-box {{ background-color: #ecf0f1; border: 1px solid #bdc3c7; padding: 20px; border-radius: 5px; margin: 20px 0; text-align: center; }}
                                                .credentials-box p {{ margin: 5px 0; font-size: 16px; }}
                                                .credentials-box strong {{ color: #e74c3c; }}
                                                .btn {{ background-color: #3498db; color: white !important; padding: 12px 25px; text-decoration: none; border-radius: 5px; font-weight: bold; display: inline-block; margin-top: 20px; }}
                                                .footer {{ background-color: #eceff1; padding: 15px; text-align: center; font-size: 12px; color: #7f8c8d; }}
                                            </style>
                                        </head>
                                        <body>
                                            <div class="email-container">
                                                <div class="header">
                                                    <h1>Welcome to LMS</h1>
                                                </div>
                                                <div class="content">
                                                    <p>Hi <strong>{acc['full_name']}</strong>,</p>
                                                    <p>Your student account has been successfully created. You can now access your courses and materials online.</p>
                                                    
                                                    <div class="credentials-box">
                                                        <p><strong>Student ID:</strong> {acc['student_id']}</p>
                                                        <p><strong>Password:</strong> {acc['password']}</p>
                                                    </div>
                                                    
                                                    <p style="text-align: center; font-size: 14px; color: #7f8c8d;">Please change your password after your first login.</p>

                                                    <div style="text-align: center;">
                                                        <a href="{link_url}" class="btn">Login to LMS</a>
                                                    </div>
                                                </div>
                                                <div class="footer">
                                                    <p>&copy; 2024 Learning Management System</p>
                                                </div>
                                            </div>
                                        </body>
                                        </html>
                                        """
                                        
                                        msg = Message(
                                            subject="Welcome to LMS - Login Credentials",
                                            recipients=[acc['email']],
                                            html=email_html
                                        )
                                        mail.send(msg)

                                    thread_conn.commit()
                                except Exception as e:
                                    print(f"Bulk email error: {e}")
                                    thread_conn.rollback()
                                finally:
                                    thread_cur.close()
                                    thread_conn.close()

                        Thread(target=send_bulk_emails_thread, args=(new_accounts,)).start()

                    flash(f'Bulk import completed! {len(new_accounts)} accounts created. Login details are being sent via email.', 'success')
                
                except Exception as e:
                    flash(f'Error reading file: {str(e)}', 'danger')
                
                return redirect(url_for('auth.signup'))

            # B. SINGLE ACCOUNT CREATION
            # -------------------------
            full_name = request.form.get('full_name')
            student_id = request.form.get('student_id')
            email = request.form.get('email')
            program_id = request.form.get('program_id')
            section_id = request.form.get('section_id')
            year_level_id = request.form.get('year_level_id')

            if not all([full_name, student_id, email, program_id, section_id, year_level_id]):
                flash('Please fill in all fields.', 'danger')
                return redirect(url_for('auth.signup'))

            cur.execute("SELECT 1 FROM users WHERE student_id = %s OR email = %s", (student_id, email))
            if cur.fetchone():
                flash('Student ID or Email already registered.', 'warning')
                return redirect(url_for('auth.signup'))

            # Generate Password
            generated_password = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(10))
            password_hash = generate_password_hash(generated_password)

            cur.execute("""
                INSERT INTO users 
                (full_name, student_id, email, password_hash, role_id, program_id, section_id, year_level_id, approved, email_verified, date_created)
                VALUES (%s, %s, %s, %s, 3, %s, %s, %s, TRUE, TRUE, NOW())
                RETURNING user_id
            """, (full_name, student_id, email, password_hash, program_id, section_id, year_level_id))
            new_user_id = cur.fetchone()['user_id']
            conn.commit()

            # Background Thread for Single Email
            app = current_app._get_current_object()
            
            def send_single_email():
                with app.app_context():
                    thread_conn = app.get_db_connection()
                    thread_cur = thread_conn.cursor()
                    try:
                        try:
                            link_url = url_for('auth.login', _external=True)
                        except:
                            link_url = "#"

                        # 1. DB Notification
                        thread_cur.execute("""
                            INSERT INTO notifications (user_id, course_id, title, message, created_at, read)
                            VALUES (%s, NULL, %s, %s, NOW(), FALSE)
                        """, (new_user_id, "Welcome to LMS", f"Welcome {full_name}! Your account has been created."))

                        # 2. HTML Email
                        email_html = f"""
                        <!DOCTYPE html>
                        <html>
                        <head>
                            <style>
                                body {{ font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif; background-color: #f4f6f8; margin: 0; padding: 0; }}
                                .email-container {{ max-width: 600px; margin: 30px auto; background-color: #ffffff; border-radius: 8px; overflow: hidden; box-shadow: 0 4px 6px rgba(0,0,0,0.1); border: 1px solid #e0e0e0; }}
                                .header {{ background-color: #2c3e50; color: #ffffff; padding: 25px; text-align: center; }}
                                .header h1 {{ margin: 0; font-size: 24px; font-weight: 600; }}
                                .content {{ padding: 30px; color: #333333; line-height: 1.6; }}
                                .credentials-box {{ background-color: #ecf0f1; border: 1px solid #bdc3c7; padding: 20px; border-radius: 5px; margin: 20px 0; text-align: center; }}
                                .credentials-box p {{ margin: 5px 0; font-size: 16px; }}
                                .credentials-box strong {{ color: #e74c3c; }}
                                .btn {{ background-color: #3498db; color: white !important; padding: 12px 25px; text-decoration: none; border-radius: 5px; font-weight: bold; display: inline-block; margin-top: 20px; }}
                                .footer {{ background-color: #eceff1; padding: 15px; text-align: center; font-size: 12px; color: #7f8c8d; }}
                            </style>
                        </head>
                        <body>
                            <div class="email-container">
                                <div class="header">
                                    <h1>Welcome to LMS</h1>
                                </div>
                                <div class="content">
                                    <p>Hi <strong>{full_name}</strong>,</p>
                                    <p>Your student account has been successfully created. You can now access your courses and materials online.</p>
                                    
                                    <div class="credentials-box">
                                        <p><strong>Student ID:</strong> {student_id}</p>
                                        <p><strong>Password:</strong> {generated_password}</p>
                                    </div>
                                    
                                    <p style="text-align: center; font-size: 14px; color: #7f8c8d;">Please change your password after your first login.</p>

                                    <div style="text-align: center;">
                                        <a href="{link_url}" class="btn">Login to LMS</a>
                                    </div>
                                </div>
                                <div class="footer">
                                    <p>&copy; 2024 Learning Management System</p>
                                </div>
                            </div>
                        </body>
                        </html>
                        """

                        msg = Message(
                            subject="Welcome to LMS - Login Credentials",
                            recipients=[email],
                            html=email_html
                        )
                        mail.send(msg)
                        
                        thread_conn.commit()
                    except Exception as e:
                        print(f"Single email error: {e}")
                        thread_conn.rollback()
                    finally:
                        thread_cur.close()
                        thread_conn.close()

            Thread(target=send_single_email).start()

            flash('Student account created successfully! Login credentials are being sent via email.', 'success')
            return redirect(url_for('auth.signup'))

    except Exception as e:
        conn.rollback()
        flash(f'An error occurred: {str(e)}', 'danger')
    finally:
        cur.close()
        conn.close()

    current_user = None
    if 'user_id' in session:
        current_user = session.get('role_id')

    return render_template(
        'add_account.html',
        programs=programs,
        year_levels=year_levels,
        sections=sections,
        selected_program_id=selected_program_id,
        user=current_user
    )

@auth.route('/get_sections/<int:program_id>')
def get_sections(program_id):
    from . import create_app
    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    
    cur.execute("SELECT section_id, section_name FROM sections WHERE program_id = %s ORDER BY section_name", (program_id,))
    sections = cur.fetchall()
    
    cur.close()
    conn.close()
    
    # Return JSON
    return {"sections": sections}

from flask import request, jsonify, session, current_app, render_template, flash, redirect, url_for
from flask_mail import Message
from threading import Thread
import psycopg2
import psycopg2.extras
import pandas as pd
import secrets
import string
import os
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash


# ========================================================
# ADMIN: ADD TEACHER ACCOUNT (Single & Bulk)
# ========================================================
@auth.route('/add_teacher_account', methods=['GET', 'POST'])
def add_teacher_account():
    # Access Control: Only Admins (Role ID 1)
    if 'user_id' not in session or session.get('role_id') != 1:
        flash('Access denied. Only admins can create teacher accounts.', 'danger')
        return redirect(url_for('auth.dashboard'))

    conn = current_app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # Fetch logged-in admin info
    cur.execute("SELECT * FROM users WHERE user_id = %s", (session['user_id'],))
    user = cur.fetchone()

    try:
        if request.method == 'POST':
            
            # ----------------------------------------
            # A. BULK IMPORT (Excel/CSV)
            # ----------------------------------------
            if 'bulk_file' in request.files and request.files['bulk_file'].filename != '':
                bulk_file = request.files['bulk_file']
                
                try:
                    filename = secure_filename(bulk_file.filename)
                    if filename.endswith(('.xlsx', '.xls')):
                        df = pd.read_excel(bulk_file)
                    elif filename.endswith('.csv'):
                        df = pd.read_csv(bulk_file)
                    else:
                        flash('Invalid file format. Upload Excel or CSV.', 'danger')
                        return redirect(url_for('auth.add_teacher_account'))

                    # Clean columns
                    df.columns = [c.strip().lower() for c in df.columns]

                    # Expected columns
                    required_columns = ['full_name', 'teacher_id', 'email']
                    if not all(col in df.columns for col in required_columns):
                        flash(f'File must contain columns: full_name, teacher_id, email', 'danger')
                        return redirect(url_for('auth.add_teacher_account'))

                    new_teachers = []

                    for _, row in df.iterrows():
                        full_name = str(row['full_name']).strip()
                        teacher_id = str(row['teacher_id']).strip()
                        email = str(row['email']).strip()

                        # Skip if exists
                        cur.execute("SELECT 1 FROM users WHERE student_id = %s OR email = %s", (teacher_id, email))
                        if cur.fetchone():
                            continue

                        # Generate Credentials
                        characters = string.ascii_letters + string.digits
                        generated_password = ''.join(secrets.choice(characters) for _ in range(10))
                        password_hash = generate_password_hash(generated_password)

                        # Insert Teacher (Role 2)
                        # Note: We store teacher_id in 'student_id' column to keep schema simple
                        cur.execute("""
                            INSERT INTO users
                            (full_name, student_id, email, password_hash, role_id, approved, email_verified, must_reset_password, date_created)
                            VALUES (%s, %s, %s, %s, 2, TRUE, TRUE, TRUE, NOW())
                            RETURNING user_id
                        """, (full_name, teacher_id, email, password_hash))
                        new_user_id = cur.fetchone()['user_id']
                        
                        new_teachers.append({
                            'user_id': new_user_id,
                            'full_name': full_name,
                            'teacher_id': teacher_id,
                            'email': email,
                            'password': generated_password
                        })

                    conn.commit()

                    # Threaded Notifications
                    if new_teachers:
                        app = current_app._get_current_object()

                        def send_bulk_teacher_emails(teachers):
                            with app.app_context():
                                thread_conn = app.get_db_connection()
                                thread_cur = thread_conn.cursor()
                                try:
                                    try:
                                        link_url = url_for('auth.login', _external=True)
                                    except:
                                        link_url = "#"

                                    for t in teachers:
                                        # 1. DB Notification
                                        thread_cur.execute("""
                                            INSERT INTO notifications (user_id, course_id, title, message, created_at, read)
                                            VALUES (%s, NULL, %s, %s, NOW(), FALSE)
                                        """, (t['user_id'], "Welcome Instructor", f"Welcome {t['full_name']}! Your teacher account is ready."))

                                        # 2. HTML Email
                                        email_html = f"""
                                        <!DOCTYPE html>
                                        <html>
                                        <head>
                                            <style>
                                                body {{ font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif; background-color: #f4f6f8; margin: 0; padding: 0; }}
                                                .email-container {{ max-width: 600px; margin: 30px auto; background-color: #ffffff; border-radius: 8px; overflow: hidden; box-shadow: 0 4px 6px rgba(0,0,0,0.1); border: 1px solid #e0e0e0; }}
                                                .header {{ background-color: #34495e; color: #ffffff; padding: 25px; text-align: center; }} /* Dark Blue for Teachers */
                                                .header h1 {{ margin: 0; font-size: 24px; font-weight: 600; }}
                                                .content {{ padding: 30px; color: #333333; line-height: 1.6; }}
                                                .credentials-box {{ background-color: #ecf0f1; border: 1px solid #bdc3c7; padding: 20px; border-radius: 5px; margin: 20px 0; text-align: center; }}
                                                .credentials-box strong {{ color: #c0392b; }}
                                                .btn {{ background-color: #2c3e50; color: white !important; padding: 12px 25px; text-decoration: none; border-radius: 5px; font-weight: bold; display: inline-block; margin-top: 20px; }}
                                                .footer {{ background-color: #eceff1; padding: 15px; text-align: center; font-size: 12px; color: #7f8c8d; }}
                                            </style>
                                        </head>
                                        <body>
                                            <div class="email-container">
                                                <div class="header">
                                                    <h1>Welcome Instructor</h1>
                                                </div>
                                                <div class="content">
                                                    <p>Hi <strong>{t['full_name']}</strong>,</p>
                                                    <p>Your instructor account has been created by the administrator.</p>
                                                    
                                                    <div class="credentials-box">
                                                        <p><strong>Teacher ID:</strong> {t['teacher_id']}</p>
                                                        <p><strong>Password:</strong> {t['password']}</p>
                                                    </div>
                                                    
                                                    <p style="text-align: center; font-size: 14px; color: #7f8c8d;">You will be asked to change your password upon your first login.</p>

                                                    <div style="text-align: center;">
                                                        <a href="{link_url}" class="btn">Login to LMS</a>
                                                    </div>
                                                </div>
                                                <div class="footer">
                                                    <p>&copy; 2024 Learning Management System</p>
                                                </div>
                                            </div>
                                        </body>
                                        </html>
                                        """
                                        
                                        msg = Message(
                                            subject="LMS Instructor Account Credentials",
                                            recipients=[t['email']],
                                            html=email_html
                                        )
                                        mail.send(msg)

                                    thread_conn.commit()
                                except Exception as e:
                                    print(f"Bulk teacher email error: {e}")
                                    thread_conn.rollback()
                                finally:
                                    thread_cur.close()
                                    thread_conn.close()

                        Thread(target=send_bulk_teacher_emails, args=(new_teachers,)).start()

                    flash(f'{len(new_teachers)} teacher accounts created successfully! Credentials sent via email.', 'success')
                    return redirect(url_for('auth.add_teacher_account'))

                except Exception as e:
                    flash(f'Error processing file: {str(e)}', 'danger')
                    return redirect(url_for('auth.add_teacher_account'))

            # ----------------------------------------
            # B. SINGLE ACCOUNT CREATION
            # ----------------------------------------
            else:
                full_name = request.form.get('full_name')
                teacher_id = request.form.get('teacher_id')
                email = request.form.get('email')

                if not all([full_name, teacher_id, email]):
                    flash('Please fill in all fields.', 'danger')
                    return redirect(url_for('auth.add_teacher_account'))

                cur.execute("SELECT 1 FROM users WHERE student_id = %s OR email = %s", (teacher_id, email))
                if cur.fetchone():
                    flash('Teacher ID or Email already exists.', 'warning')
                    return redirect(url_for('auth.add_teacher_account'))

                # Generate Password
                characters = string.ascii_letters + string.digits
                generated_password = ''.join(secrets.choice(characters) for _ in range(10))
                password_hash = generate_password_hash(generated_password)

                cur.execute("""
                    INSERT INTO users
                    (full_name, student_id, email, password_hash, role_id, approved, email_verified, must_reset_password, date_created)
                    VALUES (%s, %s, %s, %s, 2, TRUE, TRUE, TRUE, NOW())
                    RETURNING user_id
                """, (full_name, teacher_id, email, password_hash))
                new_user_id = cur.fetchone()['user_id']
                conn.commit()

                # Threaded Single Email
                app = current_app._get_current_object()

                def send_single_teacher_email():
                    with app.app_context():
                        thread_conn = app.get_db_connection()
                        thread_cur = thread_conn.cursor()
                        try:
                            try:
                                link_url = url_for('auth.login', _external=True)
                            except:
                                link_url = "#"

                            # 1. DB Notification
                            thread_cur.execute("""
                                INSERT INTO notifications (user_id, course_id, title, message, created_at, read)
                                VALUES (%s, NULL, %s, %s, NOW(), FALSE)
                            """, (new_user_id, "Welcome Instructor", f"Welcome {full_name}! Your teacher account is ready."))

                            # 2. HTML Email
                            email_html = f"""
                            <!DOCTYPE html>
                            <html>
                            <head>
                                <style>
                                    body {{ font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif; background-color: #f4f6f8; margin: 0; padding: 0; }}
                                    .email-container {{ max-width: 600px; margin: 30px auto; background-color: #ffffff; border-radius: 8px; overflow: hidden; box-shadow: 0 4px 6px rgba(0,0,0,0.1); border: 1px solid #e0e0e0; }}
                                    .header {{ background-color: #34495e; color: #ffffff; padding: 25px; text-align: center; }}
                                    .header h1 {{ margin: 0; font-size: 24px; font-weight: 600; }}
                                    .content {{ padding: 30px; color: #333333; line-height: 1.6; }}
                                    .credentials-box {{ background-color: #ecf0f1; border: 1px solid #bdc3c7; padding: 20px; border-radius: 5px; margin: 20px 0; text-align: center; }}
                                    .credentials-box strong {{ color: #c0392b; }}
                                    .btn {{ background-color: #2c3e50; color: white !important; padding: 12px 25px; text-decoration: none; border-radius: 5px; font-weight: bold; display: inline-block; margin-top: 20px; }}
                                    .footer {{ background-color: #eceff1; padding: 15px; text-align: center; font-size: 12px; color: #7f8c8d; }}
                                </style>
                            </head>
                            <body>
                                <div class="email-container">
                                    <div class="header">
                                        <h1>Welcome Instructor</h1>
                                    </div>
                                    <div class="content">
                                        <p>Hi <strong>{full_name}</strong>,</p>
                                        <p>Your instructor account has been created by the administrator.</p>
                                        
                                        <div class="credentials-box">
                                            <p><strong>Teacher ID:</strong> {teacher_id}</p>
                                            <p><strong>Password:</strong> {generated_password}</p>
                                        </div>
                                        
                                        <p style="text-align: center; font-size: 14px; color: #7f8c8d;">You will be asked to change your password upon your first login.</p>

                                        <div style="text-align: center;">
                                            <a href="{link_url}" class="btn">Login to LMS</a>
                                        </div>
                                    </div>
                                    <div class="footer">
                                        <p>&copy; 2024 Learning Management System</p>
                                    </div>
                                </div>
                            </body>
                            </html>
                            """

                            msg = Message(
                                subject="LMS Instructor Account Credentials",
                                recipients=[email],
                                html=email_html
                            )
                            mail.send(msg)

                            thread_conn.commit()
                        except Exception as e:
                            print(f"Single teacher email error: {e}")
                            thread_conn.rollback()
                        finally:
                            thread_cur.close()
                            thread_conn.close()

                Thread(target=send_single_teacher_email).start()

                flash('Teacher account created! Login credentials sent to email.', 'success')
                return redirect(url_for('auth.add_teacher_account'))

    except Exception as e:
        conn.rollback()
        flash(f'An error occurred: {str(e)}', 'danger')
    finally:
        cur.close()
        conn.close()

    return render_template('add_teacher_account.html', user=user)

# =========================
# CREATE / EDIT COURSE
# =========================
@auth.route('/create_course', methods=['GET', 'POST'])
def create_course():
    if 'user_id' not in session or session.get('role_id') != 1:
        flash("Access denied. Only admins can create courses.", "danger")
        return redirect(url_for('auth.dashboard'))

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # GET edit_course_id from query parameter if editing
    edit_course_id = request.args.get('edit')
    course_to_edit = None
    if edit_course_id:
        cur.execute("SELECT * FROM courses WHERE course_id=%s", (int(edit_course_id),))
        course_to_edit = cur.fetchone()
        if not course_to_edit:
            flash("Course not found for editing.", "warning")
            return redirect(url_for('auth.create_course'))

    # =========================
    # POST: CREATE / EDIT
    # =========================
    if request.method == 'POST':
        data = request.form
        required_fields = [
            'program_id', 'academic_year_id', 'section_id',
            'course_code', 'course_title', 'year_level_id',
            'units', 'user_id'  # teacher's user_id
        ]

        if not all(data.get(f) for f in required_fields):
            flash("All fields are required.", "danger")
            return redirect(url_for('auth.create_course', edit=data.get('course_id')))

        try:
            # Cast form values
            program_id = int(data['program_id'])
            academic_year_id = int(data['academic_year_id'])
            section_id = int(data['section_id'])
            year_level_id = int(data['year_level_id'])
            units = float(data['units'])
            teacher_user_id = int(data['user_id'])

            # Set status to 'approved' for every new/updated course
            status = 'approved'
            approved = True  # boolean column

            if data.get('course_id'):  # UPDATE
                course_id = int(data['course_id'])

                cur.execute("""
                    UPDATE courses
                    SET program_id=%s, academic_year_id=%s, section_id=%s,
                        course_code=%s, course_title=%s, year_level_id=%s,
                        units=%s, user_id=%s, status=%s, approved=%s
                    WHERE course_id=%s
                """, (
                    program_id, academic_year_id, section_id,
                    data['course_code'], data['course_title'], year_level_id,
                    units, teacher_user_id, status, approved, course_id
                ))
                flash("✅ Course updated successfully!", "success")

            else:  # CREATE
                cur.execute("""
                    INSERT INTO courses (
                        program_id, academic_year_id, section_id,
                        course_code, course_title, year_level_id, units,
                        status, date_created, user_id, approved
                    )
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,NOW(),%s,%s)
                """, (
                    program_id, academic_year_id, section_id,
                    data['course_code'], data['course_title'], year_level_id, units,
                    status, teacher_user_id, approved
                ))
                flash("✅ Course created successfully!", "success")

            conn.commit()

        except Exception as e:
            conn.rollback()
            flash(f"Server error: {str(e)}", "danger")
        finally:
            cur.close()
            conn.close()

        return redirect(url_for('auth.create_course'))


    # =========================
    # GET: Fetch dropdown lists
    # =========================
    cur.execute("SELECT * FROM programs ORDER BY program_name")
    programs = cur.fetchall()

    cur.execute("SELECT * FROM academic_years ORDER BY start_year DESC")
    academic_years = cur.fetchall()

    cur.execute("SELECT * FROM sections ORDER BY section_name")
    sections = cur.fetchall()

    cur.execute("SELECT * FROM year_levels ORDER BY year_level_id")
    year_levels = cur.fetchall()

    # Fetch teachers (users with role_id = 2)
    cur.execute("""
        SELECT user_id, full_name
        FROM users
        WHERE role_id = 2
        ORDER BY full_name
    """)
    teachers = cur.fetchall()

    # Fetch existing courses
    cur.execute("""
        SELECT c.*, 
               p.program_name, s.section_name, y.year_name AS year_level,
               a.start_year, a.end_year,
               t.full_name AS teacher_name
        FROM courses c
        LEFT JOIN programs p ON c.program_id = p.program_id
        LEFT JOIN sections s ON c.section_id = s.section_id
        LEFT JOIN year_levels y ON c.year_level_id = y.year_level_id
        LEFT JOIN academic_years a ON c.academic_year_id = a.academic_year_id
        LEFT JOIN users t ON c.user_id = t.user_id
        ORDER BY c.course_id DESC
    """)
    courses = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "create_course.html",
        programs=programs,
        academic_years=academic_years,
        sections=sections,
        year_levels=year_levels,
        teachers=teachers,
        courses=courses,
        course_to_edit=course_to_edit
    )

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@auth.route('/bulk_import_courses', methods=['GET', 'POST'])
def bulk_import_courses():
    if 'user_id' not in session or session.get('role_id') != 1:
        flash("Access denied. Only admins can import courses.", "danger")
        return redirect(url_for('auth.dashboard'))

    if request.method == 'POST':
        if 'file' not in request.files:
            flash("No file part", "danger")
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash("No selected file", "danger")
            return redirect(request.url)
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join("/tmp", filename)
            file.save(filepath)

            try:
                # Read Excel
                df = pd.read_excel(filepath)

                required_columns = ['program_id', 'academic_year_id', 'section_id', 
                                    'course_code', 'course_title', 'year_level_id', 
                                    'units', 'user_id']
                for col in required_columns:
                    if col not in df.columns:
                        flash(f"Missing required column: {col}", "danger")
                        return redirect(request.url)

                app = create_app()
                conn = app.get_db_connection()
                cur = conn.cursor()

                inserted_count = 0
                for _, row in df.iterrows():
                    # Convert types
                    program_id = int(row['program_id'])
                    academic_year_id = int(row['academic_year_id'])
                    section_id = int(row['section_id'])
                    year_level_id = int(row['year_level_id'])
                    units = float(row['units'])
                    teacher_user_id = int(row['user_id'])
                    status = 'approved'
                    approved = True

                    # Prevent duplicate section/program/year
                    cur.execute("""
                        SELECT * FROM courses
                        WHERE program_id=%s AND academic_year_id=%s AND section_id=%s
                    """, (program_id, academic_year_id, section_id))
                    if cur.fetchone():
                        continue  # skip duplicates

                    cur.execute("""
                        INSERT INTO courses (
                            program_id, academic_year_id, section_id,
                            course_code, course_title, year_level_id, units,
                            status, date_created, user_id, approved
                        )
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,NOW(),%s,%s)
                    """, (
                        program_id, academic_year_id, section_id,
                        row['course_code'], row['course_title'], year_level_id, units,
                        status, teacher_user_id, approved
                    ))
                    inserted_count += 1

                conn.commit()
                flash(f"✅ Successfully imported {inserted_count} courses!", "success")

            except Exception as e:
                conn.rollback()
                flash(f"Server error: {str(e)}", "danger")
            finally:
                cur.close()
                conn.close()
                os.remove(filepath)

            return redirect(url_for('auth.bulk_import_courses'))

        else:
            flash("Invalid file type. Please upload an Excel file.", "danger")
            return redirect(request.url)

    return render_template("bulk_import_courses.html")


# =========================
# DELETE COURSE
# =========================
@auth.route('/delete-course/<int:course_id>', methods=['POST'])
def delete_course(course_id):
    if 'user_id' not in session or session.get('role_id') != 1:
        flash("Access denied. Only admins can delete courses.", "danger")
        return redirect(url_for('auth.dashboard'))

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor()

    try:
        # Check if course exists
        cur.execute("SELECT * FROM courses WHERE course_id=%s", (course_id,))
        course = cur.fetchone()
        if not course:
            flash("❌ Course not found.", "warning")
            return redirect(url_for('auth.create_course'))

        # Delete the course
        cur.execute("DELETE FROM courses WHERE course_id=%s", (course_id,))
        conn.commit()
        flash("✅ Course deleted successfully!", "success")

    except Exception as e:
        conn.rollback()
        flash(f"Server error: {str(e)}", "danger")

    finally:
        cur.close()
        conn.close()

    return redirect(url_for('auth.create_course'))

# =========================
# MANAGE STUDENTS PAGE
# =========================
@auth.route('/course/<int:course_id>/manage', methods=['GET'])
def manage_students_page(course_id):
    if 'user_id' not in session or session.get('role_id') != 1:
        flash("Access denied. Only admins can manage students.", "danger")
        return redirect(url_for('auth.create_course'))

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # Fetch course info
    cur.execute("SELECT * FROM courses WHERE course_id=%s", (course_id,))
    course = cur.fetchone()
    if not course:
        flash("Course not found.", "warning")
        cur.close()
        conn.close()
        return redirect(url_for('auth.create_course'))

    # Filters
    selected_program_id = request.args.get('program_id')
    selected_section_id = request.args.get('section_id')
    search_query = request.args.get('search', '')

    # Pagination
    page = int(request.args.get('page', 1))
    per_page = 10
    offset = (page - 1) * per_page

    # Fetch students with Program Name and Section Name
    query = """
        SELECT u.user_id, u.full_name, u.year_level_id,
               p.program_name,
               s.section_name
        FROM users u
        LEFT JOIN programs p ON u.program_id = p.program_id
        LEFT JOIN sections s ON u.section_id = s.section_id
        WHERE u.role_id = 3
    """
    params = []

    if selected_program_id and selected_program_id != "all":
        query += " AND u.program_id = %s"
        params.append(selected_program_id)
    if selected_section_id and selected_section_id != "all":
        query += " AND u.section_id = %s"
        params.append(selected_section_id)
    if search_query:
        query += " AND u.full_name ILIKE %s"
        params.append(f"%{search_query}%")

    # Count total students for pagination
    total_query = f"SELECT COUNT(*) FROM ({query}) AS total_students"
    cur.execute(total_query, params)
    total_students = cur.fetchone()['count']

    # Fetch paginated results
    query += " ORDER BY u.full_name LIMIT %s OFFSET %s"
    params.extend([per_page, offset])
    cur.execute(query, params)
    all_students = cur.fetchall()

    # Fetch enrolled students for this course
    cur.execute("SELECT user_id FROM course_students WHERE course_id=%s", (course_id,))
    enrolled_students = cur.fetchall()
    enrolled_students_ids = [s['user_id'] for s in enrolled_students]

    # Fetch programs and sections for filter dropdowns
    cur.execute("SELECT program_id, program_name FROM programs ORDER BY program_name")
    programs = cur.fetchall()
    cur.execute("SELECT section_id, section_name FROM sections ORDER BY section_name")
    sections = cur.fetchall()

    cur.close()
    conn.close()

    total_pages = (total_students + per_page - 1) // per_page

    return render_template(
        "manage_course_students.html",
        course=course,
        all_students=all_students,
        enrolled_students_ids=enrolled_students_ids,
        programs=programs,
        sections=sections,
        selected_program_id=selected_program_id,
        selected_section_id=selected_section_id,
        search_query=search_query,
        page=page,
        total_pages=total_pages
    )

# =========================
# ADD/REMOVE STUDENT
# =========================
@auth.route('/course/<int:course_id>/students', methods=['POST'])
def manage_course_students(course_id):
    if 'user_id' not in session or session.get('role_id') != 1:
        flash("Access denied. Only admins can manage students.", "danger")
        return redirect(url_for('auth.create_course'))

    student_id = request.form.get('user_id')
    action = request.form.get('action')

    if not student_id or action not in ['add', 'remove']:
        flash("Invalid student or action.", "warning")
        return redirect(url_for('auth.manage_students_page', course_id=course_id))

    student_id = int(student_id)
    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor()

    try:
        if action == 'add':
            cur.execute("""
                INSERT INTO course_students (course_id, user_id)
                SELECT %s, %s
                WHERE NOT EXISTS (
                    SELECT 1 FROM course_students WHERE course_id=%s AND user_id=%s
                )
            """, (course_id, student_id, course_id, student_id))
            flash("✅ Student added successfully.", "success")
        elif action == 'remove':
            cur.execute("""
                DELETE FROM course_students
                WHERE course_id=%s AND user_id=%s
            """, (course_id, student_id))
            flash("✅ Student removed successfully.", "success")
        conn.commit()
    except Exception as e:
        conn.rollback()
        flash(f"Server error: {str(e)}", "danger")
    finally:
        cur.close()
        conn.close()

    return redirect(url_for('auth.manage_students_page', course_id=course_id))

@auth.route('/login', methods=['GET', 'POST'])
def login():
    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    if request.method == 'POST':
        # Get form data
        identifier = request.form.get('identifier', '').strip()  # Student ID or Teacher Email
        password = request.form.get('password', '').strip()
        captcha = request.form.get('captcha', '').strip()
        captcha_generated = request.form.get('captchaGenerated', '').strip()

        # -----------------------
        # Validate Inputs
        # -----------------------
        if not identifier:
            flash('Student ID or Email is required', 'identifierError')
            return redirect(url_for('auth.login'))

        if not password:
            flash('Password is required', 'passwordError')
            return redirect(url_for('auth.login'))

        if not captcha:
            flash('Security code is required', 'captchaError')
            return redirect(url_for('auth.login'))
        elif captcha.upper() != captcha_generated.upper():
            flash('Security code is incorrect', 'captchaError')
            return redirect(url_for('auth.login'))

        # -----------------------
        # Fetch user from DB
        # -----------------------
        if '@' in identifier:  # Teacher login
            cur.execute("""
                SELECT user_id, full_name, email, password_hash, role_id, email_verified, must_reset_password
                FROM users
                WHERE email = %s AND role_id = 2
            """, (identifier,))
        else:  # Student login
            cur.execute("""
                SELECT user_id, full_name, student_id, email, password_hash, role_id, email_verified, approved, must_reset_password
                FROM users
                WHERE student_id = %s AND role_id = 3
            """, (identifier,))

        user = cur.fetchone()

        if not user:
            flash('No account found with that ID or email.', 'identifierError')
            return redirect(url_for('auth.login'))

        # -----------------------
        # Check email verification
        # -----------------------
        if not user['email_verified']:
            flash('Your email has not been verified. Please check your email for the OTP.', 'identifierError')
            return redirect(url_for('auth.login'))

        # -----------------------
        # Check account approval (students only)
        # -----------------------
        if user['role_id'] == 3 and not user['approved']:
            flash('Your account is not yet approved by the admin. Please wait.', 'identifierError')
            return redirect(url_for('auth.login'))

        # -----------------------
        # Check password
        # -----------------------
        if not check_password_hash(user['password_hash'], password):
            flash('Incorrect password. Please try again.', 'passwordError')
            return redirect(url_for('auth.login'))

        # -----------------------
        # Successful login
        # -----------------------
        session['user_id'] = user['user_id']
        session['full_name'] = user['full_name']
        session['role_id'] = user['role_id']

        # -----------------------
        # Redirect based on role and must_reset_password
        # -----------------------
        if user['must_reset_password']:
            return redirect(url_for('auth.reset_initial_password'))

        if user['role_id'] == 3:  # Student
            return redirect(url_for('auth.dashboard'))

        if user['role_id'] == 2:  # Teacher
            return redirect(url_for('auth.teachers_dashboard'))

    cur.close()
    conn.close()
    return render_template('login.html')

@auth.route('/reset_initial_password', methods=['GET', 'POST'])
def reset_initial_password():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    user_id = session['user_id']

    # Fetch role of user
    cur.execute("SELECT role_id FROM users WHERE user_id = %s", (user_id,))
    user = cur.fetchone()
    if not user:
        flash("User not found.", "danger")
        return redirect(url_for('auth.login'))

    role_id = user['role_id']

    if request.method == 'POST':
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')

        if new_password != confirm_password:
            flash("Passwords do not match.", "danger")
            return redirect(url_for('auth.reset_initial_password'))

        if len(new_password) < 6:
            flash("Password must be at least 6 characters.", "danger")
            return redirect(url_for('auth.reset_initial_password'))

        new_hash = generate_password_hash(new_password)

        # Update password + disable reset requirement
        cur.execute("""
            UPDATE users
            SET password_hash = %s, must_reset_password = FALSE
            WHERE user_id = %s
        """, (new_hash, user_id))
        conn.commit()

        flash("Password updated successfully!", "success")

        # Redirect based on role
        if role_id == 3:  # Student
            return redirect(url_for('auth.dashboard'))
        elif role_id == 2:  # Teacher
            return redirect(url_for('auth.teachers_dashboard'))
        else:
            # Optional: for any other role, redirect to login
            return redirect(url_for('auth.login'))

    # GET: Show reset page
    return render_template("reset_initial_password.html")

@auth.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    from . import create_app
    import secrets
    from datetime import datetime, timedelta
    from flask_mail import Message
    from . import mail

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        if not email:
            flash('Email is required', 'danger')
            return redirect(url_for('auth.forgot_password'))

        # Check if user exists
        cur.execute("SELECT user_id, full_name FROM users WHERE email = %s", (email,))
        user = cur.fetchone()
        if not user:
            flash('No account found with that email.', 'danger')
            return redirect(url_for('auth.forgot_password'))

        # Generate token
        token = secrets.token_urlsafe(16)
        expiry = datetime.now() + timedelta(minutes=30)

        # Store token in DB
        cur.execute("""
            UPDATE users SET reset_token = %s, reset_token_expiry = %s WHERE user_id = %s
        """, (token, expiry, user['user_id']))
        conn.commit()

        # Send reset email
        reset_link = url_for('auth.reset_password', token=token, _external=True)
        msg = Message(
            subject="LMS Password Reset",
            recipients=[email],
            html=f"""
                <p>Hello {user['full_name']},</p>
                <p>Click the link below to reset your password (valid for 30 minutes):</p>
                <a href="{reset_link}" style="display:inline-block;padding:10px 15px;background:#3B82F6;color:white;border-radius:5px;text-decoration:none;">Reset Password</a>
                <p>If you did not request a password reset, ignore this email.</p>
            """
        )
        mail.send(msg)
        flash('Password reset link sent to your email!', 'success')
        return redirect(url_for('auth.login'))

    cur.close()
    conn.close()
    return render_template('forgot_password.html')

@auth.route('/reset_password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    from . import create_app
    from werkzeug.security import generate_password_hash
    import psycopg2.extras
    from flask import request, flash, redirect, url_for, render_template

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # Verify token exists and is valid
    cur.execute("""
        SELECT * FROM users 
        WHERE reset_token = %s AND reset_token_expiry > NOW()
    """, (token,))
    user = cur.fetchone()

    if not user:
        flash("Invalid or expired reset link.", "danger")
        cur.close()
        conn.close()
        return redirect(url_for('auth.login'))

    if request.method == 'POST':
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')

        if not new_password or not confirm_password:
            flash("Please fill in all fields.", "danger")
        elif new_password != confirm_password:
            flash("Passwords do not match.", "danger")
        else:
            password_hash = generate_password_hash(new_password)
            # Update password and remove token
            cur.execute("""
                UPDATE users 
                SET password_hash = %s, reset_token = NULL, reset_token_expiry = NULL
                WHERE user_id = %s
            """, (password_hash, user['user_id']))
            conn.commit()
            flash("Password reset successfully! You can now log in.", "success")
            cur.close()
            conn.close()
            return redirect(url_for('auth.login'))

    cur.close()
    conn.close()
    return render_template("reset_password.html", email=user['email'])

@auth.route('/admin-login', methods=['GET', 'POST'])
def adminlogin():
    from werkzeug.security import check_password_hash
    import requests
    from . import create_app

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor()

    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '').strip()
        captcha = request.form.get('captcha', '').strip()
        captcha_generated = request.form.get('captchaGenerated', '').strip()

        # ✅ Validate Email
        if not email:
            flash('Email is required', 'emailError')
            return redirect(url_for('auth.adminlogin'))

        # ✅ Validate Password
        if not password:
            flash('Password is required', 'passwordError')
            return redirect(url_for('auth.adminlogin'))

        # ✅ Validate Captcha
        if not captcha:
            flash('Security code is required', 'captchaError')
            return redirect(url_for('auth.adminlogin'))
        elif captcha.upper() != captcha_generated.upper():
            flash('Security code is incorrect', 'captchaError')
            return redirect(url_for('auth.adminlogin'))

        # ✅ Check credentials in DB by email
        cur.execute("""
            SELECT user_id, full_name, email, password_hash, role_id 
            FROM users WHERE email = %s
        """, (email,))
        user = cur.fetchone()

        if not user:
            flash('No account found with that email.', 'emailError')
            return redirect(url_for('auth.adminlogin'))

        if not check_password_hash(user[3], password):
            flash('Incorrect password. Please try again.', 'passwordError')
            return redirect(url_for('auth.adminlogin'))

        # 🔒 Ensure admin role (role_id = 1)
        if user[4] != 1:
            flash('Access denied. Admin credentials required.', 'emailError')
            return redirect(url_for('auth.adminlogin'))

        # ✅ Successful login
        session['user_id'] = user[0]
        session['full_name'] = user[1]
        session['role_id'] = user[4]

        flash('Login successful! Redirecting to admin dashboard...', 'success')
        return redirect(url_for('auth.admindashboard'))

    # 🧹 Close DB connection
    cur.close()
    conn.close()

    return render_template('admin_login.html')

@auth.route('/admin-dashboard')
def admindashboard():
    # ✅ Restrict access to admin users only
    if 'user_id' not in session or session.get('role_id') != 1:
        flash('Please log in as an admin to access this page.', 'warning')
        return redirect(url_for('auth.adminlogin'))

    from . import create_app
    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor()

    user_id = session['user_id']

    # 🧩 Fetch admin info
    cur.execute("""
        SELECT user_id, full_name, email 
        FROM users 
        WHERE user_id = %s
    """, (user_id,))
    user = cur.fetchone()

    if not user:
        flash('User not found.', 'danger')
        cur.close()
        conn.close()
        return redirect(url_for('auth.adminlogin'))

    # 🧩 Fetch all users (for totals)
    cur.execute("SELECT COUNT(*) FROM users WHERE role_id IN (2,3)")
    total_users = cur.fetchone()[0] or 1  # avoid division by zero

    # 🧩 Fetch ALL unapproved users (teachers + students)
    cur.execute("""
        SELECT user_id, full_name, email, role_id, approved, approved_at
        FROM users 
        WHERE (role_id = 2 OR role_id = 3)
        AND approved = FALSE
        ORDER BY user_id DESC
    """)
    unapproved_users = cur.fetchall()

    # 🧩 Fetch all approved teachers
    cur.execute("""
        SELECT user_id, full_name, email, approved, approved_at
        FROM users 
        WHERE role_id = 2 AND approved = TRUE
        ORDER BY full_name ASC
    """)
    teachers = cur.fetchall()

    # 🧩 Fetch all approved students
    cur.execute("""
        SELECT user_id, full_name, email, approved, approved_at
        FROM users 
        WHERE role_id = 3 AND approved = TRUE
        ORDER BY full_name ASC
    """)
    students = cur.fetchall()

    cur.execute("""
    SELECT COUNT(*) 
    FROM courses 
    WHERE status ILIKE 'approved'
    """)
    active_courses = cur.fetchone()[0]

    # 🧩 Fetch total courses
    cur.execute("SELECT COUNT(*) FROM courses")
    total_courses = cur.fetchone()[0] or 1  # avoid division by zero

    cur.close()
    conn.close()

    # ✅ Compute percentages
    total_students = len(students)
    total_teachers = len(teachers)
    total_pending = len(unapproved_users)

    student_percent = round((total_students / total_users) * 100, 1)
    teacher_percent = round((total_teachers / total_users) * 100, 1)
    pending_percent = round((total_pending / total_users) * 100, 1)

    # ✅ Active course % (dynamic)
    active_course_percent = round((active_courses / total_courses) * 100, 1)

    return render_template(
        'admin_dashboard.html',
        user={
            'user_id': user[0],
            'full_name': user[1],
            'email': user[2]
        },
        unapproved_users=unapproved_users,
        teachers=teachers,
        students=students,
        stats={
            'total_students': total_students,
            'total_teachers': total_teachers,
            'total_pending': total_pending,
            'student_percent': student_percent,
            'teacher_percent': teacher_percent,
            'pending_percent': pending_percent,
            'active_courses': active_courses,
            'active_course_percent': active_course_percent
        }
    )

# ======================================================
# ADMIN LMS - SHOW ALL COURSES (MANAGE)
# ======================================================
@auth.route('/admin_lms', methods=['GET'])
def admin_lms():
    # Require login
    if 'user_id' not in session:
        flash('Please log in first.', 'warning')
        return redirect(url_for('auth.login'))

    # Only allow admins
    if session.get('role_id') != 1:  # assuming 1 = admin
        flash('Access denied. Admins only.', 'danger')
        return redirect(url_for('auth.login'))

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # 🧑‍💼 Get current admin info
    user_id = session['user_id']
    cur.execute("""
        SELECT user_id, full_name, email, role_id
        FROM users
        WHERE user_id = %s
    """, (user_id,))
    user = cur.fetchone()  # this is the admin user object

    # 📚 Fetch dropdown options for filters
    cur.execute("SELECT program_id, program_name FROM programs ORDER BY program_name;")
    programs = cur.fetchall()

    cur.execute("""
        SELECT academic_year_id, start_year, end_year, semester
        FROM academic_years
        ORDER BY start_year DESC, semester;
    """)
    academic_years = cur.fetchall()

    cur.execute("SELECT section_id, section_name FROM sections ORDER BY section_name;")
    sections = cur.fetchall()

    # 🧾 Filters from request
    selected_program_id = request.args.get('program_id')
    selected_academic_year_id = request.args.get('academic_year_id')
    selected_semester = request.args.get('semester')
    selected_status = request.args.get('status')  # pending, approved, rejected

    # 🧠 Fetch all courses
    query = """
        SELECT c.course_id, c.course_code, c.course_title, c.course_description,
               c.units, c.year_level_id, c.status,
               u.full_name AS teacher_name, u.email AS teacher_email,
               p.program_name, s.section_name,
               CONCAT(a.start_year, '-', a.end_year, ' (', a.semester, ' Semester)') AS academic_year
        FROM courses c
        JOIN users u ON c.user_id = u.user_id
        JOIN programs p ON c.program_id = p.program_id
        JOIN academic_years a ON c.academic_year_id = a.academic_year_id
        LEFT JOIN sections s ON c.section_id = s.section_id
        WHERE 1=1
    """
    params = []

    if selected_program_id and selected_program_id != "all":
        query += " AND c.program_id = %s"
        params.append(selected_program_id)
    if selected_academic_year_id and selected_academic_year_id != "all":
        query += " AND c.academic_year_id = %s"
        params.append(selected_academic_year_id)
    if selected_semester and selected_semester != "all":
        query += " AND a.semester = %s"
        params.append(selected_semester)
    if selected_status and selected_status != "all":
        query += " AND c.status = %s"
        params.append(selected_status)

    query += " ORDER BY c.status, c.year_level_id, c.course_code;"
    cur.execute(query, params)
    courses = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        'admin_lms.html',
        user=user,  # <-- pass admin user info to template
        programs=programs,
        academic_years=academic_years,
        sections=sections,
        courses=courses,
        selected_program_id=selected_program_id,
        selected_academic_year_id=selected_academic_year_id,
        selected_semester=selected_semester,
        selected_status=selected_status
    )

@auth.route('/admin_lms_export', methods=['GET'])
def admin_lms_export():
    if 'user_id' not in session or session.get('role_id') != 1:
        flash('Access denied', 'danger')
        return redirect(url_for('auth.login'))

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    selected_program_id = request.args.get('program_id')
    selected_academic_year_id = request.args.get('academic_year_id')
    selected_semester = request.args.get('semester')
    selected_status = request.args.get('status')

    query = """
        SELECT c.course_id, c.course_code, c.course_title, c.course_description,
               c.units, c.year_level_id, c.status,
               u.full_name AS teacher_name,
               p.program_name, s.section_name,
               CONCAT(a.start_year, '-', a.end_year, ' (', a.semester, ' Semester)') AS academic_year
        FROM courses c
        JOIN users u ON c.user_id = u.user_id
        JOIN programs p ON c.program_id = p.program_id
        JOIN academic_years a ON c.academic_year_id = a.academic_year_id
        LEFT JOIN sections s ON c.section_id = s.section_id
        WHERE 1=1
    """
    params = []

    if selected_program_id and selected_program_id != "all":
        query += " AND c.program_id = %s"
        params.append(selected_program_id)
    if selected_academic_year_id and selected_academic_year_id != "all":
        query += " AND c.academic_year_id = %s"
        params.append(selected_academic_year_id)
    if selected_semester and selected_semester != "all":
        query += " AND a.semester = %s"
        params.append(selected_semester)
    if selected_status and selected_status != "all":
        query += " AND c.status = %s"
        params.append(selected_status)

    query += " ORDER BY c.status, c.year_level_id, c.course_code;"

    cur.execute(query, params)
    courses = cur.fetchall()

    cur.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Courses"

    headers = [
        "Course ID", "Course Code", "Course Title", "Description",
        "Units", "Year Level", "Status",
        "Teacher", "Program", "Section", "Academic Year"
    ]
    ws.append(headers)

    for c in courses:
        ws.append([
            c['course_id'],
            c['course_code'],
            c['course_title'],
            c['course_description'],
            c['units'],
            c['year_level_id'],
            c['status'],
            c['teacher_name'],
            c['program_name'],
            c['section_name'],
            c['academic_year']
        ])

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name="courses.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@auth.route('/student_list', methods=['GET'])
def student_list():
    if 'user_id' not in session or session.get('role_id') != 1:
        flash('Access denied', 'danger')
        return redirect(url_for('auth.login'))

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        # Fetch students with program, section, AND year_level info
        cur.execute("""
            SELECT 
                u.user_id, 
                u.full_name, 
                u.email, 
                u.status,
                u.date_created,
                u.student_id,
                u.year_level_id,
                yl.year_name,       -- Make sure HTML uses {{ student.year_name }}
                p.program_id,
                p.program_name,
                p.program_code,
                s.section_id,
                s.section_name
            FROM users u
            LEFT JOIN programs p ON u.program_id = p.program_id
            LEFT JOIN sections s ON u.section_id = s.section_id
            LEFT JOIN year_levels yl ON u.year_level_id = yl.year_level_id
            WHERE u.role_id = 3
              AND (u.approved = TRUE)
            ORDER BY u.full_name;
        """)
        students = cur.fetchall()

        # Fetch all programs
        cur.execute("""
            SELECT program_id, program_name
            FROM programs
            ORDER BY program_name;
        """)
        programs = cur.fetchall()

        # Fetch all year_levels
        cur.execute("""
            SELECT year_level_id, year_name
            FROM year_levels
            ORDER BY year_level_id;
        """)
        year_levels = cur.fetchall()

        # Dashboard stats
        total_active = sum(1 for s in students if s['status'] == 'active')
        inactive_students = sum(1 for s in students if s['status'] == 'inactive')
        pending_verifications = sum(1 for s in students if s['status'] == 'pending')

        from datetime import datetime, timedelta
        one_week_ago = datetime.now() - timedelta(days=7)
        new_enrollments = sum(1 for s in students if s['date_created'] >= one_week_ago)

    finally:
        cur.close()
        conn.close()

    return render_template(
        'student_list.html',
        students=students,
        programs=programs,
        year_levels=year_levels,
        total_active=total_active,
        inactive_students=inactive_students,
        pending_verifications=pending_verifications,
        new_enrollments=new_enrollments
    )

from openpyxl import Workbook
from flask import send_file
import io

@auth.route('/student_list_export', methods=['GET'])
def student_list_export():
    if 'user_id' not in session or session.get('role_id') != 1:
        flash('Access denied', 'danger')
        return redirect(url_for('auth.login'))

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        # Fetch all approved students
        cur.execute("""
            SELECT 
                u.user_id, 
                u.full_name, 
                u.email, 
                u.status,
                u.date_created,
                u.student_id,
                p.program_name,
                p.program_code,
                s.section_name
            FROM users u
            LEFT JOIN programs p ON u.program_id = p.program_id
            LEFT JOIN sections s ON u.section_id = s.section_id
            WHERE u.role_id = 3
              AND (u.approved = TRUE)
            ORDER BY u.full_name;
        """)
        students = cur.fetchall()

    finally:
        cur.close()
        conn.close()

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Student List"

    # Header row
    headers = [
        "User ID", "Student ID", "Full Name", "Email", "Status",
        "Program", "Program Code", "Section", "Date Created"
    ]
    ws.append(headers)

    # Data rows
    for s in students:
        ws.append([
            s['user_id'],
            s['student_id'],
            s['full_name'],
            s['email'],
            s['status'],
            s['program_name'],
            s['program_code'],
            s['section_name'],
            s['date_created'].strftime("%Y-%m-%d %H:%M:%S")
        ])

    # Save to memory
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name="student_list.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ==========================
# ACTIVATE STUDENT
# ==========================
@auth.route('/activate_student/<int:user_id>', methods=['POST'])
def activate_student(user_id):
    if 'user_id' not in session or session.get('role_id') != 1:
        return jsonify({"success": False, "message": "Access denied"}), 403

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        cur.execute("""
            UPDATE users
            SET status = 'active'
            WHERE user_id = %s AND role_id = 3
            RETURNING status;
        """, (user_id,))
        updated = cur.fetchone()
        conn.commit()
        if updated:
            return jsonify({
                "success": True,
                "message": "✅ Student activated successfully!",
                "status": updated['status']
            })
        return jsonify({"success": False, "message": "Student not found."}), 404
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "message": f"Error: {str(e)}"}), 500
    finally:
        cur.close()
        conn.close()

# ==========================
# DEACTIVATE STUDENT
# ==========================
@auth.route('/deactivate_student/<int:user_id>', methods=['POST'])
def deactivate_student(user_id):
    if 'user_id' not in session or session.get('role_id') != 1:
        return jsonify({"success": False, "message": "Access denied"}), 403

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        cur.execute("""
            UPDATE users
            SET status = 'inactive'
            WHERE user_id = %s AND role_id = 3
            RETURNING status;
        """, (user_id,))
        updated = cur.fetchone()
        conn.commit()
        if updated:
            return jsonify({
                "success": True,
                "message": "✅ Student deactivated successfully!",
                "status": updated['status']
            })
        return jsonify({"success": False, "message": "Student not found."}), 404
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "message": f"Error: {str(e)}"}), 500
    finally:
        cur.close()
        conn.close()

# ==========================
# EDIT STUDENT (VIEW FORM)
# ==========================
@auth.route('/edit_student/<int:user_id>', methods=['GET'])
def edit_student(user_id):
    if 'user_id' not in session or session.get('role_id') != 1:
        flash("Access denied", "danger")
        return redirect(url_for('auth.login'))

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        # Fetch student info with program and section
        cur.execute("""
            SELECT 
                u.user_id,
                u.full_name,
                u.email,
                u.student_id,
                u.status,
                u.program_id,
                u.section_id,
                p.program_name,
                s.section_name
            FROM users u
            LEFT JOIN programs p ON u.program_id = p.program_id
            LEFT JOIN sections s ON u.section_id = s.section_id
            WHERE u.user_id = %s AND u.role_id = 3;
        """, (user_id,))
        student = cur.fetchone()

        if not student:
            flash("Student not found.", "danger")
            return redirect(url_for('student_list'))

        # Fetch all programs for dropdown
        cur.execute("SELECT program_id, program_name FROM programs ORDER BY program_name;")
        programs = cur.fetchall()

        # Fetch all sections for student's program
        sections = []
        if student['program_id']:
            cur.execute("""
                SELECT section_id, section_name 
                FROM sections 
                WHERE program_id = %s
                ORDER BY section_name;
            """, (student['program_id'],))
            sections = cur.fetchall()

    finally:
        cur.close()
        conn.close()

    return render_template(
        "edit_student.html",
        student=student,
        programs=programs,
        sections=sections  # pass sections to template
    )

# ==========================
# UPDATE STUDENT
# ==========================
@auth.route('/update_student/<int:user_id>', methods=['POST'])
def update_student(user_id):
    if 'user_id' not in session or session.get('role_id') != 1:
        flash("Access denied", "danger")
        return redirect(url_for('auth.login'))

    full_name = request.form.get("full_name")
    email = request.form.get("email")
    student_id = request.form.get("student_id")
    program_id = request.form.get("program_id") or None  # convert empty string to None
    section_id = request.form.get("section_id") or None  # convert empty string to None
    status = request.form.get("status")

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            UPDATE users
            SET 
                full_name = %s,
                email = %s,
                student_id = %s,
                program_id = %s,
                section_id = %s,
                status = %s
            WHERE user_id = %s AND role_id = 3
            RETURNING user_id;
        """, (full_name, email, student_id, program_id, section_id, status, user_id))

        updated = cur.fetchone()
        conn.commit()

        if updated:
            flash("Student record updated successfully!", "success")
        else:
            flash("Student not found.", "danger")

    except Exception as e:
        conn.rollback()
        flash(f"Error updating student: {str(e)}", "danger")

    finally:
        cur.close()
        conn.close()

    # ✅ Use correct blueprint prefix
    return redirect(url_for("auth.student_list"))


from flask import request, jsonify, session, current_app
# Assuming 'auth' is your Blueprint

# ========================================================
# SINGLE PROMOTE STUDENT
# ========================================================

@auth.route('/promote_student/<int:user_id>', methods=['POST'])
def promote_student(user_id):
    # 1. Check Authorization
    if 'user_id' not in session or session.get('role_id') != 1:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    conn = current_app.get_db_connection()
    cur = conn.cursor()

    try:
        # 2. SQL Query
        # Updates the user to the lowest available year_level that is greater than their current one.
        # The 'AND EXISTS' clause prevents setting the ID to NULL if they are already at the max level.
        query = """
            UPDATE users u
            SET year_level_id = (
                SELECT MIN(nl.year_level_id)
                FROM year_levels nl
                WHERE nl.year_level_id > u.year_level_id
            )
            WHERE u.user_id = %s
            AND EXISTS (
                SELECT 1
                FROM year_levels nl
                WHERE nl.year_level_id > u.year_level_id
            )
            RETURNING u.year_level_id;
        """
        
        cur.execute(query, (user_id,))
        updated_row = cur.fetchone()
        conn.commit()

        if updated_row:
            return jsonify({'success': True, 'message': 'Student promoted successfully.'})
        else:
            return jsonify({'success': False, 'message': 'Student is already at the max year level or could not be found.'})

    except Exception as e:
        conn.rollback()
        print(f"Error promoting student: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        cur.close()
        conn.close()


# ========================================================
# BULK PROMOTE STUDENTS
# ========================================================

@auth.route('/bulk_promote_students', methods=['POST'])
def bulk_promote_students():
    # 1. Check Authorization
    if 'user_id' not in session or session.get('role_id') != 1:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    data = request.get_json()
    raw_ids = data.get('user_ids', [])

    if not raw_ids:
        return jsonify({'success': False, 'message': 'No students selected'})

    # 2. Validate IDs
    try:
        # Create a list of integers to pass to PostgreSQL array
        user_ids = [int(uid) for uid in raw_ids]
    except ValueError:
        return jsonify({'success': False, 'message': 'Invalid ID format provided.'})

    conn = current_app.get_db_connection()
    cur = conn.cursor()

    try:
        # 3. SQL Query
        # We use 'ANY(%s)' to handle the list of IDs.
        # We use the same subquery logic as single promote to ensure consistency.
        query = """
            UPDATE users u
            SET year_level_id = (
                SELECT MIN(nl.year_level_id)
                FROM year_levels nl
                WHERE nl.year_level_id > u.year_level_id
            )
            WHERE u.user_id = ANY(%s)
            AND EXISTS (
                SELECT 1
                FROM year_levels nl
                WHERE nl.year_level_id > u.year_level_id
            )
            RETURNING u.user_id;
        """
        
        # Pass the list as a single argument for ANY(%s)
        cur.execute(query, (user_ids,))
        updated_rows = cur.fetchall()
        conn.commit()
        
        count = len(updated_rows)

        if count > 0:
            return jsonify({'success': True, 'message': f'{count} students promoted successfully.'})
        else:
            return jsonify({'success': False, 'message': 'No students were promoted (they may already be at max level).'})

    except Exception as e:
        conn.rollback()
        print(f"Error bulk promoting: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        cur.close()
        conn.close()

# ==========================
# GET TEACHER LIST (role = 2) with program info
# ==========================
@auth.route('/teacher_list', methods=['GET'])
def teacher_list():
    if 'user_id' not in session or session.get('role_id') != 1:
        flash('Access denied', 'danger')
        return redirect(url_for('auth.login'))

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        # Fetch teachers with program info INCLUDING only approved teachers
        cur.execute("""
            SELECT 
                u.user_id, 
                u.full_name, 
                u.email, 
                u.status,
                u.date_created,
                p.program_id,
                p.program_name,
                p.program_code
            FROM users u
            LEFT JOIN programs p ON u.program_id = p.program_id
            WHERE u.role_id = 2
              AND (u.approved = TRUE)   
            ORDER BY u.full_name;
        """)
        teachers = cur.fetchall()

        # Fetch all programs for the dropdown filter
        cur.execute("""
            SELECT program_id, program_name
            FROM programs
            ORDER BY program_name;
        """)
        programs = cur.fetchall()

        # Dashboard stats
        total_active = sum(1 for t in teachers if t['status'] == 'active')
        inactive_teachers = sum(1 for t in teachers if t['status'] == 'inactive')
        pending_verifications = sum(1 for t in teachers if t['status'] == 'pending')

        from datetime import datetime, timedelta
        one_week_ago = datetime.now() - timedelta(days=7)
        new_enrollments = sum(1 for t in teachers if t['date_created'] >= one_week_ago)

    finally:
        cur.close()
        conn.close()

    return render_template(
        'teacher_list.html',
        teachers=teachers,
        programs=programs,
        total_active=total_active,
        inactive_teachers=inactive_teachers,
        pending_verifications=pending_verifications,
        new_enrollments=new_enrollments
    )

from openpyxl import Workbook
from flask import send_file
import io

@auth.route('/teacher_list_export', methods=['GET'])
def teacher_list_export():
    if 'user_id' not in session or session.get('role_id') != 1:
        flash('Access denied', 'danger')
        return redirect(url_for('auth.login'))

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        # Fetch all approved teachers
        cur.execute("""
            SELECT 
                u.user_id, 
                u.full_name, 
                u.email, 
                u.status,
                u.date_created,
                p.program_name,
                p.program_code
            FROM users u
            LEFT JOIN programs p ON u.program_id = p.program_id
            WHERE u.role_id = 2
              AND (u.approved = TRUE)
            ORDER BY u.full_name;
        """)
        teachers = cur.fetchall()

    finally:
        cur.close()
        conn.close()

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Teacher List"

    # Header row
    headers = [
        "User ID", "Full Name", "Email", "Status",
        "Program", "Program Code", "Date Created"
    ]
    ws.append(headers)

    # Data rows
    for t in teachers:
        ws.append([
            t['user_id'],
            t['full_name'],
            t['email'],
            t['status'],
            t['program_name'],
            t['program_code'],
            t['date_created'].strftime("%Y-%m-%d %H:%M:%S")
        ])

    # Save to memory
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name="teacher_list.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ==========================
# ACTIVATE TEACHER
# ==========================
@auth.route('/activate_teacher/<int:user_id>', methods=['POST'])
def activate_teacher(user_id):
    if 'user_id' not in session or session.get('role_id') != 1:
        return jsonify({"success": False, "message": "Access denied"}), 403

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        cur.execute("""
            UPDATE users
            SET status = 'active'
            WHERE user_id = %s AND role_id = 2
            RETURNING status;
        """, (user_id,))
        updated = cur.fetchone()
        conn.commit()
        if updated:
            return jsonify({
                "success": True,
                "message": "✅ Teacher activated successfully!",
                "status": updated['status']
            })
        return jsonify({"success": False, "message": "Teacher not found."}), 404
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "message": f"Error: {str(e)}"}), 500
    finally:
        cur.close()
        conn.close()

# ==========================
# DEACTIVATE TEACHER
# ==========================
@auth.route('/deactivate_teacher/<int:user_id>', methods=['POST'])
def deactivate_teacher(user_id):
    if 'user_id' not in session or session.get('role_id') != 1:
        return jsonify({"success": False, "message": "Access denied"}), 403

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        cur.execute("""
            UPDATE users
            SET status = 'inactive'
            WHERE user_id = %s AND role_id = 2
            RETURNING status;
        """, (user_id,))
        updated = cur.fetchone()
        conn.commit()
        if updated:
            return jsonify({
                "success": True,
                "message": "✅ Teacher deactivated successfully!",
                "status": updated['status']
            })
        return jsonify({"success": False, "message": "Teacher not found."}), 404
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "message": f"Error: {str(e)}"}), 500
    finally:
        cur.close()
        conn.close()

@auth.route('/admin-programs')
def adminprograms():
    # ✅ Restrict access to admin users only
    if 'user_id' not in session or session.get('role_id') != 1:
        flash('Please log in as an admin to access this page.', 'warning')
        return redirect(url_for('auth.adminlogin'))

    from . import create_app
    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor()

    user_id = session['user_id']

    # 🧩 Fetch admin info
    cur.execute("""
        SELECT user_id, full_name, email 
        FROM users 
        WHERE user_id = %s
    """, (user_id,))
    user = cur.fetchone()

    if not user:
        flash('User not found.', 'danger')
        cur.close()
        conn.close()
        return redirect(url_for('auth.adminlogin'))

        # 🧩 Fetch all programs (exclude deleted)
    cur.execute("""
        SELECT program_id, program_code, program_name, description
        FROM programs
        WHERE deleted = FALSE
        ORDER BY program_name ASC
    """)
    programs = cur.fetchall()


    total_programs = len(programs)

    cur.close()
    conn.close()

    return render_template(
        'admin_programs.html',
        user={
            'user_id': user[0],
            'full_name': user[1],
            'email': user[2]
        },
        programs=programs,
        stats={
            'total_programs': total_programs
        }
    )

@auth.route('/create-program', methods=['POST'])
def create_program():
    if 'user_id' not in session or session.get('role_id') != 1:
        flash('Access denied.', 'danger')
        return redirect(url_for('auth.adminlogin'))

    program_code = request.form.get('program_code')
    program_name = request.form.get('program_name')
    description = request.form.get('description')

    from . import create_app
    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO programs (program_code, program_name, description)
        VALUES (%s, %s, %s)
    """, (program_code, program_name, description))

    conn.commit()
    cur.close()
    conn.close()

    flash(f'Program "{program_name}" created successfully!', 'success')
    return redirect(url_for('auth.adminprograms'))

# Edit Program
@auth.route('/edit-program/<int:program_id>', methods=['POST'])
def edit_program(program_id):
    if 'user_id' not in session or session.get('role_id') != 1:
        flash('Access denied.', 'danger')
        return redirect(url_for('auth.adminlogin'))

    program_code = request.form.get('program_code')
    program_name = request.form.get('program_name')
    description = request.form.get('description')

    from . import create_app
    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE programs
        SET program_code = %s, program_name = %s, description = %s
        WHERE program_id = %s
    """, (program_code, program_name, description, program_id))

    conn.commit()
    cur.close()
    conn.close()

    flash(f'Program "{program_name}" updated successfully!', 'success')
    return redirect(url_for('auth.adminprograms'))

# Soft Delete Program
@auth.route('/delete-program/<int:program_id>', methods=['POST'])
def delete_program(program_id):
    if 'user_id' not in session or session.get('role_id') != 1:
        flash('Access denied.', 'danger')
        return redirect(url_for('auth.adminlogin'))

    from . import create_app
    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor()

    # Mark program as deleted
    cur.execute("""
        UPDATE programs
        SET deleted = TRUE
        WHERE program_id = %s
    """, (program_id,))

    conn.commit()
    cur.close()
    conn.close()

    flash('Program deleted successfully!', 'success')
    return redirect(url_for('auth.adminprograms'))

# ========================
# SECTION MANAGEMENT ROUTES
# ========================
@auth.route('/admin-sections')
def adminsections():
    # ✅ Restrict access to admin users only
    if 'user_id' not in session or session.get('role_id') != 1:
        flash('Please log in as an admin to access this page.', 'warning')
        return redirect(url_for('auth.adminlogin'))

    from . import create_app
    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor()

    user_id = session['user_id']

    # 🧩 Fetch admin info
    cur.execute("""
        SELECT user_id, full_name, email 
        FROM users 
        WHERE user_id = %s
    """, (user_id,))
    user = cur.fetchone()

    if not user:
        flash('User not found.', 'danger')
        cur.close()
        conn.close()
        return redirect(url_for('auth.adminlogin'))

    # 🧩 Fetch all sections (including year_name)
    cur.execute("""
        SELECT 
            s.section_id,
            s.section_name,
            yl.year_name AS year_name,
            s.semester,
            p.program_name,
            u.full_name AS adviser_name
        FROM sections s
        JOIN programs p ON s.program_id = p.program_id
        LEFT JOIN users u ON s.adviser_id = u.user_id
        LEFT JOIN year_levels yl ON s.year_level_id = yl.year_level_id
        WHERE s.deleted = FALSE
        ORDER BY p.program_name ASC, yl.year_name ASC, s.section_name ASC
    """)
    sections = cur.fetchall()

    # 🧩 Fetch all programs (for dropdown)
    cur.execute("""
        SELECT program_id, program_name 
        FROM programs 
        WHERE deleted = FALSE
        ORDER BY program_name ASC
    """)
    programs = cur.fetchall()

    # 🧩 Fetch all teachers (for adviser dropdown)
    cur.execute("""
        SELECT user_id, full_name
        FROM users
        WHERE role_id = 2 -- Teacher
        ORDER BY full_name ASC
    """)
    advisers = cur.fetchall()

    # 🧩 Fetch all year levels (for dropdown)
    cur.execute("""
        SELECT year_level_id, year_name
        FROM year_levels
        ORDER BY year_name ASC
    """)
    year_levels = cur.fetchall()

    total_sections = len(sections)

    cur.close()
    conn.close()

    return render_template(
        'admin_sections.html',
        user={
            'user_id': user[0],
            'full_name': user[1],
            'email': user[2]
        },
        sections=sections,
        programs=programs,
        advisers=advisers,
        year_levels=year_levels,
        stats={
            'total_sections': total_sections
        }
    )

# =====================
# CREATE SECTION
# =====================
@auth.route('/create-section', methods=['POST'])
def create_section():
    if 'user_id' not in session or session.get('role_id') != 1:
        flash('Access denied.', 'danger')
        return redirect(url_for('auth.adminlogin'))

    program_id = request.form.get('program_id')
    section_name = request.form.get('section_name')
    year_level_id = request.form.get('year_level_id')  # changed
    semester = request.form.get('semester')
    adviser_id = request.form.get('adviser_id') or None

    from . import create_app
    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO sections (program_id, section_name, year_level_id, semester, adviser_id)
        VALUES (%s, %s, %s, %s, %s)
    """, (program_id, section_name, year_level_id, semester, adviser_id))

    conn.commit()
    cur.close()
    conn.close()

    flash(f'Section "{section_name}" created successfully!', 'success')
    return redirect(url_for('auth.adminsections'))

# =====================
# EDIT SECTION
# =====================
@auth.route('/edit-section/<int:section_id>', methods=['POST'])
def edit_section(section_id):
    if 'user_id' not in session or session.get('role_id') != 1:
        flash('Access denied.', 'danger')
        return redirect(url_for('auth.adminlogin'))

    program_id = request.form.get('program_id')
    section_name = request.form.get('section_name')
    year_level_id = request.form.get('year_level_id')  # changed
    semester = request.form.get('semester')
    adviser_id = request.form.get('adviser_id') or None

    from . import create_app
    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE sections
        SET program_id = %s, section_name = %s, year_level_id = %s, semester = %s, adviser_id = %s
        WHERE section_id = %s
    """, (program_id, section_name, year_level_id, semester, adviser_id, section_id))

    conn.commit()
    cur.close()
    conn.close()

    flash(f'Section "{section_name}" updated successfully!', 'success')
    return redirect(url_for('auth.adminsections'))

# =====================
# SOFT DELETE SECTION
# =====================
@auth.route('/delete-section/<int:section_id>', methods=['POST'])
def delete_section(section_id):
    if 'user_id' not in session or session.get('role_id') != 1:
        flash('Access denied.', 'danger')
        return redirect(url_for('auth.adminlogin'))

    from . import create_app
    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor()

    # Soft delete the section
    cur.execute("""
        UPDATE sections
        SET deleted = TRUE
        WHERE section_id = %s
    """, (section_id,))

    conn.commit()
    cur.close()
    conn.close()

    flash('Section deleted successfully!', 'success')
    return redirect(url_for('auth.adminsections'))

# ========================
# ACADEMIC YEAR MANAGEMENT ROUTES
# ========================

@auth.route('/admin-academicyear')
def admin_academicyear():
    # ✅ Restrict access to admin users only
    if 'user_id' not in session or session.get('role_id') != 1:
        flash('Please log in as an admin to access this page.', 'warning')
        return redirect(url_for('auth.adminlogin'))

    from . import create_app
    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor()

    # 🧩 Fetch admin info
    user_id = session['user_id']
    cur.execute("SELECT user_id, full_name, email FROM users WHERE user_id = %s", (user_id,))
    user = cur.fetchone()

    # 🧩 Fetch all academic years
    cur.execute("""
        SELECT academic_year_id, start_year, end_year, semester, status, date_created
        FROM academic_years
        WHERE status != 'archived'
        ORDER BY start_year DESC, semester
    """)
    academic_years = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        'admin_academic_years.html',
        user={
            'user_id': user[0],
            'full_name': user[1],
            'email': user[2]
        },
        academic_years=academic_years
    )

# =====================
# CREATE ACADEMIC YEAR
# =====================
@auth.route('/create-academic-year', methods=['POST'])
def create_academic_year():
    if 'user_id' not in session or session.get('role_id') != 1:
        flash('Access denied.', 'danger')
        return redirect(url_for('auth.adminlogin'))

    start_year = request.form.get('start_year')
    end_year = request.form.get('end_year')
    semester = request.form.get('semester')
    status = request.form.get('status', 'inactive')  # Default to inactive

    from . import create_app
    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor()

    # Ensure no duplicates
    cur.execute("""
        SELECT COUNT(*) FROM academic_years
        WHERE start_year = %s AND end_year = %s AND semester = %s
    """, (start_year, end_year, semester))
    if cur.fetchone()[0] > 0:
        flash('This academic year already exists.', 'warning')
        cur.close()
        conn.close()
        return redirect(url_for('auth.admin_academicyear'))

    # Insert new record
    cur.execute("""
        INSERT INTO academic_years (start_year, end_year, semester, status)
        VALUES (%s, %s, %s, %s)
    """, (start_year, end_year, semester, status))

    conn.commit()
    cur.close()
    conn.close()

    flash('Academic year created successfully!', 'success')
    return redirect(url_for('auth.admin_academicyear'))

# =====================
# EDIT ACADEMIC YEAR
# =====================
@auth.route('/edit-academic-year/<int:academic_year_id>', methods=['POST'])
def edit_academic_year(academic_year_id):
    if 'user_id' not in session or session.get('role_id') != 1:
        flash('Access denied.', 'danger')
        return redirect(url_for('auth.adminlogin'))

    start_year = request.form.get('start_year')
    end_year = request.form.get('end_year')
    semester = request.form.get('semester')
    status = request.form.get('status')

    from . import create_app
    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE academic_years
        SET start_year = %s, end_year = %s, semester = %s, status = %s
        WHERE academic_year_id = %s
    """, (start_year, end_year, semester, status, academic_year_id))

    conn.commit()
    cur.close()
    conn.close()

    flash('Academic year updated successfully!', 'success')
    return redirect(url_for('auth.admin_academicyear'))

# =====================
# ACTIVATE / DEACTIVATE
# =====================
@auth.route('/toggle-academic-year/<int:academic_year_id>/<string:action>', methods=['POST'])
def toggle_academic_year(academic_year_id, action):
    if 'user_id' not in session or session.get('role_id') != 1:
        flash('Access denied.', 'danger')
        return redirect(url_for('auth.adminlogin'))

    new_status = 'active' if action == 'activate' else 'inactive'

    from . import create_app
    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor()

    # Optional: Deactivate all others before activating new one
    if new_status == 'active':
        cur.execute("UPDATE academic_years SET status = 'inactive' WHERE status = 'active'")

    cur.execute("""
        UPDATE academic_years
        SET status = %s
        WHERE academic_year_id = %s
    """, (new_status, academic_year_id))

    conn.commit()
    cur.close()
    conn.close()

    flash(f'Academic year has been {new_status}.', 'success')
    return redirect(url_for('auth.admin_academicyear'))


# =====================
# DELETE ACADEMIC YEAR (ARCHIVE)
# =====================
@auth.route('/delete-academic-year/<int:academic_year_id>', methods=['POST'])
def delete_academic_year(academic_year_id):
    if 'user_id' not in session or session.get('role_id') != 1:
        flash('Access denied.', 'danger')
        return redirect(url_for('auth.adminlogin'))

    from . import create_app
    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE academic_years
        SET status = 'archived'
        WHERE academic_year_id = %s
    """, (academic_year_id,))

    conn.commit()
    cur.close()
    conn.close()

    flash('Academic year archived successfully.', 'success')
    return redirect(url_for('auth.admin_academicyear'))


from datetime import date, datetime, timedelta

@auth.route('/dashboard', methods=['GET', 'POST'])
def dashboard():
    # ===============================
    # AUTHENTICATION
    # ===============================
    if 'user_id' not in session:
        flash('Please log in to access the dashboard.', 'warning')
        return redirect(url_for('auth.login'))

    allowed_roles = [1, 2, 3]
    if session.get('role_id') not in allowed_roles:
        flash('Access denied. You do not have permission.', 'danger')
        return redirect(url_for('auth.login'))

    # ===============================
    # DATABASE CONNECTION
    # ===============================
    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        # FETCH USER INFO
        cur.execute("""
            SELECT u.user_id, u.full_name, u.student_id, u.email, u.role_id,
                   u.program_id, u.section_id,
                   p.program_code, s.section_name
            FROM users u
            LEFT JOIN programs p ON u.program_id = p.program_id
            LEFT JOIN sections s ON u.section_id = s.section_id
            WHERE u.user_id = %s
        """, (session['user_id'],))
        user = cur.fetchone()
        if not user:
            flash('User not found.', 'danger')
            return redirect(url_for('auth.login'))

        # FETCH TODAY'S MEETINGS FOR USER'S PROGRAM & SECTION
        cur.execute("""
            SELECT 
                m.meeting_id, m.title, m.description, m.room_code,
                m.scheduled_date, m.scheduled_time,
                u.full_name AS created_by_name,
                c.course_code, c.course_title,
                p.program_name, s.section_name
            FROM meetings m
            LEFT JOIN users u ON m.created_by = u.user_id
            LEFT JOIN courses c ON m.course_id = c.course_id
            LEFT JOIN programs p ON m.program_id = p.program_id
            LEFT JOIN sections s ON m.section_id = s.section_id
            WHERE 
                m.scheduled_date = %s
                AND m.program_id = %s
                AND m.section_id = %s
            ORDER BY m.scheduled_time ASC
        """, (date.today(), user['program_id'], user['section_id']))
        meetings_today = cur.fetchall()

        # ===============================
        # FILTER MEETINGS BASED ON TIME
        # Remove meetings where current time > scheduled_time + 60 minutes
        # ===============================
        now = datetime.now()
        filtered_meetings = []
        for meeting in meetings_today:
            scheduled_datetime = datetime.combine(meeting['scheduled_date'], meeting['scheduled_time'])
            if now <= scheduled_datetime + timedelta(minutes=60):
                filtered_meetings.append(meeting)

        meetings_today = filtered_meetings

        # FETCH UNREAD NOTIFICATIONS (optional)
        cur.execute("""
            SELECT n.notification_id, n.course_id, n.title, n.message, n.read, n.created_at,
                   c.course_code
            FROM notifications n
            LEFT JOIN courses c ON n.course_id = c.course_id
            WHERE n.user_id = %s AND n.read = FALSE
            ORDER BY n.created_at DESC
        """, (session['user_id'],))
        notifications = cur.fetchall()
        unread_count = len(notifications)

    finally:
        cur.close()
        conn.close()

    # HANDLE JOIN MEETING
    if request.method == "POST":
        room_id = request.form.get("roomID")
        if room_id:
            return redirect(f"/meeting?roomID={room_id}")

    # RENDER DASHBOARD
    return render_template(
        'dashboard.html',
        user={
            'user_id': user['user_id'],
            'full_name': user['full_name'],
            'student_id': user.get('student_id', ''),
            'email': user['email'],
            'role_id': user['role_id'],
            'program_code': user.get('program_code', ''),
            'section': user.get('section_name', '')
        },
        notifications=notifications,
        unread_count=unread_count,
        meetings_today=meetings_today
    )

# ===============================
# CLEAR ALL NOTIFICATIONS (Soft Delete)
# ===============================
@auth.route('/notifications/clear_all', methods=['POST'])
def clear_all_notifications():
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'User not logged in.'}), 401

    user_id = session['user_id']

    try:
        app = create_app()
        conn = app.get_db_connection()
        cur = conn.cursor()

        # Mark all notifications as read (soft delete)
        cur.execute("""
            UPDATE notifications
            SET read = TRUE
            WHERE user_id = %s AND read = FALSE;
        """, (user_id,))

        conn.commit()
        cur.close()
        conn.close()

        return jsonify({'status': 'success', 'message': 'All notifications cleared.'})

    except Exception as e:
        print("Error clearing notifications:", e)
        return jsonify({'status': 'error', 'message': str(e)}), 500



import random
from collections import defaultdict
from datetime import datetime, timezone, timedelta
from flask import Blueprint, render_template, request, redirect, url_for, session, flash, jsonify
import psycopg2.extras

# Define Philippine Timezone (UTC+8)
PH_TZ = timezone(timedelta(hours=8))

from flask import current_app # Ensure this is imported

@auth.route('/create_meeting/<int:course_id>', methods=['GET', 'POST'])
def create_meeting(course_id):
    # 1. Check Login
    if 'user_id' not in session:
        flash('Please log in to schedule a meeting.', 'warning')
        return redirect(url_for('auth.login'))

    if session.get('role_id') not in [1, 2]:
        flash('Access denied.', 'danger')
        return redirect(url_for('auth.dashboard'))

    # FIX 1: Use 'current_app' instead of creating a new app instance
    conn = current_app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        # 2. Get User & Course Info
        cur.execute("SELECT user_id, full_name, role_id, email FROM users WHERE user_id = %s", (session['user_id'],))
        current_user = cur.fetchone()

        cur.execute("""
            SELECT c.course_id, c.course_code, c.course_title,
                   c.program_id, c.section_id,
                   p.program_name, s.section_name
            FROM courses c
            LEFT JOIN programs p ON c.program_id = p.program_id
            LEFT JOIN sections s ON c.section_id = s.section_id
            WHERE c.course_id = %s
        """, (course_id,))
        course = cur.fetchone()

        if not course:
            flash('Course not found.', 'danger')
            return redirect(url_for('auth.dashboard'))

        # Default random code
        room_code = str(random.randint(100000, 999999))

        # 3. Handle Form Submission
        if request.method == 'POST':
            action = request.form.get('action')
            meeting_id = request.form.get('meeting_id')
            is_instant = request.form.get('is_instant') == 'true'

            title = request.form.get('title')
            description = request.form.get('description')
            room_code_input = request.form.get('room_code') or room_code
            date_input = request.form.get('date')
            time_input = request.form.get('time')

            if not all([title, room_code_input, date_input, time_input]):
                flash('Please fill out all required fields.', 'warning')
            else:
                # Parse Date/Time
                try:
                    dt_str = f"{date_input} {time_input}"
                    naive_dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M")
                    ph_dt = naive_dt.replace(tzinfo=PH_TZ)
                except ValueError:
                    ph_dt = datetime.now(PH_TZ)

                if action == 'edit' and meeting_id:
                    # Update Existing
                    cur.execute("""
                        UPDATE meetings
                        SET title=%s, description=%s, room_code=%s,
                            scheduled_date=%s, scheduled_time=%s
                        WHERE meeting_id=%s AND course_id=%s
                    """, (
                        title, description, room_code_input,
                        ph_dt.date(), ph_dt.time(),
                        meeting_id, course_id
                    ))
                    conn.commit()
                    flash('Meeting updated successfully!', 'success')

                else:
                    # Create New
                    # FIX 2: Explicitly insert 'FALSE' for deleted column
                    cur.execute("""
                        INSERT INTO meetings (
                            title, description, room_code, scheduled_date, scheduled_time,
                            created_by, program_id, section_id, course_id, deleted
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, FALSE)
                        RETURNING meeting_id;
                    """, (
                        title, description, room_code_input,
                        ph_dt.date(), ph_dt.time(),
                        session['user_id'],
                        course['program_id'],
                        course['section_id'],
                        course_id
                    ))
                    # Consume the ID to clear cursor
                    new_id = cur.fetchone()
                    conn.commit()

                    # Notify Students (if not instant)
                    if not is_instant:
                        cur.execute("""
                            SELECT u.user_id 
                            FROM enrollments e
                            JOIN users u ON e.user_id = u.user_id
                            WHERE e.course_id = %s AND u.role_id = 3
                        """, (course_id,))
                        students = cur.fetchall()
                        
                        teacher_name = current_user['full_name']
                        for s in students:
                            cur.execute("""
                                INSERT INTO notifications (user_id, course_id, title, message, created_at, read)
                                VALUES (%s, %s, %s, %s, NOW(), FALSE)
                            """, (s['user_id'], course_id, "New Meeting", f"{teacher_name} scheduled '{title}'."))
                        conn.commit()
                        flash('Meeting scheduled!', 'success')

                    if is_instant:
                        return redirect(url_for('auth.join_meeting', room_code=room_code_input))

            return redirect(url_for('auth.create_meeting', course_id=course_id))

        # 4. Handle Delete
        delete_id = request.args.get('delete_id')
        if delete_id:
            cur.execute("UPDATE meetings SET deleted=TRUE WHERE meeting_id=%s", (delete_id,))
            conn.commit()
            flash('Meeting deleted.', 'success')
            return redirect(url_for('auth.create_meeting', course_id=course_id))

        # 5. Fetch Data for Display
        # FIX 3: Robust WHERE clause to handle NULLs and FALSE
        cur.execute("""
            SELECT m.meeting_id, m.title, m.description, m.room_code,
                   m.scheduled_date, m.scheduled_time,
                   u.full_name AS created_by_name
            FROM meetings m
            LEFT JOIN users u ON m.created_by = u.user_id
            WHERE (m.deleted IS FALSE OR m.deleted IS NULL) 
            AND m.course_id = %s
            ORDER BY m.scheduled_date DESC, m.scheduled_time DESC
        """, (course_id,))
        meetings = cur.fetchall()

        # Get Attendance
        cur.execute("""
            SELECT ma.meeting_id, u.full_name, ma.joined_at, ma.left_at
            FROM meeting_attendance ma
            JOIN users u ON ma.user_id = u.user_id
            JOIN meetings m ON ma.meeting_id = m.meeting_id
            WHERE m.course_id = %s
            ORDER BY ma.joined_at ASC
        """, (course_id,))
        raw_attendance = cur.fetchall()

        # Process Attendance
        attendance_map = defaultdict(list)
        for record in raw_attendance:
            mid = record['meeting_id']
            
            # Safe timezone conversion
            joined_iso = "-"
            if record['joined_at']:
                dt = record['joined_at']
                if dt.tzinfo is None: dt = dt.replace(tzinfo=timezone.utc)
                joined_iso = dt.astimezone(PH_TZ).isoformat()

            left_iso = "-"
            if record['left_at']:
                dt = record['left_at']
                if dt.tzinfo is None: dt = dt.replace(tzinfo=timezone.utc)
                left_iso = dt.astimezone(PH_TZ).isoformat()

            attendance_map[mid].append({
                'name': record['full_name'],
                'joined': joined_iso,
                'left': left_iso
            })

        attendance_dict = dict(attendance_map)

    except Exception as e:
        print(f"Error in create_meeting: {e}")
        # flash(f"Error: {e}", "danger") # Uncomment this to see errors on screen
        meetings = []
        attendance_dict = {}
        room_code = ""
    finally:
        cur.close()
        conn.close()

    return render_template(
        'create_meeting.html',
        course=course,
        meetings=meetings,
        room_code=room_code,
        user=current_user,
        attendance=attendance_dict,
        programs=[], sections=[], courses=[]
    )

# Add this to auth.py

from datetime import datetime, timedelta, date, time
import pytz # Ensure you have pytz installed

@auth.route('/view_meetings/<int:course_id>', methods=['GET'])
def view_meetings(course_id):
    # 1. Check Login
    if 'user_id' not in session:
        flash('Please log in to view meetings.', 'warning')
        return redirect(url_for('auth.login'))

    conn = current_app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        # 2. Get User & Course Info
        cur.execute("SELECT user_id, full_name, role_id, email FROM users WHERE user_id = %s", (session['user_id'],))
        current_user = cur.fetchone()

        cur.execute("""
            SELECT c.course_id, c.course_code, c.course_title, s.section_name
            FROM courses c
            LEFT JOIN sections s ON c.section_id = s.section_id
            WHERE c.course_id = %s
        """, (course_id,))
        course = cur.fetchone()

        if not course:
            flash('Course not found.', 'danger')
            if session.get('role_id') == 3:
                return redirect(url_for('auth.student_dashboard'))
            return redirect(url_for('auth.teacher_dashboard'))

        # 3. Fetch All Meetings (Non-deleted)
        cur.execute("""
            SELECT m.meeting_id, m.title, m.description, m.room_code,
                   m.scheduled_date, m.scheduled_time,
                   u.full_name AS created_by_name
            FROM meetings m
            LEFT JOIN users u ON m.created_by = u.user_id
            WHERE (m.deleted IS FALSE OR m.deleted IS NULL) 
            AND m.course_id = %s
            ORDER BY m.scheduled_date ASC, m.scheduled_time ASC
        """, (course_id,))
        all_meetings = cur.fetchall()

        # 4. Filter: Active vs Archived Logic
        active_meetings = []
        archived_meetings = []
        
        # Define Timezone (Philippines)
        ph_tz = pytz.timezone('Asia/Manila')
        now = datetime.now(ph_tz)

        for meeting in all_meetings:
            # Combine Date and Time from DB into a full datetime object
            # Note: psycopg2 usually returns datetime.date and datetime.time objects
            sch_date = meeting['scheduled_date']
            sch_time = meeting['scheduled_time']
            
            if sch_date and sch_time:
                # Create a naive datetime
                dt_naive = datetime.combine(sch_date, sch_time)
                # Localize it to PH time
                meeting_start = ph_tz.localize(dt_naive)
                
                # Calculate "Expiration" (Start Time + 60 Minutes)
                meeting_end_buffer = meeting_start + timedelta(minutes=60)

                if now > meeting_end_buffer:
                    # It has been more than 60 mins since start -> Archive
                    archived_meetings.append(meeting)
                else:
                    # Upcoming or currently happening -> Active
                    active_meetings.append(meeting)

        # Sort Active: Nearest upcoming first
        # Sort Archived: Most recent past first
        active_meetings.sort(key=lambda x: (x['scheduled_date'], x['scheduled_time']))
        archived_meetings.sort(key=lambda x: (x['scheduled_date'], x['scheduled_time']), reverse=True)

    except Exception as e:
        print(f"Error in view_meetings: {e}")
        active_meetings = []
        archived_meetings = []
        course = None
    finally:
        cur.close()
        conn.close()

    return render_template(
        'view_meetings.html',
        course=course,
        meetings=active_meetings,   # Only send active ones to the main loop
        archived=archived_meetings, # Send these if you want to display a "Past Meetings" section
        user=current_user
    )

from flask import Blueprint, session, render_template, flash, redirect, url_for, request, current_app
# ==========================
# JOIN MEETING
# ==========================
@auth.route('/join_meeting/<room_code>')
def join_meeting(room_code):
    if 'user_id' not in session:
        flash('Please log in to join the meeting.', 'warning')
        return redirect(url_for('auth.login'))

    user_id = session['user_id']
    conn = current_app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        # 1. Fetch meeting info
        cur.execute("""
            SELECT m.meeting_id, m.title AS meeting_title,
                   c.course_id, c.course_code, c.course_title
            FROM meetings m
            JOIN courses c ON m.course_id = c.course_id
            WHERE m.room_code=%s AND m.deleted=FALSE
        """, (room_code,))
        meeting = cur.fetchone()

        if not meeting:
            flash('Meeting not found.', 'danger')
            return redirect(url_for('auth.dashboard'))
        
        # 2. Fetch User Role (Teacher=2, Student=3)
        cur.execute("SELECT role_id FROM users WHERE user_id = %s", (user_id,))
        user_data = cur.fetchone()
        # Default to 3 (Student) if something goes wrong
        user_role = user_data['role_id'] if user_data else 3

        meeting_id = meeting['meeting_id']
        now = datetime.now(PH_TZ)

        # 3. Track attendance (joined_at)
        cur.execute("""
            SELECT * FROM meeting_attendance 
            WHERE meeting_id=%s AND user_id=%s
        """, (meeting_id, user_id))
        attendance = cur.fetchone()

        if not attendance:
            # Insert new record with joined_at, leave left_at as NULL
            cur.execute("""
                INSERT INTO meeting_attendance (meeting_id, user_id, joined_at)
                VALUES (%s, %s, %s)
            """, (meeting_id, user_id, now))
        else:
            # Re-joining: Update joined_at and RESET left_at to NULL (active again)
            cur.execute("""
                UPDATE meeting_attendance
                SET joined_at=%s, left_at=NULL
                WHERE meeting_id=%s AND user_id=%s
            """, (now, meeting_id, user_id))

        conn.commit()

    finally:
        cur.close()
        conn.close()

    # 4. Pass 'role' to the HTML template
    return render_template(
        'join_meeting.html',
        room_code=room_code,
        user_name=session.get('full_name', 'Guest'),
        role=user_role,  # <--- PASSING THE ROLE HERE
        course={
            'course_id': meeting['course_id'],
            'course_code': meeting['course_code'],
            'course_title': meeting['course_title']
        }
    )

# ==========================
# LEAVE MEETING (No changes needed here, but kept for context)
# ==========================
@auth.route('/leave_meeting/<room_code>', methods=['POST'])
def leave_meeting(room_code):
    if 'user_id' not in session:
        return {"status": "error", "message": "Unauthorized"}, 401

    user_id = session['user_id']
    conn = current_app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        cur.execute("SELECT meeting_id FROM meetings WHERE room_code=%s AND deleted=FALSE", (room_code,))
        meeting = cur.fetchone()
        if not meeting:
            return {"status": "error", "message": "Meeting not found"}, 404

        meeting_id = meeting['meeting_id']
        now = datetime.now(PH_TZ)

        cur.execute("""
            UPDATE meeting_attendance
            SET left_at=%s
            WHERE meeting_id=%s AND user_id=%s
        """, (now, meeting_id, user_id))
        conn.commit()

    finally:
        cur.close()
        conn.close()

    return {"status": "success", "message": "Left meeting recorded"}

# =========================
# STUDENT LMS - SHOW ENROLLED COURSES WITH FILTERS
# =========================
@auth.route('/lms', methods=['GET', 'POST'])
def lms():
    if 'user_id' not in session:
        flash('Please log in first.', 'warning')
        return redirect(url_for('auth.login'))

    # Only allow students
    if session.get('role_id') != 3:
        flash('Access denied. Students only.', 'danger')
        return redirect(url_for('auth.login'))

    from . import create_app
    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    user_id = session['user_id']

    # -------------------------
    # Fetch student info
    # -------------------------
    cur.execute("""
        SELECT u.user_id, u.full_name, u.email, u.program_id, u.year_level_id, s.section_name
        FROM users u
        LEFT JOIN sections s ON u.section_id = s.section_id
        WHERE u.user_id = %s
    """, (user_id,))
    student = cur.fetchone()
    if not student:
        flash("Student record not found.", "danger")
        return redirect(url_for('auth.login'))

    # -------------------------
    # Fetch academic years for filter (only for enrolled courses)
    # -------------------------
    cur.execute("""
        SELECT DISTINCT a.academic_year_id, a.start_year, a.end_year, a.semester
        FROM courses c
        JOIN course_students cs ON c.course_id = cs.course_id
        JOIN academic_years a ON c.academic_year_id = a.academic_year_id
        WHERE cs.user_id = %s
          AND c.status = 'approved'
          AND a.status = 'active'
        ORDER BY a.start_year DESC, a.semester
    """, (user_id,))
    academic_years = cur.fetchall()

    # -------------------------
    # Fetch all year levels (optional, for display/filter)
    # -------------------------
    cur.execute("SELECT year_level_id, year_name FROM year_levels ORDER BY year_level_id;")
    year_levels = cur.fetchall()

    # -------------------------
    # Get filter values from form or query parameters
    # -------------------------
    selected_academic_year_id = request.form.get('academic_year_id') or request.args.get('academic_year_id')
    selected_semester = request.form.get('semester') or request.args.get('semester')

    # -------------------------
    # Fetch only courses where student is enrolled
    # -------------------------
    query = """
        SELECT c.course_id, c.course_code, c.course_title, c.course_description,
               c.units, c.year_level_id, yl.year_name AS year_level, c.status,
               p.program_name, s.section_name,
               u.full_name AS instructor_name,
               CONCAT(a.start_year, '-', a.end_year, ' (', a.semester, ' Semester)') AS academic_year,
               a.semester
        FROM courses c
        JOIN course_students cs ON c.course_id = cs.course_id
        JOIN programs p ON c.program_id = p.program_id
        JOIN sections s ON c.section_id = s.section_id
        JOIN academic_years a ON c.academic_year_id = a.academic_year_id
        JOIN year_levels yl ON c.year_level_id = yl.year_level_id
        LEFT JOIN users u ON c.user_id = u.user_id
        WHERE cs.user_id = %s
          AND c.status = 'approved'
          AND a.status = 'active'
    """
    params = [user_id]

    # Apply filters only on enrolled courses
    if selected_academic_year_id and selected_academic_year_id != "all":
        query += " AND c.academic_year_id = %s"
        params.append(selected_academic_year_id)
    if selected_semester and selected_semester != "all":
        query += " AND a.semester = %s"
        params.append(selected_semester)

    query += " ORDER BY c.year_level_id, c.course_code;"
    cur.execute(query, params)
    courses = cur.fetchall()

    # -------------------------
    # Close connection
    # -------------------------
    cur.close()
    conn.close()

    # -------------------------
    # Render LMS page
    # -------------------------
    return render_template(
        'lms.html',
        user=student,
        courses=courses,
        academic_years=academic_years,
        year_levels=year_levels,
        selected_academic_year_id=selected_academic_year_id,
        selected_semester=selected_semester
    )

from flask import Blueprint, session, flash, redirect, url_for, render_template, jsonify, current_app
from datetime import datetime, timedelta, date, time
import pytz
import psycopg2.extras

PH_TZ = pytz.timezone('Asia/Manila')

def ensure_aware_datetime(val):
    """ Convert dates/datetimes to PH timezone-aware datetime """
    if val is None:
        return None
    if isinstance(val, date) and not isinstance(val, datetime):
        val = datetime.combine(val, time.min)
    if isinstance(val, datetime):
        if val.tzinfo is None:
            return PH_TZ.localize(val)
        return val.astimezone(PH_TZ)
    return None


# ========================================================
# MATERIALS ROUTE (FULLY FIXED VERSION)
# ========================================================
@auth.route('/materials/<int:course_id>')
def materials(course_id):
    if 'user_id' not in session:
        flash('Please log in first.', 'warning')
        return redirect(url_for('auth.login'))

    user_id = session['user_id']
    role_id = session.get('role_id')

    conn = current_app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    now = datetime.now(PH_TZ)  # Current PH time

    try:
        # -----------------------------------------------------------
        # Fetch user
        # -----------------------------------------------------------
        cur.execute("SELECT user_id, full_name, email, role_id FROM users WHERE user_id=%s", (user_id,))
        user = cur.fetchone()

        # -----------------------------------------------------------
        # Fetch course
        # -----------------------------------------------------------
        cur.execute("""
            SELECT c.course_id, c.course_code, c.course_title, s.section_name
            FROM courses c
            LEFT JOIN sections s ON c.section_id = s.section_id
            WHERE c.course_id = %s
        """, (course_id,))
        course = cur.fetchone()

        if not course:
            flash('Course not found.', 'danger')
            return redirect(url_for('auth.lms'))

        # -----------------------------------------------------------
        # Fetch OBTLP (correct table)
        # -----------------------------------------------------------
        cur.execute("""
            SELECT * FROM obtlp_files
            WHERE course_id = %s AND (is_deleted IS DISTINCT FROM TRUE)
            ORDER BY upload_date DESC LIMIT 1
        """, (course_id,))
        obtlp = cur.fetchone()

        if obtlp and obtlp.get('file_path'):
            obtlp['static_path'] = url_for('static', filename=obtlp['file_path'])

        # -----------------------------------------------------------
        # Fetch Materials
        # -----------------------------------------------------------
        cur.execute("""
            SELECT material_id, title, description, file_path, material_type, upload_date, term
            FROM materials
            WHERE course_id=%s AND is_deleted=FALSE
            ORDER BY upload_date DESC
        """, (course_id,))
        materials = cur.fetchall()

        for m in materials:
            m['static_path'] = url_for('static', filename=m['file_path'])
            m['preview_type'] = 'pdf'

        # -----------------------------------------------------------
        # ROLE LOGIC
        # -----------------------------------------------------------
        quizzes, assignments, activities = [], [], []

        # ===========================================================
        # STUDENT VIEW
        # ===========================================================
        if role_id == 3:

            # ----------------------------------------------------------
            # QUIZZES (UNTOUCHED — as requested)
            # ----------------------------------------------------------
            cur.execute("""
                SELECT q.quiz_id, q.title, q.total_points, q.date_published, q.duration_minutes, q.material_id,
                       q.publish_at, q.max_attempts, q.published,
                       qe.extended_due_date, qe.extra_attempts,
                       (SELECT COUNT(*) FROM quiz_submissions qs WHERE qs.quiz_id=q.quiz_id AND qs.student_id=%s) AS attempt_count,
                       (SELECT total_score FROM quiz_submissions qs WHERE qs.quiz_id=q.quiz_id AND qs.student_id=%s 
                        ORDER BY submission_date DESC LIMIT 1) AS student_score
                FROM quizzes q
                LEFT JOIN quiz_exceptions qe ON q.quiz_id=qe.quiz_id AND qe.student_id=%s
                WHERE q.course_id=%s AND q.is_deleted=FALSE
                ORDER BY COALESCE(q.publish_at, q.date_published) DESC
            """, (user_id, user_id, user_id, course_id))
            raw_quizzes = cur.fetchall()

            for quiz in raw_quizzes:
                if not quiz['published']:
                    continue

                publish_time = ensure_aware_datetime(quiz['publish_at'] or quiz['date_published']) or now
                duration = quiz['duration_minutes'] or 0

                std_end = publish_time + timedelta(minutes=duration) if duration > 0 else now + timedelta(days=3650)

                if quiz['extended_due_date']:
                    ext = ensure_aware_datetime(quiz['extended_due_date'])
                    if ext > std_end:
                        std_end = ext

                is_expired = now > std_end
                total_allowed = (quiz['max_attempts'] or 1) + (quiz['extra_attempts'] or 0)

                quiz['attempts_left'] = total_allowed - quiz['attempt_count']
                quiz['expired'] = is_expired
                quiz['can_take'] = (not is_expired) and (quiz['attempt_count'] < total_allowed)
                quiz['submitted'] = quiz['attempt_count'] > 0

                quizzes.append(quiz)

            # ----------------------------------------------------------
            # ASSIGNMENTS (FIXED — publish logic + timer)
            # ----------------------------------------------------------
            cur.execute("""
                SELECT a.assignment_id, a.title, a.description, a.due_date, a.points, 
                       a.material_id, a.publish_at, a.published,
                       ae.extended_due_date,
                       s.submission_id IS NOT NULL AS submitted,
                       s.grade, s.feedback, s.file_path
                FROM assignments a
                LEFT JOIN submissions s ON s.assignment_id=a.assignment_id AND s.student_id=%s
                LEFT JOIN assignment_exceptions ae ON a.assignment_id=ae.assignment_id AND ae.student_id=%s
                WHERE a.course_id=%s AND a.is_deleted=FALSE
                ORDER BY a.due_date ASC
            """, (user_id, user_id, course_id))

            assignments = cur.fetchall()

            for a in assignments:

                # ❗ Hide until published
                if not a.get('published', False):
                    a['hidden'] = True
                    a['can_submit'] = False
                    continue

                a['hidden'] = False

                # Countdown starts when assignment is published
                publish_time = ensure_aware_datetime(a.get('publish_at'))
                if publish_time is None:
                    publish_time = now
                a['publish_time'] = publish_time

                # Effective due date (with exceptions)
                eff_due = ensure_aware_datetime(a['extended_due_date'] or a['due_date'])
                a['effective_due_date'] = eff_due

                # Timer calculation
                if eff_due:
                    remaining = eff_due - now
                    a['remaining_time_seconds'] = max(0, int(remaining.total_seconds()))
                    a['expired'] = now > eff_due and not a['submitted']
                else:
                    a['remaining_time_seconds'] = None
                    a['expired'] = False

                a['can_submit'] = not a['expired']

                if a['file_path']:
                    a['file_static'] = url_for('static', filename=a['file_path'])

            # ----------------------------------------------------------
            # ACTIVITIES (UNCHANGED — per your request)
            # ----------------------------------------------------------
            cur.execute("""
                SELECT la.activity_id, la.title, la.description, la.due_date, la.points, 
                       la.material_id, la.publish_at,
                       ae.extended_due_date,
                       s.submission_id IS NOT NULL AS submitted,
                       s.grade, s.feedback, s.file_path
                FROM learning_activities la
                LEFT JOIN learning_activity_submissions s 
                    ON s.activity_id=la.activity_id AND s.student_id=%s
                LEFT JOIN activity_exceptions ae 
                    ON la.activity_id=ae.activity_id AND ae.student_id=%s
                WHERE la.course_id=%s
                ORDER BY la.due_date ASC
            """, (user_id, user_id, course_id))

            activities = cur.fetchall()

            for act in activities:
                eff_due = ensure_aware_datetime(act['extended_due_date'] or act['due_date'])
                act['effective_due_date'] = eff_due
                act['expired'] = (eff_due and now > eff_due) and not act['submitted']

                if act['file_path']:
                    act['file_static'] = url_for('static', filename=act['file_path'])


        # ===========================================================
        # TEACHER VIEW (UNCHANGED)
        # ===========================================================
        else:
            cur.execute("SELECT * FROM quizzes WHERE course_id=%s AND is_deleted=FALSE", (course_id,))
            quizzes = cur.fetchall()

            cur.execute("SELECT * FROM assignments WHERE course_id=%s AND is_deleted=FALSE", (course_id,))
            assignments = cur.fetchall()

            cur.execute("SELECT * FROM learning_activities WHERE course_id=%s", (course_id,))
            activities = cur.fetchall()

    except Exception as e:
        print(f"Error in materials route: {e}")
        obtlp = None
        materials = []
    finally:
        cur.close()
        conn.close()

    return render_template(
        'materials.html',
        user=user,
        course=course,
        materials=materials,
        obtlp=obtlp,
        quizzes=quizzes,
        assignments=assignments,
        activities=activities,
        course_id=course_id
    )


# ========================================================
# GET QUIZ DATA (API LOGIC)
# ========================================================
@auth.route('/api/get_quiz_data/<int:quiz_id>', methods=['GET'])
def get_quiz_data(quiz_id):
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401

    user_id = session['user_id']
    conn = current_app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    
    now = datetime.now(PH_TZ)

    try:
        # Fetch quiz info
        cur.execute("""
            SELECT q.quiz_id, q.title, q.description, q.duration_minutes,
                   q.published, q.date_published, q.publish_at, q.max_attempts,
                   qe.extra_attempts, qe.extended_due_date
            FROM quizzes q
            LEFT JOIN quiz_exceptions qe ON q.quiz_id = qe.quiz_id AND qe.student_id = %s
            WHERE q.quiz_id=%s
        """, (user_id, quiz_id))
        quiz = cur.fetchone()

        if not quiz:
            return jsonify({'status': 'error', 'message': 'Quiz not found'}), 404
        if not quiz['published']:
            return jsonify({'status': 'error', 'message': 'This quiz is not published yet.'}), 403

        # Determine start time
        publish_time = ensure_aware_datetime(quiz['publish_at'] or quiz['date_published'])
        if not publish_time:
            return jsonify({'status': 'error', 'message': 'Configuration Error: Quiz is published but has no date set.'}), 400
        if now < publish_time:
            return jsonify({'status': 'error', 'message': 'Quiz is scheduled for a future date.'}), 403

        # Check attempts
        cur.execute("SELECT COUNT(*) as count FROM quiz_submissions WHERE quiz_id=%s AND student_id=%s", (quiz_id, user_id))
        attempts_done = cur.fetchone()['count']
        total_allowed = (quiz['max_attempts'] or 1) + (quiz['extra_attempts'] or 0)
        if attempts_done >= total_allowed:
            return jsonify({'status': 'error', 'message': 'Maximum attempts reached.'}), 403

        # Calculate absolute end time
        if quiz['duration_minutes'] > 0:
            standard_end_time = publish_time + timedelta(minutes=quiz['duration_minutes'])
        else:
            standard_end_time = now + timedelta(days=3650)  # unlimited

        final_end_time = standard_end_time
        if quiz['extended_due_date']:
            ext_due = ensure_aware_datetime(quiz['extended_due_date'])
            if ext_due > final_end_time:
                final_end_time = ext_due

        # Block if expired
        if now > final_end_time:
            return jsonify({'status': 'error', 'message': 'Time is up! This quiz has expired.'}), 403

        # Fetch questions
        cur.execute("""
            SELECT question_id, question_text, question_type, points,
                   option_a, option_b, option_c, option_d
            FROM quiz_questions
            WHERE quiz_id=%s
            ORDER BY question_id
        """, (quiz_id,))
        questions = cur.fetchall()

        return jsonify({
            'status': 'success',
            'title': quiz['title'],
            'description': quiz['description'],
            'end_time': final_end_time.isoformat(),  # <-- return absolute end time
            'questions': questions
        })

    finally:
        cur.close()
        conn.close()


# ========================================================
# 3. SUBMIT QUIZ (UPDATED: ATTEMPT CHECK & DEADLINE)
# ========================================================
@auth.route('/api/submit_quiz', methods=['POST'])
def submit_quiz():
    
    def normalize_tf(text):
        if not text: return ""
        clean = str(text).strip().upper()
        if len(clean) > 0:
            return clean[0] 
        return ""

    def normalize_text(text):
        if not text: return ""
        return str(text).strip().lower().replace(" ", "")

    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401

    student_id = session['user_id']
    
    try:
        data = request.get_json(force=True)
        quiz_id = data.get('quiz_id')
        answers = data.get('answers', [])
    except:
        return jsonify({'status': 'error', 'message': 'Invalid Data'}), 400

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        # 1. Fetch Constraints (Attempts & Deadlines)
        cur.execute("""
            SELECT q.max_attempts, q.publish_at, q.date_published, q.duration_minutes,
                   qe.extra_attempts, qe.extended_due_date
            FROM quizzes q
            LEFT JOIN quiz_exceptions qe ON q.quiz_id = qe.quiz_id AND qe.student_id = %s
            WHERE q.quiz_id = %s
        """, (student_id, quiz_id))
        quiz_constraints = cur.fetchone()

        if not quiz_constraints:
            return jsonify({'status': 'error', 'message': 'Quiz not found'}), 404

        # 2. Check Attempt Limits
        cur.execute("SELECT COUNT(*) as count FROM quiz_submissions WHERE quiz_id=%s AND student_id=%s", (quiz_id, student_id))
        attempts_done = cur.fetchone()['count']

        base_attempts = quiz_constraints['max_attempts'] or 1
        extra_attempts = quiz_constraints['extra_attempts'] or 0
        
        if attempts_done >= (base_attempts + extra_attempts):
             return jsonify({'status': 'error', 'message': 'Submission rejected: Maximum attempts reached.'}), 403

        # 3. Check Deadline (Security Check)
        now = datetime.now()
        extended_due = quiz_constraints['extended_due_date']
        start_time = quiz_constraints['publish_at'] or quiz_constraints['date_published']
        duration = quiz_constraints['duration_minutes'] or 0

        # Allow a small buffer (e.g., 2 minutes) for network latency
        buffer_time = timedelta(minutes=2)

        if extended_due:
            if now > (extended_due + buffer_time):
                return jsonify({'status': 'error', 'message': 'Submission rejected: Time Limit Exceeded (Exception).'}), 403
        elif start_time:
            end_time = start_time + timedelta(minutes=duration)
            if now > (end_time + buffer_time):
                return jsonify({'status': 'error', 'message': 'Submission rejected: Time Limit Exceeded.'}), 403

        # 4. Create Submission
        cur.execute("""
            INSERT INTO quiz_submissions (quiz_id, student_id, submission_date, total_score) 
            VALUES (%s, %s, NOW(), 0) RETURNING submission_id
        """, (quiz_id, student_id))
        submission_id = cur.fetchone()['submission_id']
        
        total_score = 0

        for ans in answers:
            qid = ans.get('question_id')
            if not qid: continue

            cur.execute("""
                SELECT question_type, correct_answer, correct_answer_text, points 
                FROM quiz_questions WHERE question_id=%s
            """, (qid,))
            q = cur.fetchone()
            if not q: continue

            qtype = (q['question_type'] or "").lower().strip().replace(" ", "_")
            points = q['points'] or 0
            
            db_correct_answer = (q['correct_answer'] or "").strip()
            db_correct_text   = (q['correct_answer_text'] or "").strip()

            val_choice = (ans.get('answer_choice') or "").strip()
            val_text   = (ans.get('answer_text') or "").strip()
            student_input = val_text if val_text else val_choice
            
            score = 0

            # --- SCORING LOGIC ---
            if qtype == 'multiple_choice':
                if val_choice.upper() == db_correct_answer.upper():
                    score = points

            elif qtype == 'true_false':
                std_norm = normalize_tf(student_input)
                db_norm  = normalize_tf(db_correct_answer) 
                if std_norm == db_norm:
                    score = points

            elif qtype in ['identification', 'enumeration']:
                if normalize_text(student_input) == normalize_text(db_correct_text):
                    score = points
            
            # Essay (Manual Grading) - score remains 0 initially
            elif qtype == 'essay':
                score = 0 

            total_score += score

            # Insert Answer
            cur.execute("""
                INSERT INTO quiz_answers (submission_id, question_id, answer_choice, answer_text, score) 
                VALUES (%s,%s,%s,%s,%s)
            """, (submission_id, qid, (val_choice or None), (val_text or None), score))

        # Final Update
        cur.execute("UPDATE quiz_submissions SET total_score=%s WHERE submission_id=%s", (total_score, submission_id))
        conn.commit()

        return jsonify({'status': 'success', 'score': total_score})

    except Exception as e:
        conn.rollback()
        print(f"Submit Quiz Error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        cur.close()
        conn.close() 

from flask import Blueprint, jsonify, session, request
from datetime import datetime
import psycopg2
import psycopg2.extras
from website import create_app  # Ensure this import matches your project structure

@auth.route('/api/review_quiz/<int:quiz_id>')
def review_quiz(quiz_id):
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401

    student_id = session['user_id']

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        # =========================================================
        # 1. FETCH SUBMISSION (From 'quiz_submissions')
        # =========================================================
        cur.execute("""
            SELECT submission_id, total_score, submission_date
            FROM quiz_submissions
            WHERE quiz_id = %s AND student_id = %s
            ORDER BY submission_date DESC
            LIMIT 1
        """, (quiz_id, student_id))

        submission = cur.fetchone()

        if not submission:
            return jsonify({'status': 'error', 'message': 'No attempt found'}), 404

        submission_id = submission['submission_id']

        # =========================================================
        # 2. FETCH QUESTIONS (From 'quiz_questions')
        # =========================================================
        # Added correct_answer_text because submit_quiz uses it for Identification
        cur.execute("""
            SELECT question_id, question_text, correct_answer, correct_answer_text, points, question_type
            FROM quiz_questions
            WHERE quiz_id = %s
            ORDER BY question_id ASC
        """, (quiz_id,))
        questions = cur.fetchall()

        # =========================================================
        # 3. BUILD REVIEW DATA
        # =========================================================
        questions_data = []
        max_score = 0

        for q in questions:
            q_points = q['points'] or 0
            max_score += q_points
            q_type = (q['question_type'] or "").lower().strip()

            # Determine which correct answer to show the student
            if q_type in ['identification', 'enumeration']:
                display_correct_answer = q['correct_answer_text']
            else:
                display_correct_answer = q['correct_answer']

            # =========================================================
            # 4. FETCH STUDENT ANSWER (From 'quiz_answers')
            # =========================================================
            # Note: Changed table from 'student_answers' to 'quiz_answers'
            # Note: Fetching 'answer_choice' and 'answer_text' to determine what they typed
            cur.execute("""
                SELECT answer_text, answer_choice, score
                FROM quiz_answers
                WHERE submission_id = %s AND question_id = %s
                LIMIT 1
            """, (submission_id, q['question_id']))
            
            answer = cur.fetchone()

            if answer:
                # Determine what the user actually answered
                val_choice = answer['answer_choice']
                val_text = answer['answer_text']
                
                # Priority: Text > Choice (matches submit_quiz logic)
                if val_text and str(val_text).strip():
                    user_answer = val_text
                elif val_choice and str(val_choice).strip():
                    user_answer = val_choice
                else:
                    user_answer = ""
                
                earned = float(answer['score'] or 0)
                
                # Logic to determine if correct for the UI (Green/Red checkmark)
                # If earned points equals max points for the question, it's correct.
                is_correct = (earned == float(q_points)) and (q_points > 0)
            else:
                user_answer = "No Answer"
                is_correct = False
                earned = 0.0

            questions_data.append({
                'question_id': q['question_id'],
                'question_text': q['question_text'],
                'correct_answer': display_correct_answer,
                'user_answer': user_answer,
                'is_correct': is_correct,
                'points': q_points,
                'earned_points': earned,
                'type': q['question_type']
            })

        # =========================================================
        # 5. FINAL RESPONSE
        # =========================================================
        response = {
            'status': 'success',
            'total_score': float(submission['total_score']),
            'max_score': float(max_score),
            # Pass if score is >= 50%
            'passed': float(submission['total_score']) >= (float(max_score) * 0.5),
            'date_taken': submission['submission_date'].strftime('%b %d, %Y %I:%M %p'),
            'questions': questions_data
        }

        return jsonify(response)

    except Exception as e:
        print(f"------------ SERVER ERROR IN REVIEW_QUIZ ------------")
        print(e)
        return jsonify({'status': 'error', 'message': str(e)}), 500

    finally:
        if cur: cur.close()
        if conn: conn.close()

UPLOAD_ASSIGNMENTS_FOLDER = 'static/uploads/assignments'
UPLOAD_ACTIVITIES_FOLDER = 'static/uploads/activities'

from flask import session, request, redirect, url_for, flash, current_app
from werkzeug.utils import secure_filename
from flask_mail import Message
from threading import Thread # Import Threading
import os
import psycopg2
import psycopg2.extras
from datetime import datetime, date, time

# ========================================================
# 1. SUBMIT ASSIGNMENT ROUTE (FIXED)
# ========================================================
@auth.route('/submit_assignment/<int:course_id>', methods=['POST'])
def submit_assignment(course_id):
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
        
    user_id = session['user_id']
    raw_assign_id = request.form.get('assignment_id')
    file = request.files.get('file')

    if not raw_assign_id or not file or file.filename == '':
        flash('Missing file or Assignment ID.', 'danger')
        return redirect(url_for('auth.materials', course_id=course_id))

    try:
        assignment_id = int(raw_assign_id)
    except ValueError:
        flash('Invalid Assignment ID.', 'danger')
        return redirect(url_for('auth.materials', course_id=course_id))

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        # 1. Check Duplicate
        cur.execute("SELECT submission_id FROM submissions WHERE assignment_id=%s AND student_id=%s", (assignment_id, user_id))
        if cur.fetchone():
            flash('You have already submitted this assignment.', 'warning')
            return redirect(url_for('auth.materials', course_id=course_id))

        # 2. Check Expiration (With Date Type Fix)
        cur.execute("""
            SELECT a.due_date, ae.extended_due_date 
            FROM assignments a
            LEFT JOIN assignment_exceptions ae ON a.assignment_id = ae.assignment_id AND ae.student_id = %s
            WHERE a.assignment_id = %s
        """, (user_id, assignment_id))
        check = cur.fetchone()

        if check:
            eff_deadline = check['extended_due_date'] if check['extended_due_date'] else check['due_date']
            
            # --- FIX FOR TYPE ERROR ---
            if eff_deadline:
                # If it's a simple date (no time), convert it to end-of-day datetime
                if isinstance(eff_deadline, date) and not isinstance(eff_deadline, datetime):
                    eff_deadline = datetime.combine(eff_deadline, time.max)
                
                if datetime.now() > eff_deadline:
                    flash('Submission rejected: This assignment is past due.', 'danger')
                    return redirect(url_for('auth.materials', course_id=course_id))
            # --------------------------

        # 3. Save File
        filename = secure_filename(file.filename)
        base_path = os.path.join(current_app.root_path, 'static', 'uploads', 'assignments')
        os.makedirs(base_path, exist_ok=True)
        
        disk_path = os.path.join(base_path, filename)
        if os.path.exists(disk_path):
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            filename = f"{timestamp}_{filename}"
            disk_path = os.path.join(base_path, filename)

        file.save(disk_path)
        db_path = f"uploads/assignments/{filename}"

        # 4. Insert Record (Added Feedback column explicitly)
        cur.execute("""
            INSERT INTO submissions 
            (assignment_id, student_id, file_path, submission_date, grade, feedback, status) 
            VALUES (%s, %s, %s, NOW(), NULL, NULL, 'Submitted')
        """, (assignment_id, user_id, db_path))
        
        conn.commit()
        flash('Assignment submitted successfully!', 'success')

    except Exception as e:
        conn.rollback()
        print(f"Assign Submit Error: {e}")
        flash(f"Error submitting assignment: {str(e)}", 'danger')
    finally:
        cur.close()
        conn.close()

    return redirect(url_for('auth.materials', course_id=course_id))



# ========================================================
# 2. SUBMIT ACTIVITY ROUTE (FIXED)
# ========================================================
@auth.route('/submit_activity/<int:course_id>', methods=['POST'])
def submit_activity(course_id):
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
        
    user_id = session['user_id']
    raw_activity_id = request.form.get('activity_id')
    file = request.files.get('file')

    if not raw_activity_id or not file or file.filename == '':
        flash('Missing file or Activity ID.', 'danger')
        return redirect(url_for('auth.materials', course_id=course_id))

    try:
        activity_id = int(raw_activity_id)
    except ValueError:
        flash('Invalid Activity ID.', 'danger')
        return redirect(url_for('auth.materials', course_id=course_id))

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        # 1. Check Duplicate
        cur.execute("SELECT submission_id FROM learning_activity_submissions WHERE activity_id=%s AND student_id=%s", (activity_id, user_id))
        if cur.fetchone():
            flash('You have already submitted this activity.', 'warning')
            return redirect(url_for('auth.materials', course_id=course_id))

        # 2. Check Expiration (With Date Type Fix)
        cur.execute("""
            SELECT la.due_date, ae.extended_due_date 
            FROM learning_activities la
            LEFT JOIN activity_exceptions ae ON la.activity_id = ae.activity_id AND ae.student_id = %s
            WHERE la.activity_id = %s
        """, (user_id, activity_id))
        check = cur.fetchone()

        if check:
            eff_deadline = check['extended_due_date'] if check['extended_due_date'] else check['due_date']
            
            # --- FIX FOR TYPE ERROR ---
            if eff_deadline:
                if isinstance(eff_deadline, date) and not isinstance(eff_deadline, datetime):
                    eff_deadline = datetime.combine(eff_deadline, time.max)
                
                if datetime.now() > eff_deadline:
                    flash('Submission rejected: This activity is past due.', 'danger')
                    return redirect(url_for('auth.materials', course_id=course_id))
            # --------------------------

        # 3. Save File
        filename = secure_filename(file.filename)
        base_path = os.path.join(current_app.root_path, 'static', 'uploads', 'activities')
        os.makedirs(base_path, exist_ok=True)
        
        disk_path = os.path.join(base_path, filename)
        if os.path.exists(disk_path):
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            filename = f"{timestamp}_{filename}"
            disk_path = os.path.join(base_path, filename)

        file.save(disk_path)
        db_path = f"uploads/activities/{filename}"

        # 4. Insert Record (Added Feedback column explicitly)
        cur.execute("""
            INSERT INTO learning_activity_submissions 
            (activity_id, student_id, file_path, created_at, updated_at, grade, feedback) 
            VALUES (%s, %s, %s, NOW(), NOW(), NULL, NULL)
        """, (activity_id, user_id, db_path))
        
        conn.commit()
        flash('Activity submitted successfully!', 'success')

    except Exception as e:
        conn.rollback()
        print(f"Activity Submit Error: {e}")
        flash(f"Error submitting activity: {str(e)}", 'danger')
    finally:
        cur.close()
        conn.close()

    return redirect(url_for('auth.materials', course_id=course_id))

# ===============================
# VIEW ANNOUNCEMENTS (Student Only)
# ===============================
@auth.route('/course_overview/<int:course_id>', methods=['GET'])
def course_overview(course_id):
    # 1️⃣ Authentication
    if 'user_id' not in session:
        flash('Please log in first.', 'warning')
        return redirect(url_for('auth.login'))

    # 2️⃣ Authorization: students only
    if session.get('role_id') != 3:
        flash('Access denied. Students only.', 'danger')
        return redirect(url_for('auth.login'))

    user_id = session['user_id']

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        # ===============================
        # FETCH LOGGED-IN USER INFO
        # ===============================
        cur.execute("""
            SELECT user_id, full_name, email
            FROM users
            WHERE user_id = %s;
        """, (user_id,))
        user = cur.fetchone()

        # ===============================
        # FETCH COURSE INFO (with teacher, sem/year, section)
        # ===============================
        cur.execute("""
            SELECT 
                c.course_id, 
                c.course_code, 
                c.course_title, 
                c.course_description,
                p.program_name,
                s.section_name,   -- ✅ Added section name
                a.start_year, 
                a.end_year, 
                a.semester,
                CONCAT(a.start_year, '-', a.end_year, ' (', a.semester, ' Semester)') AS academic_period,
                u.full_name AS instructor_name
            FROM courses c
            JOIN programs p ON c.program_id = p.program_id
            JOIN academic_years a ON c.academic_year_id = a.academic_year_id
            LEFT JOIN users u ON c.user_id = u.user_id
            LEFT JOIN sections s ON c.section_id = s.section_id  -- ✅ Added join
            WHERE c.course_id = %s;
        """, (course_id,))
        course = cur.fetchone()
        if not course:
            flash('Course not found.', 'danger')
            return redirect(url_for('auth.dashboard'))

        # ===============================
        # FETCH ANNOUNCEMENTS
        # ===============================
        cur.execute("""
            SELECT announcement_id, title, content, date_posted
            FROM announcements
            WHERE course_id = %s
            ORDER BY date_posted DESC;
        """, (course_id,))
        announcements = cur.fetchall()

        # ===============================
        # FETCH UNREAD NOTIFICATIONS
        # ===============================
        cur.execute("""
            SELECT n.notification_id, n.course_id, n.title, n.message, n.read, n.created_at,
                   c.course_code
            FROM notifications n
            LEFT JOIN courses c ON n.course_id = c.course_id
            WHERE n.user_id = %s AND n.read = FALSE AND n.course_id = %s
            ORDER BY n.created_at DESC;
        """, (user_id, course_id))
        notifications = cur.fetchall()

        # Limit to max 20
        notifications = notifications[:20]
        unread_count = len(notifications)

    finally:
        cur.close()
        conn.close()

    return render_template(
        'course_overview.html',
        user=user,
        course=course,
        announcements=announcements,
        notifications=notifications,
        unread_count=unread_count
    )

# ===============================
# VIEW PERFORMANCE OVERVIEW (Student)
# ===============================
@auth.route('/performance_overview/<int:course_id>', methods=['GET'])
def performance_overview(course_id):
    # Check if user is logged in
    if 'user_id' not in session:
        flash('Please log in first.', 'warning')
        return redirect(url_for('auth.login'))

    # Only students
    if session.get('role_id') != 3:
        flash('Access denied. Students only.', 'danger')
        return redirect(url_for('auth.login'))

    user_id = session['user_id']
    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        # =======================================
        # 1️⃣ USER INFO
        # =======================================
        cur.execute("""
            SELECT user_id, full_name, email
            FROM users
            WHERE user_id = %s;
        """, (user_id,))
        user = cur.fetchone()

        # =======================================
        # 2️⃣ COURSE INFO + SECTION NAME
        # =======================================
        cur.execute("""
            SELECT 
                c.course_id,
                c.course_code,
                c.course_title,
                p.program_name,
                s.section_name,
                a.start_year,
                a.end_year,
                a.semester,
                CONCAT(a.start_year, '-', a.end_year, ' (', a.semester, ' Semester)') AS academic_period
            FROM courses c
            JOIN programs p ON c.program_id = p.program_id
            JOIN academic_years a ON c.academic_year_id = a.academic_year_id
            LEFT JOIN sections s ON c.section_id = s.section_id
            WHERE c.course_id = %s;
        """, (course_id,))
        course = cur.fetchone()

        # =======================================
        # 3️⃣ ASSIGNMENTS
        # =======================================
        cur.execute("""
            SELECT a.title AS assignment_title,
                   s.submission_date,
                   s.grade,
                   s.remarks
            FROM assignments a
            LEFT JOIN submissions s 
                ON a.assignment_id = s.assignment_id 
               AND s.student_id = %s
            WHERE a.course_id = %s
            ORDER BY a.due_date ASC;
        """, (user_id, course_id))
        assignments = cur.fetchall()

        # =======================================
        # 4️⃣ QUIZZES
        # =======================================
        cur.execute("""
            SELECT q.title AS quiz_title,
                   qs.submission_date,
                   qs.total_score,
                   COALESCE(
                       (SELECT SUM(points) 
                        FROM quiz_questions 
                        WHERE quiz_id = q.quiz_id),
                       25
                   ) AS max_points
            FROM quizzes q
            LEFT JOIN quiz_submissions qs 
                ON q.quiz_id = qs.quiz_id 
               AND qs.student_id = %s
            WHERE q.course_id = %s
            ORDER BY q.date_created ASC;
        """, (user_id, course_id))
        quizzes = cur.fetchall()

        # =======================================
        # 5️⃣ LEARNING ACTIVITIES (FIXED)
        # =======================================
        cur.execute("""
            SELECT la.activity_id,
                   la.title,
                   la.description,
                   la.points,
                   la.due_date,
                   la.publish_at,
                   las.submission_id,
                   las.submission_date,        -- FIXED COLUMN
                   las.grade AS student_grade,
                   las.feedback
            FROM learning_activities la
            LEFT JOIN learning_activity_submissions las 
                ON las.activity_id = la.activity_id 
               AND las.student_id = %s
            WHERE la.course_id = %s
              AND la.published = TRUE
              AND (la.publish_at IS NULL OR la.publish_at <= NOW())
            ORDER BY la.due_date ASC;
        """, (user_id, course_id))
        activities = cur.fetchall()

        # =======================================
        # 6️⃣ GRADE COMPUTATIONS
        # =======================================
        total_points_earned = 0.0
        total_points_possible = 0.0

        # ----- ASSIGNMENTS -----
        assignment_scores = []
        for a in assignments:
            if a.get('grade') is not None:
                g = float(a['grade'])
                assignment_scores.append(g)
                total_points_earned += g
                total_points_possible += 100.0

        assignment_avg = round(sum(assignment_scores) / len(assignment_scores), 2) if assignment_scores else 0.0
        assignment_status = "Above average" if assignment_avg >= 85 else "Needs improvement"

        # ----- QUIZZES -----
        quiz_total_earned = 0.0
        quiz_total_possible = 0.0
        for q in quizzes:
            if q.get('total_score') is not None:
                ts = float(q['total_score'])
                mp = float(q['max_points'])
                quiz_total_earned += ts
                quiz_total_possible += mp
                total_points_earned += ts
                total_points_possible += mp

        quiz_avg = round((quiz_total_earned / quiz_total_possible) * 100.0, 2) if quiz_total_possible > 0 else 0.0
        quiz_status = "Good performance" if quiz_avg >= 85 else "Needs improvement"

        # ----- LEARNING ACTIVITIES -----
        activity_total_earned = 0.0
        activity_total_possible = 0.0

        for act in activities:
            if act.get('student_grade') is not None:
                sg = float(act['student_grade'])
                pts = float(act['points'])
                activity_total_earned += sg
                activity_total_possible += pts
                total_points_earned += sg
                total_points_possible += pts

        activity_avg = round((activity_total_earned / activity_total_possible) * 100.0, 2) if activity_total_possible > 0 else 0.0
        activity_status = "Good performance" if activity_avg >= 85 else "Needs improvement"

        # ----- OVERALL -----
        overall_grade = round((total_points_earned / total_points_possible) * 100.0, 2) if total_points_possible > 0 else 0.0

        # Letter Grade
        if overall_grade >= 90:
            grade_letter = "A"
        elif overall_grade >= 85:
            grade_letter = "B+"
        elif overall_grade >= 80:
            grade_letter = "B"
        elif overall_grade >= 75:
            grade_letter = "C"
        else:
            grade_letter = "F"

    finally:
        cur.close()
        conn.close()

    # =======================================
    # RENDER PAGE
    # =======================================
    return render_template(
        'student_performance.html',
        user=user,
        course_id=course_id,
        course=course,
        assignments=assignments,
        quizzes=quizzes,
        activities=activities,
        overall_grade=overall_grade,
        grade_letter=grade_letter,
        total_points=int(total_points_earned),
        max_points=int(total_points_possible),
        assignment_avg=assignment_avg,
        assignment_status=assignment_status,
        quiz_avg=quiz_avg,
        quiz_status=quiz_status,
        activity_avg=activity_avg,
        activity_status=activity_status
    )


@auth.route('/teachers-dashboard')
def teachers_dashboard():
    if 'user_id' not in session or session.get('role_id') != 2:
        flash('Please log in as a teacher to access this page.', 'warning')
        return redirect(url_for('auth.login'))

    # Fetch teacher info
    from . import create_app
    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor()
    user_id = session['user_id']
    cur.execute("SELECT user_id, full_name, student_id, email FROM users WHERE user_id = %s", (user_id,))
    user = cur.fetchone()
    cur.close()
    conn.close()

    if not user:
        flash('User not found.', 'danger')
        return redirect(url_for('auth.login'))

    return render_template('teachers_dashboard.html', user={
        'user_id': user[0],
        'full_name': user[1],
        'student_id': user[2],
        'email': user[3]
    })

# === File upload configuration ===
UPLOAD_FOLDER = 'static/uploads'
ALLOWED_EXTENSIONS = {'pdf', 'docx', 'pptx', 'mp4', 'jpg', 'png'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ======================================================
# TEACHER LMS - SHOW ALL COURSES ASSIGNED TO THE TEACHER
# ======================================================
@auth.route('/teacher_lms', methods=['GET'])
def teacher_lms():
    # Require login
    if 'user_id' not in session:
        flash('Please log in first.', 'warning')
        return redirect(url_for('auth.login'))

    # Only allow teachers
    if session.get('role_id') != 2:
        flash('Access denied. Teachers only.', 'danger')
        return redirect(url_for('auth.login'))

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    user_id = session['user_id']

    # 🧑‍🏫 Fetch teacher info
    cur.execute("""
        SELECT user_id, full_name, email, role_id, program_id
        FROM users
        WHERE user_id = %s
    """, (user_id,))
    user = cur.fetchone()

    # 📚 Fetch dropdowns
    cur.execute("SELECT program_id, program_name FROM programs ORDER BY program_name;")
    programs = cur.fetchall()

    cur.execute("""
        SELECT academic_year_id, start_year, end_year, semester
        FROM academic_years
        WHERE status = 'active'
        ORDER BY start_year DESC, semester;
    """)
    academic_years = cur.fetchall()

    cur.execute("SELECT section_id, section_name FROM sections ORDER BY section_name;")
    sections = cur.fetchall()

    cur.execute("SELECT year_level_id, year_name FROM year_levels ORDER BY year_level_id;")
    year_levels = cur.fetchall()

    # 🧾 Filters
    selected_academic_year_id = request.args.get('academic_year_id')
    selected_semester = request.args.get('semester')

    # 🧠 Fetch teacher’s courses (active academic years only)
    query = """
        SELECT c.course_id, c.course_code, c.course_title, c.course_description,
               c.units, yl.year_name AS year_level, c.status,
               p.program_name, s.section_name,
               CONCAT(a.start_year, '-', a.end_year, ' (', a.semester, ' Semester)') AS academic_year
        FROM courses c
        JOIN programs p ON c.program_id = p.program_id
        JOIN academic_years a ON c.academic_year_id = a.academic_year_id
        JOIN year_levels yl ON c.year_level_id = yl.year_level_id
        LEFT JOIN sections s ON c.section_id = s.section_id
        WHERE c.user_id = %s AND a.status = 'active'
    """
    params = [user_id]

    if selected_academic_year_id and selected_academic_year_id != "all":
        query += " AND c.academic_year_id = %s"
        params.append(selected_academic_year_id)

    if selected_semester and selected_semester != "all":
        query += " AND a.semester = %s"
        params.append(selected_semester)

    query += " ORDER BY yl.year_level_id, c.course_code;"
    cur.execute(query, params)
    courses = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        'teacher_lms.html',
        user=user,
        programs=programs,
        academic_years=academic_years,
        sections=sections,
        year_levels=year_levels,
        courses=courses,
        selected_academic_year_id=selected_academic_year_id,
        selected_semester=selected_semester
    )

from flask import current_app
from threading import Thread

ALLOWED_EXTENSIONS = {'pdf'}

def allowed_pdf(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

from flask import render_template, redirect, url_for, session, request, flash, current_app
import psycopg2.extras
import os
from werkzeug.utils import secure_filename

@auth.route('/upload_obtlp', methods=['GET', 'POST'])
def upload_obtlp():
    if 'user_id' not in session:
        flash('Please log in first.', 'warning')
        return redirect(url_for('auth.login'))

    if session.get('role_id') != 2:
        flash('Access denied. Teachers only.', 'danger')
        return redirect(url_for('auth.login'))

    app = create_app()
    conn = app.get_db_connection()

    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        # Teacher info
        cur.execute("SELECT user_id, full_name, email FROM users WHERE user_id=%s", (session['user_id'],))
        user = cur.fetchone()

        # Fetch all programs
        cur.execute("SELECT program_id, program_name FROM programs ORDER BY program_name")
        programs = cur.fetchall()

        # Fetch courses with sections for dropdown
        cur.execute("""
            SELECT c.course_id, c.course_title, c.program_id, s.section_name
            FROM courses c
            LEFT JOIN sections s ON c.section_id = s.section_id
            WHERE c.user_id = %s
            ORDER BY c.course_title
        """, (session['user_id'],))
        courses = cur.fetchall()

        # --- Soft Delete ---
        delete_id = request.args.get('delete_id')
        if delete_id:
            cur.execute("UPDATE obtlp_files SET is_deleted = TRUE WHERE id=%s", (delete_id,))
            conn.commit()
            flash("OBTLP file deleted (soft delete).", "success")
            return redirect(url_for('auth.upload_obtlp'))

        # --- Edit Existing File ---
        edit_id = request.args.get('edit_id')
        file_to_edit = None
        if edit_id:
            cur.execute("""
                SELECT * FROM obtlp_files
                WHERE id=%s AND uploaded_by=%s AND is_deleted IS DISTINCT FROM TRUE
            """, (edit_id, session['user_id']))
            file_to_edit = cur.fetchone()

        # --- Handle Upload / Update ---
        if request.method == 'POST':
            program_id = request.form.get('program_id')
            course_id = request.form.get('course_id')
            title = request.form.get('title', '').strip()
            description = request.form.get('description', '').strip()
            files = request.files.getlist('files')
            existing_file_id = request.form.get('file_id')  # for edits

            if not program_id or not course_id:
                flash("Please select a program and course.", "danger")
                return redirect(url_for('auth.upload_obtlp'))

            save_folder = os.path.join(current_app.root_path, 'static', 'obtlp')
            os.makedirs(save_folder, exist_ok=True)

            uploaded_files = []

            # If editing, update title/description/course/program without changing file
            if existing_file_id and not files[0].filename:
                cur.execute("""
                    UPDATE obtlp_files
                    SET program_id=%s, course_id=%s, title=%s, description=%s
                    WHERE id=%s AND uploaded_by=%s
                """, (program_id, course_id, title, description, existing_file_id, session['user_id']))
                conn.commit()
                flash("OBTLP file updated successfully!", "success")
                return redirect(url_for('auth.upload_obtlp'))

            # Otherwise, handle new file upload(s)
            for file in files:
                if file and allowed_pdf(file.filename):
                    filename = secure_filename(file.filename)
                    base, ext = os.path.splitext(filename)
                    counter = 1
                    file_path = os.path.join(save_folder, filename)
                    while os.path.exists(file_path):
                        filename = f"{base}_{counter}{ext}"
                        file_path = os.path.join(save_folder, filename)
                        counter += 1
                    file.save(file_path)
                    relative_path = os.path.join('obtlp', filename).replace("\\", "/")

                    cur.execute("""
                        INSERT INTO obtlp_files (program_id, course_id, title, description, file_path, file_type, uploaded_by)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """, (program_id, course_id, title, description, relative_path, ext[1:], session['user_id']))

                    uploaded_files.append(filename)

            conn.commit()
            if uploaded_files:
                flash(f"{len(uploaded_files)} file(s) uploaded successfully!", "success")
            return redirect(url_for('auth.upload_obtlp'))

        # --- Fetch uploaded files ---
        cur.execute("""
            SELECT o.*, c.course_title, c.section_id, s.section_name, p.program_name
            FROM obtlp_files o
            LEFT JOIN courses c ON o.course_id = c.course_id
            LEFT JOIN programs p ON o.program_id = p.program_id
            LEFT JOIN sections s ON c.section_id = s.section_id
            WHERE o.uploaded_by = %s AND (o.is_deleted IS DISTINCT FROM TRUE)
            ORDER BY o.upload_date DESC
        """, (session['user_id'],))
        obtlp_files = cur.fetchall()
        for f in obtlp_files:
            f['static_path'] = url_for('static', filename=f['file_path'])

    conn.close()
    return render_template(
        'upload_obtlp.html', 
        user=user, 
        programs=programs, 
        courses=courses, 
        obtlp_files=obtlp_files,
        file_to_edit=file_to_edit
    )


from flask import Blueprint, render_template, request, redirect, url_for, flash, session, current_app, jsonify
import psycopg2
import psycopg2.extras
from psycopg2 import sql # Added for dynamic exception handling
import os
from werkzeug.utils import secure_filename
from threading import Thread
from flask_mail import Message # Assuming you are using Flask-Mail based on your snippet

def allowed_pdf(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() == 'pdf'

# ========================================================
# UPLOAD MATERIALS (With Email & DB Notifications)
# ========================================================
@auth.route('/upload_materials/<int:course_id>', methods=['GET', 'POST'])
def upload_materials(course_id):
    if 'user_id' not in session:
        flash('Please log in first.', 'warning')
        return redirect(url_for('auth.login'))
    
    # Role check (Teacher = 2)
    if session.get('role_id') != 2:
        flash('Access denied. Teachers only.', 'danger')
        return redirect(url_for('auth.login'))

    conn = current_app.get_db_connection()
    
    # Initialize to avoid UnboundLocalError on fail
    course = None
    materials = []
    user = None

    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            # --- 1. Fetch current user (Teacher) ---
            cur.execute("SELECT user_id, full_name, email FROM users WHERE user_id=%s", (session['user_id'],))
            user = cur.fetchone()
            teacher_name = user['full_name']

            # --- 2. Fetch course info ---
            cur.execute("""
                SELECT c.course_id, c.course_code, c.course_title, s.section_name
                FROM courses c
                LEFT JOIN sections s ON c.section_id = s.section_id
                WHERE c.course_id=%s
            """, (course_id,))
            course = cur.fetchone()

            # --- 3. Handle delete material ---
            delete_id = request.args.get('delete_material_id')
            if delete_id:
                cur.execute("SELECT file_path FROM materials WHERE material_id=%s AND course_id=%s", (delete_id, course_id))
                mat = cur.fetchone()
                if mat:
                    try: 
                        os.remove(os.path.join(current_app.root_path, 'static', mat['file_path']))
                    except Exception: 
                        pass
                    cur.execute("DELETE FROM materials WHERE material_id=%s", (delete_id,))
                    conn.commit()
                    flash('Material deleted successfully.', 'success')
                return redirect(url_for('auth.upload_materials', course_id=course_id))

            # --- 4. Handle upload ---
            if request.method == 'POST' and 'files' in request.files:
                files = request.files.getlist('files')
                title = request.form.get('material_title', '').strip()
                description = request.form.get('material_description', '').strip()
                term = request.form.get('term', 'Prelims')

                save_folder = os.path.join(current_app.root_path, 'static', 'uploads')
                os.makedirs(save_folder, exist_ok=True)

                files_uploaded = False

                for file in files:
                    if file and allowed_pdf(file.filename): # Ensure allowed_pdf is defined in your utils
                        filename = secure_filename(file.filename)
                        # Handle duplicate filenames
                        base, ext = os.path.splitext(filename)
                        counter = 1
                        file_path = os.path.join(save_folder, filename)
                        while os.path.exists(file_path):
                            filename = f"{base}_{counter}{ext}"
                            file_path = os.path.join(save_folder, filename)
                            counter += 1
                        
                        file.save(file_path)
                        db_path = os.path.join('uploads', filename).replace('\\', '/')

                        cur.execute("""
                            INSERT INTO materials (course_id, title, description, term, file_path, material_type, upload_date, is_deleted)
                            VALUES (%s, %s, %s, %s, %s, 'pdf', NOW(), FALSE)
                        """, (course_id, title, description, term, db_path))
                        files_uploaded = True
                
                conn.commit()

                # ===========================================
                # NOTIFICATION LOGIC (If upload successful)
                # ===========================================
                if files_uploaded:
                    # A. Fetch Enrolled Students
                    cur.execute("""
                        SELECT u.user_id, u.full_name, u.email
                        FROM users u
                        JOIN courses c ON u.section_id = c.section_id
                        WHERE u.role_id = 3 AND c.course_id = %s;
                    """, (course_id,))
                    students = cur.fetchall()

                    # B. Prepare Data for Thread
                    course_code = course['course_code']
                    mat_title = title if title else "New Course Material"
                    mat_desc = description

                    # C. Define Thread Function
                    app = current_app._get_current_object()

                    def notify_students_upload():
                        with app.app_context():
                            thread_conn = app.get_db_connection()
                            thread_cur = thread_conn.cursor()
                            
                            try:
                                link_url = url_for('auth.login', _external=True)
                            except:
                                link_url = "#"

                            try:
                                for student in students:
                                    # 1. Insert DB Notification
                                    thread_cur.execute("""
                                        INSERT INTO notifications (user_id, course_id, title, message, created_at, read)
                                        VALUES (%s, %s, %s, %s, NOW(), FALSE)
                                    """, (
                                        student['user_id'],
                                        course_id,
                                        f"New Material: {mat_title}",
                                        f"{teacher_name} uploaded new material in {course_code}."
                                    ))

                                    # 2. HTML Email Content
                                    email_html = f"""
                                    <!DOCTYPE html>
                                    <html>
                                    <head>
                                        <style>
                                            body {{ font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif; background-color: #f4f6f8; margin: 0; padding: 0; }}
                                            .email-container {{ max-width: 600px; margin: 30px auto; background-color: #ffffff; border-radius: 8px; overflow: hidden; box-shadow: 0 4px 6px rgba(0,0,0,0.1); border: 1px solid #e0e0e0; }}
                                            .header {{ background-color: #3498db; color: #ffffff; padding: 25px; text-align: center; }} /* Blue for Material */
                                            .header h1 {{ margin: 0; font-size: 24px; font-weight: 600; }}
                                            .content {{ padding: 30px; color: #333333; line-height: 1.6; }}
                                            .info-box {{ background-color: #f8f9fa; border-left: 5px solid #3498db; padding: 15px 20px; margin: 20px 0; border-radius: 4px; }}
                                            .info-box h3 {{ margin-top: 0; color: #2980b9; }}
                                            .btn-container {{ text-align: center; margin-top: 30px; margin-bottom: 20px; }}
                                            .btn {{ background-color: #27ae60; color: white !important; padding: 12px 25px; text-decoration: none; border-radius: 5px; font-weight: bold; display: inline-block; }}
                                            .footer {{ background-color: #eceff1; padding: 15px; text-align: center; font-size: 12px; color: #7f8c8d; }}
                                        </style>
                                    </head>
                                    <body>
                                        <div class="email-container">
                                            <div class="header">
                                                <h1>New Material Uploaded</h1>
                                            </div>
                                            <div class="content">
                                                <p>Hi <strong>{student['full_name']}</strong>,</p>
                                                <p>Your instructor, <strong>{teacher_name}</strong>, has uploaded new learning material for the course <strong>{course_code}</strong>.</p>
                                                
                                                <div class="info-box">
                                                    <h3>{mat_title}</h3>
                                                    <p>{mat_desc if mat_desc else 'No description provided.'}</p>
                                                </div>

                                                <p>Log in to the portal to view or download the files.</p>
                                                
                                                <div class="btn-container">
                                                    <a href="{link_url}" class="btn">View Materials</a>
                                                </div>
                                            </div>
                                            <div class="footer">
                                                <p>This is an automated notification from your Learning Management System.</p>
                                                <p>&copy; 2024 LMS Team</p>
                                            </div>
                                        </div>
                                    </body>
                                    </html>
                                    """

                                    # 3. Send Email
                                    try:
                                        msg = Message(
                                            subject=f"New Material: {mat_title} ({course_code})",
                                            recipients=[student['email']],
                                            html=email_html
                                        )
                                        mail.send(msg)
                                    except Exception as e:
                                        print(f"Failed to send email to {student['email']}: {e}")

                                thread_conn.commit()
                            except Exception as e:
                                print(f"Error in material notification thread: {e}")
                                thread_conn.rollback()
                            finally:
                                thread_cur.close()
                                thread_conn.close()

                    # D. Start Thread
                    Thread(target=notify_students_upload).start()

                flash('Materials uploaded and students notified!', 'success')
                return redirect(url_for('auth.upload_materials', course_id=course_id))

            # --- 5. Fetch materials ---
            cur.execute("""
                SELECT * FROM materials
                WHERE course_id=%s AND is_deleted=FALSE
                ORDER BY 
                    CASE WHEN term='Prelims' THEN 1
                         WHEN term='Midterms' THEN 2
                         WHEN term='Finals' THEN 3
                         ELSE 4 END,
                    upload_date DESC
            """, (course_id,))
            materials = cur.fetchall()

            # --- 6. Fetch quizzes ---
            cur.execute("""
                SELECT q.*, 
                       (SELECT COUNT(*) FROM quiz_questions WHERE quiz_id=q.quiz_id) as total_questions
                FROM quizzes q
                WHERE q.course_id=%s AND is_deleted=FALSE AND material_id IS NOT NULL
                ORDER BY q.date_created DESC
            """, (course_id,))
            all_quizzes = cur.fetchall()

            # --- 7. Fetch assignments ---
            cur.execute("""
                SELECT a.*, 
                       (SELECT COUNT(*) FROM submissions WHERE assignment_id=a.assignment_id) as sub_count
                FROM assignments a
                WHERE a.course_id=%s AND is_deleted=FALSE AND material_id IS NOT NULL
                ORDER BY a.date_created DESC
            """, (course_id,))
            all_assignments = cur.fetchall()

            # --- 8. Fetch learning activities ---
            cur.execute("""
                SELECT la.*,
                       (SELECT COUNT(*) FROM learning_activity_submissions s WHERE s.activity_id = la.activity_id) as sub_count
                FROM learning_activities la
                WHERE la.course_id=%s AND la.material_id IS NOT NULL
                ORDER BY la.created_at DESC
            """, (course_id,))
            all_activities = cur.fetchall()

            # --- 9. Attach data to materials ---
            for m in materials:
                m['static_path'] = url_for('static', filename=m['file_path'])
                
                # Map based on material_id
                m['quizzes'] = [q for q in all_quizzes if q['material_id'] == m['material_id']]
                m['assignments'] = [a for a in all_assignments if a['material_id'] == m['material_id']]
                m['activities'] = [la for la in all_activities if la['material_id'] == m['material_id']]

    except Exception as e:
        print(f"Error in upload_materials: {e}")
        flash(f"An error occurred: {str(e)}", "danger")
    finally:
        conn.close()

    return render_template('upload_materials.html',
                           course=course,
                           materials=materials,
                           user=user,
                           course_id=course_id)

@auth.route('/edit_material/<int:material_id>', methods=['POST'])
def edit_material(material_id):
    conn = current_app.get_db_connection()
    try:
        title = request.form.get('material_title')
        desc = request.form.get('material_description')
        term = request.form.get('term') 

        with conn.cursor() as cur:
            cur.execute("""
                UPDATE materials 
                SET title = %s, description = %s, term = %s
                WHERE material_id = %s
            """, (title, desc, term, material_id))
            conn.commit()
            
        return jsonify({'status': 'success', 'message': 'Material updated successfully!'})
    except Exception as e:
        conn.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

@auth.route('/create_quiz/<int:course_id>/<int:material_id>', methods=['GET'])
def create_quiz(course_id, material_id):
    conn = current_app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute("SELECT * FROM materials WHERE course_id=%s AND is_deleted=FALSE ORDER BY date_uploaded DESC;", (course_id,))
        materials = cur.fetchall()
        
        return render_template('upload_materials.html', 
                               course_id=course_id, 
                               material_id=material_id, 
                               materials=materials)
    except Exception as e:
        return f"Error loading page: {str(e)}", 500
    finally:
        cur.close()
        conn.close()

# ==========================================
# 1. GET QUIZ DETAILS (For Editing)
# ==========================================
@auth.route('/get_quiz_details/<int:quiz_id>', methods=['GET'])
def get_quiz_details(quiz_id):
    conn = current_app.get_db_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            # UPDATED: Fetch max_attempts and publish_at
            cur.execute("""
                SELECT quiz_id, title, description, duration_minutes, course_id, material_id,
                       max_attempts, TO_CHAR(publish_at, 'YYYY-MM-DD"T"HH24:MI') as publish_at
                FROM quizzes WHERE quiz_id = %s
            """, (quiz_id,))
            quiz = cur.fetchone()
            
            if not quiz:
                return jsonify({'status': 'error', 'message': 'Quiz not found'}), 404

            # Get Questions
            cur.execute("""
                SELECT question_text, question_type, option_a, option_b, option_c, option_d, 
                       correct_answer, correct_answer_text, points 
                FROM quiz_questions WHERE quiz_id = %s ORDER BY question_id ASC
            """, (quiz_id,))
            questions = cur.fetchall()
            
            quiz['questions'] = questions
            return jsonify({'status': 'success', 'data': quiz})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

# ==========================================
# 2. QUIZ ACTIONS (Publish / Delete)
# ==========================================
@auth.route('/quiz_action/<int:quiz_id>/<action>', methods=['POST'])
def quiz_action(quiz_id, action):
    conn = current_app.get_db_connection()
    try:
        with conn.cursor() as cur:
            if action == 'publish':
                # Set published = TRUE
                # Also set date_published = now() if it's currently NULL
                cur.execute("""
                    UPDATE quizzes
                    SET published = TRUE,
                        date_published = COALESCE(date_published, NOW())
                    WHERE quiz_id = %s
                """, (quiz_id,))
                msg = "Quiz Published!"
                
            elif action == 'unpublish':
                cur.execute("UPDATE quizzes SET published = FALSE WHERE quiz_id = %s", (quiz_id,))
                msg = "Quiz Unpublished."
                
            elif action == 'delete':
                cur.execute("UPDATE quizzes SET is_deleted = TRUE WHERE quiz_id = %s", (quiz_id,))
                msg = "Quiz Deleted."
                
            else:
                return jsonify({'status': 'error', 'message': 'Invalid action'}), 400
            
            conn.commit()
            return jsonify({'status': 'success', 'message': msg})
    
    except Exception as e:
        conn.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500
    
    finally:
        conn.close()

from flask import request, jsonify, session, current_app, render_template, flash, redirect, url_for
from flask_mail import Message
from threading import Thread
import psycopg2
import psycopg2.extras
# ==========================================
# 3. UPDATE SUBMIT API (Handle Create/Edit)
# ==========================================
@auth.route('/submit_quiz_api', methods=['POST'])
def submit_quiz_api():
    conn = current_app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        data = request.get_json(force=True)
        
        quiz_id = data.get('quiz_id') # None if creating
        course_id = int(data.get('course_id'))
        
        # Handle material_id safely
        raw_mat = data.get('material_id')
        material_id = int(raw_mat) if raw_mat and str(raw_mat) != '0' else None
        
        title = data.get('title')
        description = data.get('description', '')
        duration = int(data.get('duration', 0))
        questions = data.get('questions', [])

        # --- UPDATED: New Fields ---
        max_attempts = int(data.get('max_attempts', 1))
        publish_at = data.get('publish_at') # 'YYYY-MM-DD HH:MM'
        if publish_at == "": publish_at = None
        # ---------------------------

        if not title or not questions:
            return jsonify({'status':'error', 'message':'Missing title or questions.'}), 400

        # ==================================================
        # PREPARE DATA FOR NOTIFICATION (Teacher & Course)
        # ==================================================
        cur.execute("SELECT full_name FROM users WHERE user_id = %s", (session['user_id'],))
        teacher_res = cur.fetchone()
        teacher_name = teacher_res['full_name'] if teacher_res else "Instructor"

        cur.execute("SELECT course_code FROM courses WHERE course_id = %s", (course_id,))
        course_res = cur.fetchone()
        course_code = course_res['course_code'] if course_res else "Course"

        created_new_quiz = False # Flag to trigger notification

        # ==================================================
        # 1. HANDLE EDIT (UPDATE)
        # ==================================================
        if quiz_id:
            cur.execute("""
                UPDATE quizzes 
                SET title=%s, description=%s, duration_minutes=%s, material_id=%s,
                    max_attempts=%s, publish_at=%s
                WHERE quiz_id=%s
            """, (title, description, duration, material_id, max_attempts, publish_at, quiz_id))
            
            # Check if students have already taken this quiz.
            cur.execute("SELECT COUNT(*) as count FROM quiz_submissions WHERE quiz_id = %s", (quiz_id,))
            submission_count = cur.fetchone()['count']

            if submission_count > 0:
                conn.commit()
                return jsonify({
                    'status': 'success', 
                    'message': 'Quiz Details updated! (Questions were locked because students have already taken this quiz.)'
                })

            # If NO submissions, it is safe to delete and re-insert questions
            cur.execute("DELETE FROM quiz_questions WHERE quiz_id=%s", (quiz_id,))
        
        # ==================================================
        # 2. HANDLE CREATE (INSERT)
        # ==================================================
        else:
            cur.execute("""
                INSERT INTO quizzes (course_id, material_id, title, description, total_points, 
                                     published, date_created, is_deleted, duration_minutes,
                                     max_attempts, publish_at)
                VALUES (%s, %s, %s, %s, 0, FALSE, NOW(), FALSE, %s, %s, %s) 
                RETURNING quiz_id;
            """, (course_id, material_id, title, description, duration, max_attempts, publish_at))
            quiz_id = cur.fetchone()['quiz_id']
            created_new_quiz = True # Mark as new for notification

        # ==================================================
        # 3. INSERT QUESTIONS & ANSWERS
        # ==================================================
        total_points = 0
        for q in questions:
            q_type = q.get('type','').strip().lower().replace(' ','_')
            points = int(q.get('points', 1))
            total_points += points

            cur.execute("""
                INSERT INTO quiz_questions 
                (quiz_id, question_text, question_type, option_a, option_b, option_c, option_d, correct_answer, correct_answer_text, points)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                quiz_id, 
                q.get('text'), 
                q_type, 
                q.get('a'), q.get('b'), q.get('c'), q.get('d'),
                q.get('correct'),      
                q.get('correct_text'), 
                points
            ))

        # Update Total Points
        cur.execute("UPDATE quizzes SET total_points=%s WHERE quiz_id=%s;", (total_points, quiz_id))
        
        conn.commit()

        # ==================================================
        # 4. NOTIFICATION LOGIC (IF NEW QUIZ CREATED)
        # ==================================================
        if created_new_quiz:
            # A. Fetch Students
            cur.execute("""
                SELECT u.user_id, u.full_name, u.email
                FROM users u
                JOIN courses c ON u.section_id = c.section_id
                WHERE u.role_id = 3 AND c.course_id = %s;
            """, (course_id,))
            students = cur.fetchall()

            # B. Thread Function
            app = current_app._get_current_object()

            def notify_students_quiz():
                with app.app_context():
                    thread_conn = app.get_db_connection()
                    thread_cur = thread_conn.cursor()
                    
                    try:
                        link_url = url_for('auth.login', _external=True)
                    except:
                        link_url = "#"

                    try:
                        for student in students:
                            # 1. DB Notification
                            thread_cur.execute("""
                                INSERT INTO notifications (user_id, course_id, title, message, created_at, read)
                                VALUES (%s, %s, %s, %s, NOW(), FALSE)
                            """, (
                                student['user_id'],
                                course_id,
                                f"New Quiz: {title}",
                                f"{teacher_name} posted a new quiz in {course_code}."
                            ))

                            # 2. HTML Email (Purple Theme for Quizzes)
                            email_html = f"""
                            <!DOCTYPE html>
                            <html>
                            <head>
                                <style>
                                    body {{ font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif; background-color: #f4f6f8; margin: 0; padding: 0; }}
                                    .email-container {{ max-width: 600px; margin: 30px auto; background-color: #ffffff; border-radius: 8px; overflow: hidden; box-shadow: 0 4px 6px rgba(0,0,0,0.1); border: 1px solid #e0e0e0; }}
                                    .header {{ background-color: #8e44ad; color: #ffffff; padding: 25px; text-align: center; }} /* Purple for Quiz */
                                    .header h1 {{ margin: 0; font-size: 24px; font-weight: 600; }}
                                    .content {{ padding: 30px; color: #333333; line-height: 1.6; }}
                                    .info-box {{ background-color: #f8f9fa; border-left: 5px solid #8e44ad; padding: 15px 20px; margin: 20px 0; border-radius: 4px; }}
                                    .info-box h3 {{ margin-top: 0; color: #8e44ad; }}
                                    .btn-container {{ text-align: center; margin-top: 30px; margin-bottom: 20px; }}
                                    .btn {{ background-color: #27ae60; color: white !important; padding: 12px 25px; text-decoration: none; border-radius: 5px; font-weight: bold; display: inline-block; }}
                                    .footer {{ background-color: #eceff1; padding: 15px; text-align: center; font-size: 12px; color: #7f8c8d; }}
                                </style>
                            </head>
                            <body>
                                <div class="email-container">
                                    <div class="header">
                                        <h1>New Quiz Posted</h1>
                                    </div>
                                    <div class="content">
                                        <p>Hi <strong>{student['full_name']}</strong>,</p>
                                        <p>Your instructor, <strong>{teacher_name}</strong>, has added a new quiz to the course <strong>{course_code}</strong>.</p>
                                        
                                        <div class="info-box">
                                            <h3>{title}</h3>
                                            <p>{description if description else 'No description provided.'}</p>
                                            <p><strong>Duration:</strong> {duration} minutes</p>
                                        </div>

                                        <p>Please log in to the portal to check the availability and start the quiz.</p>
                                        
                                        <div class="btn-container">
                                            <a href="{link_url}" class="btn">Go to Quiz</a>
                                        </div>
                                    </div>
                                    <div class="footer">
                                        <p>This is an automated notification from your Learning Management System.</p>
                                        <p>&copy; 2024 LMS Team</p>
                                    </div>
                                </div>
                            </body>
                            </html>
                            """

                            # 3. Send Email
                            try:
                                msg = Message(
                                    subject=f"New Quiz: {title} ({course_code})",
                                    recipients=[student['email']],
                                    html=email_html
                                )
                                mail.send(msg)
                            except Exception as e:
                                print(f"Failed to send email to {student['email']}: {e}")

                        thread_conn.commit()
                    except Exception as e:
                        print(f"Error in quiz notification thread: {e}")
                        thread_conn.rollback()
                    finally:
                        thread_cur.close()
                        thread_conn.close()

            Thread(target=notify_students_quiz).start()
        
        action_word = "updated" if data.get('quiz_id') else "created"
        message = f'Quiz {action_word} successfully!'
        if created_new_quiz:
             message += " Notifications sent."

        return jsonify({'status':'success', 'message': message})

    except Exception as e:
        conn.rollback()
        print(f"Error in submit_quiz_api: {e}")
        return jsonify({'status':'error', 'message': str(e)}), 500
    finally:
        cur.close()
        conn.close()

@auth.route('/view_essay_answers/<int:quiz_id>', methods=['GET'])
def get_essay_answers(quiz_id):
    from . import create_app
    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        cur.execute("""
            SELECT qa.answer_id, qs.submission_id, qq.quiz_id,
                   u.full_name AS student_name, qq.question_text,
                   qa.answer_text, qa.score, qq.points, qs.submission_date,
                   qa.essay_grade_done
            FROM quiz_answers qa
            JOIN quiz_questions qq ON qa.question_id = qq.question_id
            JOIN quiz_submissions qs ON qa.submission_id = qs.submission_id
            JOIN users u ON qs.student_id = u.user_id
            WHERE qq.question_type='essay' AND qq.quiz_id=%s
            ORDER BY qs.submission_date DESC
        """, (quiz_id,))
        answers = cur.fetchall()

        return jsonify({"status": "success", "answers": answers})
    except Exception as e:
        print(e)
        return jsonify({"status": "error", "message": str(e)})
    finally:
        cur.close()
        conn.close()

from flask import request, jsonify, session, current_app, render_template, flash, redirect, url_for
from flask_mail import Message
from threading import Thread
import psycopg2
import psycopg2.extras


@auth.route('/grade_essay_answer', methods=['POST'])
def grade_essay_answer():
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Please log in first.'}), 401

    if session.get('role_id') != 2:  # teacher only
        return jsonify({'status': 'error', 'message': 'Access denied. Teachers only.'}), 403

    conn = current_app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    
    try:
        data = request.get_json(force=True)
        answer_id = int(data.get('answer_id', 0))
        submission_id = int(data.get('submission_id', 0))
        score_to_set = float(data.get('score', -1))

        if answer_id <= 0 or submission_id <= 0 or score_to_set < 0:
            return jsonify({'status': 'error', 'message': 'Invalid input values.'}), 400

        # 1. Update essay score and mark graded
        cur.execute("""
            UPDATE quiz_answers
            SET score = %s,
                essay_grade_done = TRUE
            WHERE answer_id = %s AND submission_id = %s
            RETURNING score, question_id
        """, (score_to_set, answer_id, submission_id))
        result = cur.fetchone()
        
        if not result:
            conn.rollback()
            return jsonify({'status': 'error', 'message': 'Answer not found.'}), 404

        updated_score = result['score']
        question_id = result['question_id']

        # 2. Recalculate total score for the submission
        cur.execute("""
            SELECT COALESCE(SUM(score),0) AS total_score, student_id
            FROM quiz_answers qa
            JOIN quiz_submissions qs ON qa.submission_id = qs.submission_id
            WHERE qa.submission_id = %s
            GROUP BY qs.student_id
        """, (submission_id,))
        total_result = cur.fetchone()
        
        if total_result:
            total_score = total_result['total_score']
            student_id = total_result['student_id']

            cur.execute("""
                UPDATE quiz_submissions
                SET total_score = %s
                WHERE submission_id = %s
            """, (total_score, submission_id))
        else:
            # Fallback if something weird happens
            conn.rollback()
            return jsonify({'status': 'error', 'message': 'Could not calculate total score.'}), 500

        conn.commit()

        # ===========================================
        # NOTIFICATION LOGIC
        # ===========================================

        # A. Fetch Student Info
        cur.execute("SELECT email, full_name FROM users WHERE user_id = %s", (student_id,))
        student = cur.fetchone()
        
        # B. Fetch Quiz, Question & Course Info
        cur.execute("""
            SELECT qq.question_text, q.title AS quiz_title, c.course_title, c.course_code, c.course_id
            FROM quiz_questions qq
            JOIN quizzes q ON qq.quiz_id = q.quiz_id
            JOIN courses c ON q.course_id = c.course_id
            WHERE qq.question_id = %s
        """, (question_id,))
        quiz_info = cur.fetchone()

        # C. Fetch Teacher Info (Current User)
        cur.execute("SELECT full_name FROM users WHERE user_id = %s", (session['user_id'],))
        teacher_res = cur.fetchone()
        teacher_name = teacher_res['full_name'] if teacher_res else "Instructor"

        if student and quiz_info:
            student_email = student['email']
            student_name = student['full_name']
            
            question_text = quiz_info['question_text']
            # Truncate question text if it's too long for the subject/preview
            short_q = (question_text[:50] + '...') if len(question_text) > 50 else question_text
            
            quiz_title = quiz_info['quiz_title']
            course_code = quiz_info['course_code']
            course_id = quiz_info['course_id']

            # D. Threaded Notification
            app = current_app._get_current_object()

            def notify_essay_grade():
                with app.app_context():
                    thread_conn = app.get_db_connection()
                    thread_cur = thread_conn.cursor()
                    
                    try:
                        # 1. Login Link
                        try:
                            link_url = url_for('auth.login', _external=True)
                        except:
                            link_url = "#"

                        # 2. Insert DB Notification
                        thread_cur.execute("""
                            INSERT INTO notifications (user_id, course_id, title, message, created_at, read)
                            VALUES (%s, %s, %s, %s, NOW(), FALSE)
                        """, (
                            student_id,
                            course_id,
                            f"Essay Graded: {quiz_title}",
                            f"Your essay for '{short_q}' in {quiz_title} has been graded. Score: {updated_score}"
                        ))

                        # 3. HTML Email Content
                        email_html = f"""
                        <!DOCTYPE html>
                        <html>
                        <head>
                            <style>
                                body {{ font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif; background-color: #f4f6f8; margin: 0; padding: 0; }}
                                .email-container {{ max-width: 600px; margin: 30px auto; background-color: #ffffff; border-radius: 8px; overflow: hidden; box-shadow: 0 4px 6px rgba(0,0,0,0.1); border: 1px solid #e0e0e0; }}
                                .header {{ background-color: #27ae60; color: #ffffff; padding: 25px; text-align: center; }}
                                .header h1 {{ margin: 0; font-size: 24px; font-weight: 600; }}
                                .content {{ padding: 30px; color: #333333; line-height: 1.6; }}
                                .info-box {{ background-color: #f8f9fa; border-left: 5px solid #27ae60; padding: 15px 20px; margin: 20px 0; border-radius: 4px; }}
                                .grade-badge {{ background-color: #27ae60; color: white; padding: 4px 10px; border-radius: 12px; font-weight: bold; font-size: 14px; margin-left: 5px; }}
                                .btn-container {{ text-align: center; margin-top: 30px; margin-bottom: 20px; }}
                                .btn {{ background-color: #3498db; color: white !important; padding: 12px 25px; text-decoration: none; border-radius: 5px; font-weight: bold; display: inline-block; }}
                                .footer {{ background-color: #eceff1; padding: 15px; text-align: center; font-size: 12px; color: #7f8c8d; }}
                            </style>
                        </head>
                        <body>
                            <div class="email-container">
                                <div class="header">
                                    <h1>Essay Graded</h1>
                                </div>
                                <div class="content">
                                    <p>Hi <strong>{student_name}</strong>,</p>
                                    <p>Your essay answer in the quiz <strong>{quiz_title}</strong> (Course: {course_code}) has been graded by {teacher_name}.</p>
                                    
                                    <div class="info-box">
                                        <p><strong>Question:</strong> {short_q}</p>
                                        <p style="margin-top: 10px; font-size: 16px;"><strong>Score Awarded:</strong> <span class="grade-badge">{updated_score}</span></p>
                                        <hr style="border: 0; border-top: 1px solid #eee; margin: 10px 0;">
                                        <p><strong>New Total Quiz Score:</strong> {total_score}</p>
                                    </div>

                                    <p>Check your LMS portal for full details.</p>
                                    
                                    <div class="btn-container">
                                        <a href="{link_url}" class="btn">View Result</a>
                                    </div>
                                </div>
                                <div class="footer">
                                    <p>This is an automated notification from your Learning Management System.</p>
                                    <p>&copy; 2024 LMS Team</p>
                                </div>
                            </div>
                        </body>
                        </html>
                        """

                        # 4. Send Email
                        try:
                            msg = Message(
                                subject=f"Essay Graded: {quiz_title} ({course_code})",
                                recipients=[student_email],
                                html=email_html
                            )
                            mail.send(msg)
                        except Exception as e:
                            print(f"Email failed for {student_email}: {e}")

                        thread_conn.commit()
                    except Exception as e:
                        print(f"Error in essay notification thread: {e}")
                        thread_conn.rollback()
                    finally:
                        thread_cur.close()
                        thread_conn.close()

            Thread(target=notify_essay_grade).start()

        return jsonify({
            'status': 'success',
            'message': 'Essay graded successfully! Notification sent.',
            'essay_score': updated_score,
            'total_score': total_score
        })

    except Exception as e:
        if conn:
            conn.rollback()
        print(e)
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        if cur: cur.close()
        if conn: conn.close()

from flask import request, jsonify, session, current_app, render_template, flash, redirect, url_for
from flask_mail import Message
from threading import Thread
import psycopg2
import psycopg2.extras

# ========================================================
# 1. API: MANAGE ASSIGNMENTS (Create/Edit/Delete/Publish)
# ========================================================
@auth.route('/manage_assignment_api', methods=['POST'])
def manage_assignment_api():
    conn = current_app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    
    try:
        data = request.get_json(force=True)
        action = data.get('action')

        # ---------------------------------------------------------
        # ACTION: CREATE or EDIT
        # ---------------------------------------------------------
        if action in ['create', 'edit']:
            assignment_id = data.get('assignment_id') 
            course_id = data.get('course_id')
            
            raw_material_id = data.get('material_id')
            material_id = None
            if raw_material_id and str(raw_material_id) != '0':
                try:
                    material_id = int(raw_material_id)
                except ValueError:
                    material_id = None

            title = data.get('title')
            description = data.get('description', '')
            due_date = data.get('due_date')
            points = int(data.get('points', 100))
            
            publish_at = data.get('publish_at') 
            if publish_at == "": publish_at = None

            if not title or not due_date:
                return jsonify({'status': 'error', 'message': 'Title and Due Date are required.'}), 400

            if action == 'edit' and assignment_id:
                cur.execute("""
                    UPDATE assignments 
                    SET title = %s, description = %s, due_date = %s, 
                        points = %s, material_id = %s, publish_at = %s
                    WHERE assignment_id = %s
                """, (title, description, due_date, points, material_id, publish_at, assignment_id))
                message = "Assignment updated successfully."
            
            else:
                cur.execute("""
                    INSERT INTO assignments 
                    (course_id, material_id, title, description, due_date, points, published, is_deleted, date_created, publish_at)
                    VALUES (%s, %s, %s, %s, %s, %s, FALSE, FALSE, NOW(), %s)
                    RETURNING assignment_id;
                """, (course_id, material_id, title, description, due_date, points, publish_at))
                message = "Assignment created successfully."

        # ---------------------------------------------------------
        # ACTION: PUBLISH / UNPUBLISH
        # ---------------------------------------------------------
        elif action == 'publish':
            aid = data.get('assignment_id')
            status_str = data.get('status') # 'publish' or 'unpublish'
            is_published = True if status_str == 'publish' else False
            
            # 1. Update DB Status
            cur.execute("""
                UPDATE assignments 
                SET published = %s, 
                    date_published = CASE WHEN %s THEN NOW() ELSE NULL END 
                WHERE assignment_id = %s
            """, (is_published, is_published, aid))
            
            # 2. NOTIFICATION LOGIC (Only if Publishing)
            if is_published:
                # A. Get Teacher Name
                cur.execute("SELECT full_name FROM users WHERE user_id = %s;", (session['user_id'],))
                teacher_user = cur.fetchone()
                teacher_name = teacher_user['full_name']

                # B. Get Assignment Details
                cur.execute("""
                    SELECT a.title, a.description, c.course_id, c.course_code
                    FROM assignments a
                    JOIN courses c ON a.course_id = c.course_id
                    WHERE a.assignment_id = %s;
                """, (aid,))
                assign_data = cur.fetchone()

                if assign_data:
                    course_id = assign_data['course_id']
                    assign_title = assign_data['title']
                    assign_desc = assign_data['description']
                    course_code = assign_data['course_code']

                    # C. Get Students
                    cur.execute("""
                        SELECT u.user_id, u.full_name, u.email
                        FROM users u
                        JOIN courses c ON u.section_id = c.section_id
                        WHERE u.role_id = 3 AND c.course_id = %s;
                    """, (course_id,))
                    students = cur.fetchall()

                    # D. Threaded Notification
                    app = current_app._get_current_object()

                    def notify_students():
                        with app.app_context():
                            thread_conn = app.get_db_connection()
                            thread_cur = thread_conn.cursor()
                            
                            try:
                                link_url = url_for('auth.login', _external=True)
                            except:
                                link_url = "#"

                            try:
                                for student in students:
                                    # DB Notification
                                    thread_cur.execute("""
                                        INSERT INTO notifications (user_id, course_id, title, message, created_at, read)
                                        VALUES (%s, %s, %s, %s, NOW(), FALSE)
                                    """, (
                                        student['user_id'],
                                        course_id,
                                        f"New Assignment: {assign_title}",
                                        f"{teacher_name} posted a new assignment in {course_code}."
                                    ))

                                    # HTML EMAIL CONTENT
                                    email_html = f"""
                                    <!DOCTYPE html>
                                    <html>
                                    <head>
                                        <style>
                                            body {{ font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif; background-color: #f4f6f8; margin: 0; padding: 0; }}
                                            .email-container {{ max-width: 600px; margin: 30px auto; background-color: #ffffff; border-radius: 8px; overflow: hidden; box-shadow: 0 4px 6px rgba(0,0,0,0.1); border: 1px solid #e0e0e0; }}
                                            .header {{ background-color: #2c3e50; color: #ffffff; padding: 25px; text-align: center; }}
                                            .header h1 {{ margin: 0; font-size: 24px; font-weight: 600; }}
                                            .content {{ padding: 30px; color: #333333; line-height: 1.6; }}
                                            .info-box {{ background-color: #f8f9fa; border-left: 5px solid #3498db; padding: 15px 20px; margin: 20px 0; border-radius: 4px; }}
                                            .info-box h3 {{ margin-top: 0; color: #2c3e50; }}
                                            .btn-container {{ text-align: center; margin-top: 30px; margin-bottom: 20px; }}
                                            .btn {{ background-color: #27ae60; color: white !important; padding: 12px 25px; text-decoration: none; border-radius: 5px; font-weight: bold; display: inline-block; }}
                                            .footer {{ background-color: #eceff1; padding: 15px; text-align: center; font-size: 12px; color: #7f8c8d; }}
                                        </style>
                                    </head>
                                    <body>
                                        <div class="email-container">
                                            <div class="header">
                                                <h1>New Assignment Posted</h1>
                                            </div>
                                            <div class="content">
                                                <p>Hi <strong>{student['full_name']}</strong>,</p>
                                                <p>Your instructor, <strong>{teacher_name}</strong>, has just published a new assignment for the course <strong>{course_code}</strong>.</p>
                                                
                                                <div class="info-box">
                                                    <h3>{assign_title}</h3>
                                                    <p>{assign_desc}</p>
                                                </div>

                                                <p>Please log in to the portal to view the due date, download materials, and submit your work.</p>
                                                
                                                <div class="btn-container">
                                                    <a href="{link_url}" class="btn">View Assignment</a>
                                                </div>
                                            </div>
                                            <div class="footer">
                                                <p>This is an automated notification from your Learning Management System.</p>
                                                <p>&copy; 2024 LMS Team</p>
                                            </div>
                                        </div>
                                    </body>
                                    </html>
                                    """

                                    # Send Email
                                    try:
                                        msg = Message(
                                            subject=f"New Assignment: {assign_title} ({course_code})",
                                            recipients=[student['email']],
                                            html=email_html
                                        )
                                        mail.send(msg)
                                    except Exception as e:
                                        print(f"Failed to send email to {student['email']}: {e}")

                                thread_conn.commit()
                            except Exception as e:
                                print(f"Error in notification thread: {e}")
                                thread_conn.rollback()
                            finally:
                                thread_cur.close()
                                thread_conn.close()

                    Thread(target=notify_students).start()

            message = f"Assignment {status_str}ed."
            if is_published:
                message += " Notifications sent."

        # ---------------------------------------------------------
        # ACTION: DELETE
        # ---------------------------------------------------------
        elif action == 'delete':
            aid = data.get('assignment_id')
            cur.execute("UPDATE assignments SET is_deleted = TRUE WHERE assignment_id = %s", (aid,))
            message = "Assignment deleted."

        # ---------------------------------------------------------
        # ACTION: GRADE SUBMISSION
        # ---------------------------------------------------------
        elif action == 'grade':
            sub_id = data.get('submission_id')
            grade = data.get('grade')
            feedback = data.get('feedback', '')
            
            # 1. Update the Grade in DB
            cur.execute("""
                UPDATE submissions 
                SET grade = %s, feedback = %s, status = 'Graded' 
                WHERE submission_id = %s
            """, (grade, feedback, sub_id))
            
            # 2. Fetch Submission & Student Details
            # UPDATED: Changed s.user_id to s.student_id (or inferred via join)
            cur.execute("""
                SELECT u.user_id, u.full_name, u.email, 
                       a.title AS assignment_title, c.course_code, c.course_id
                FROM submissions s
                JOIN users u ON s.student_id = u.user_id  -- FIXED: Changed from s.user_id
                JOIN assignments a ON s.assignment_id = a.assignment_id
                JOIN courses c ON a.course_id = c.course_id
                WHERE s.submission_id = %s
            """, (sub_id,))
            sub_info = cur.fetchone()

            # Get Teacher Name
            cur.execute("SELECT full_name FROM users WHERE user_id = %s", (session.get('user_id'),))
            teacher_res = cur.fetchone()
            teacher_name = teacher_res['full_name'] if teacher_res else "Instructor"

            # 3. Threaded Email/Notification
            if sub_info:
                app = current_app._get_current_object()
                
                def notify_grade():
                    with app.app_context():
                        thread_conn = app.get_db_connection()
                        thread_cur = thread_conn.cursor()

                        # We selected u.user_id, so we use 'user_id' from the result
                        student_id = sub_info['user_id']
                        student_name = sub_info['full_name']
                        student_email = sub_info['email']
                        assign_title = sub_info['assignment_title']
                        course_code = sub_info['course_code']
                        course_id = sub_info['course_id']
                        
                        try:
                            # Generate Login Link
                            try:
                                link_url = url_for('auth.login', _external=True)
                            except:
                                link_url = "#"

                            # Insert DB Notification
                            thread_cur.execute("""
                                INSERT INTO notifications (user_id, course_id, title, message, created_at, read)
                                VALUES (%s, %s, %s, %s, NOW(), FALSE)
                            """, (
                                student_id,
                                course_id,
                                f"Grade Posted: {assign_title}",
                                f"Your assignment '{assign_title}' has been graded. Grade: {grade}"
                            ))

                            # HTML Email Template for Grading
                            email_html = f"""
                            <!DOCTYPE html>
                            <html>
                            <head>
                                <style>
                                    body {{ font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif; background-color: #f4f6f8; margin: 0; padding: 0; }}
                                    .email-container {{ max-width: 600px; margin: 30px auto; background-color: #ffffff; border-radius: 8px; overflow: hidden; box-shadow: 0 4px 6px rgba(0,0,0,0.1); border: 1px solid #e0e0e0; }}
                                    .header {{ background-color: #27ae60; color: #ffffff; padding: 25px; text-align: center; }}
                                    .header h1 {{ margin: 0; font-size: 24px; font-weight: 600; }}
                                    .content {{ padding: 30px; color: #333333; line-height: 1.6; }}
                                    .info-box {{ background-color: #f8f9fa; border-left: 5px solid #27ae60; padding: 15px 20px; margin: 20px 0; border-radius: 4px; }}
                                    .grade-badge {{ background-color: #27ae60; color: white; padding: 4px 10px; border-radius: 12px; font-weight: bold; font-size: 14px; margin-left: 5px; }}
                                    .btn-container {{ text-align: center; margin-top: 30px; margin-bottom: 20px; }}
                                    .btn {{ background-color: #3498db; color: white !important; padding: 12px 25px; text-decoration: none; border-radius: 5px; font-weight: bold; display: inline-block; }}
                                    .footer {{ background-color: #eceff1; padding: 15px; text-align: center; font-size: 12px; color: #7f8c8d; }}
                                </style>
                            </head>
                            <body>
                                <div class="email-container">
                                    <div class="header">
                                        <h1>Assignment Graded</h1>
                                    </div>
                                    <div class="content">
                                        <p>Hi <strong>{student_name}</strong>,</p>
                                        <p>Your submission for the assignment <strong>{assign_title}</strong> in course <strong>{course_code}</strong> has been graded by {teacher_name}.</p>
                                        
                                        <div class="info-box">
                                            <p style="margin: 0; font-size: 16px;"><strong>Grade Received:</strong> <span class="grade-badge">{grade}</span></p>
                                            <p style="margin-top: 10px;"><strong>Feedback:</strong><br><em>{feedback if feedback else 'No feedback provided.'}</em></p>
                                        </div>

                                        <p>You can log in to your portal to view more details.</p>
                                        
                                        <div class="btn-container">
                                            <a href="{link_url}" class="btn">View Grade</a>
                                        </div>
                                    </div>
                                    <div class="footer">
                                        <p>This is an automated notification from your Learning Management System.</p>
                                        <p>&copy; 2024 LMS Team</p>
                                    </div>
                                </div>
                            </body>
                            </html>
                            """

                            # Send Email
                            msg = Message(
                                subject=f"Grade Posted: {assign_title} ({course_code})",
                                recipients=[student_email],
                                html=email_html
                            )
                            mail.send(msg)

                            thread_conn.commit()
                        except Exception as e:
                            print(f"Error in grade notification thread: {e}")
                            thread_conn.rollback()
                        finally:
                            thread_cur.close()
                            thread_conn.close()

                Thread(target=notify_grade).start()

            message = "Grade saved and student notified."

        else:
            return jsonify({'status': 'error', 'message': 'Invalid action'}), 400

        conn.commit()
        return jsonify({'status': 'success', 'message': message})

    except Exception as e:
        conn.rollback()
        print(f"Assignment API Error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        cur.close()
        conn.close()
# ========================================================
# 2. API: GET DETAILS (For the Edit Modal)
# ========================================================
@auth.route('/get_assignment_details/<int:assignment_id>', methods=['GET'])
def get_assignment_details(assignment_id):
    conn = current_app.get_db_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            # UPDATED: Fetch publish_at
            cur.execute("""
                SELECT assignment_id, title, description, 
                       TO_CHAR(due_date, 'YYYY-MM-DD"T"HH24:MI') as due_date, 
                       points, material_id, course_id,
                       TO_CHAR(publish_at, 'YYYY-MM-DD"T"HH24:MI') as publish_at
                FROM assignments 
                WHERE assignment_id = %s
            """, (assignment_id,))
            data = cur.fetchone()
            
            if data:
                return jsonify({'status': 'success', 'data': data})
            return jsonify({'status': 'error', 'message': 'Assignment not found'}), 404
    finally:
        conn.close()

@auth.route('/view_assignment_submissions/<int:assignment_id>', methods=['GET'])
def view_assignment_submissions(assignment_id):
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401

    conn = current_app.get_db_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT s.submission_id, 
                       u.full_name as student_name, 
                       s.submission_date, 
                       s.file_path, 
                       s.grade, 
                       COALESCE(s.feedback, '') as feedback, 
                       COALESCE(s.remarks, '') as remarks
                FROM submissions s
                JOIN users u ON s.student_id = u.user_id
                WHERE s.assignment_id = %s
                ORDER BY s.submission_date DESC
            """, (assignment_id,))
            subs = cur.fetchall()
            
            for s in subs:
                if s['submission_date']:
                    s['submission_date'] = s['submission_date'].strftime('%Y-%m-%d %H:%M')
                else:
                    s['submission_date'] = "N/A"
            
            return jsonify({'status': 'success', 'submissions': subs})

    except Exception as e:
        print(f"Error fetching submissions: {e}") 
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

from flask import request, jsonify, session, current_app, render_template, flash, redirect, url_for
from flask_mail import Message
from threading import Thread
import psycopg2
import psycopg2.extras
 

# ========================================================
# 1. API: MANAGE ACTIVITIES (Create/Edit/Delete/Publish/Grade)
# ========================================================
@auth.route('/manage_activity_api', methods=['POST'])
def manage_activity_api():
    conn = current_app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        data = request.get_json(force=True)
        action = data.get('action')

        # ---------------------------------------------------------
        # ACTION: CREATE or EDIT
        # ---------------------------------------------------------
        if action in ['create', 'edit']:
            act_id = data.get('activity_id')
            course_id = data.get('course_id')
            mat_id = data.get('material_id')
            material_id = int(mat_id) if mat_id and str(mat_id) != '0' else None
            
            title = data.get('title')
            desc = data.get('description')
            due = data.get('due_date')
            points = data.get('points', 100)
            
            publish_at = data.get('publish_at')
            if publish_at == "": publish_at = None

            if action == 'edit' and act_id:
                cur.execute("""
                    UPDATE learning_activities 
                    SET title=%s, description=%s, due_date=%s, points=%s, publish_at=%s, material_id=%s
                    WHERE activity_id=%s
                """, (title, desc, due, points, publish_at, material_id, act_id))
                msg = "Activity updated."
            else:
                cur.execute("""
                    INSERT INTO learning_activities (course_id, material_id, title, description, due_date, points, published, created_at, publish_at)
                    VALUES (%s, %s, %s, %s, %s, %s, FALSE, NOW(), %s)
                """, (course_id, material_id, title, desc, due, points, publish_at))
                msg = "Activity created."

        # ---------------------------------------------------------
        # ACTION: PUBLISH / UNPUBLISH
        # ---------------------------------------------------------
        elif action in ['publish', 'unpublish']:
            aid = data.get('activity_id')
            status = True if action == 'publish' else False
            
            # 1. Update DB Status
            cur.execute("UPDATE learning_activities SET published=%s WHERE activity_id=%s", (status, aid))
            
            # 2. NOTIFICATION LOGIC (Only if Publishing)
            if status: # if action == 'publish'
                # A. Get Teacher Name
                cur.execute("SELECT full_name FROM users WHERE user_id = %s;", (session['user_id'],))
                teacher_user = cur.fetchone()
                teacher_name = teacher_user['full_name']

                # B. Get Activity Details
                cur.execute("""
                    SELECT a.title, a.description, c.course_id, c.course_code
                    FROM learning_activities a
                    JOIN courses c ON a.course_id = c.course_id
                    WHERE a.activity_id = %s;
                """, (aid,))
                act_data = cur.fetchone()

                if act_data:
                    course_id = act_data['course_id']
                    act_title = act_data['title']
                    act_desc = act_data['description']
                    course_code = act_data['course_code']

                    # C. Get Students
                    cur.execute("""
                        SELECT u.user_id, u.full_name, u.email
                        FROM users u
                        JOIN courses c ON u.section_id = c.section_id
                        WHERE u.role_id = 3 AND c.course_id = %s;
                    """, (course_id,))
                    students = cur.fetchall()

                    # D. Threaded Notification
                    app = current_app._get_current_object()

                    def notify_students():
                        with app.app_context():
                            thread_conn = app.get_db_connection()
                            thread_cur = thread_conn.cursor()
                            
                            try:
                                link_url = url_for('auth.login', _external=True)
                            except:
                                link_url = "#"

                            try:
                                for student in students:
                                    # DB Notification
                                    thread_cur.execute("""
                                        INSERT INTO notifications (user_id, course_id, title, message, created_at, read)
                                        VALUES (%s, %s, %s, %s, NOW(), FALSE)
                                    """, (
                                        student['user_id'],
                                        course_id,
                                        f"New Activity: {act_title}",
                                        f"{teacher_name} posted a new activity in {course_code}."
                                    ))

                                    # HTML EMAIL CONTENT
                                    email_html = f"""
                                    <!DOCTYPE html>
                                    <html>
                                    <head>
                                        <style>
                                            body {{ font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif; background-color: #f4f6f8; margin: 0; padding: 0; }}
                                            .email-container {{ max-width: 600px; margin: 30px auto; background-color: #ffffff; border-radius: 8px; overflow: hidden; box-shadow: 0 4px 6px rgba(0,0,0,0.1); border: 1px solid #e0e0e0; }}
                                            .header {{ background-color: #e67e22; color: #ffffff; padding: 25px; text-align: center; }} /* Orange for Activity */
                                            .header h1 {{ margin: 0; font-size: 24px; font-weight: 600; }}
                                            .content {{ padding: 30px; color: #333333; line-height: 1.6; }}
                                            .info-box {{ background-color: #f8f9fa; border-left: 5px solid #e67e22; padding: 15px 20px; margin: 20px 0; border-radius: 4px; }}
                                            .info-box h3 {{ margin-top: 0; color: #d35400; }}
                                            .btn-container {{ text-align: center; margin-top: 30px; margin-bottom: 20px; }}
                                            .btn {{ background-color: #27ae60; color: white !important; padding: 12px 25px; text-decoration: none; border-radius: 5px; font-weight: bold; display: inline-block; }}
                                            .footer {{ background-color: #eceff1; padding: 15px; text-align: center; font-size: 12px; color: #7f8c8d; }}
                                        </style>
                                    </head>
                                    <body>
                                        <div class="email-container">
                                            <div class="header">
                                                <h1>New Activity Posted</h1>
                                            </div>
                                            <div class="content">
                                                <p>Hi <strong>{student['full_name']}</strong>,</p>
                                                <p>Your instructor, <strong>{teacher_name}</strong>, has just published a new activity for the course <strong>{course_code}</strong>.</p>
                                                
                                                <div class="info-box">
                                                    <h3>{act_title}</h3>
                                                    <p>{act_desc}</p>
                                                </div>

                                                <p>Please log in to the portal to view details and submit your work.</p>
                                                
                                                <div class="btn-container">
                                                    <a href="{link_url}" class="btn">View Activity</a>
                                                </div>
                                            </div>
                                            <div class="footer">
                                                <p>This is an automated notification from your Learning Management System.</p>
                                                <p>&copy; 2024 LMS Team</p>
                                            </div>
                                        </div>
                                    </body>
                                    </html>
                                    """

                                    # Send Email
                                    try:
                                        msg = Message(
                                            subject=f"New Activity: {act_title} ({course_code})",
                                            recipients=[student['email']],
                                            html=email_html
                                        )
                                        mail.send(msg)
                                    except Exception as e:
                                        print(f"Failed to send email to {student['email']}: {e}")

                                thread_conn.commit()
                            except Exception as e:
                                print(f"Error in notification thread: {e}")
                                thread_conn.rollback()
                            finally:
                                thread_cur.close()
                                thread_conn.close()

                    Thread(target=notify_students).start()

            msg = f"Activity {action}ed."
            if status:
                msg += " Notifications sent."

        # ---------------------------------------------------------
        # ACTION: DELETE
        # ---------------------------------------------------------
        elif action == 'delete':
            aid = data.get('activity_id')
            cur.execute("DELETE FROM learning_activities WHERE activity_id=%s", (aid,))
            msg = "Activity deleted."

        # ---------------------------------------------------------
        # ACTION: GRADE SUBMISSION
        # ---------------------------------------------------------
        elif action == 'grade':
            sub_id = data.get('submission_id')
            grade = data.get('grade')
            feedback = data.get('feedback')
            
            # 1. Update Grade
            cur.execute("""
                UPDATE learning_activity_submissions 
                SET grade=%s, feedback=%s, updated_at=NOW() 
                WHERE submission_id=%s
            """, (grade, feedback, sub_id))

            # 2. Fetch Submission & Student Details
            # Joining users on student_id (assuming standard naming convention for submission tables)
            cur.execute("""
                SELECT u.user_id, u.full_name, u.email, 
                       a.title AS activity_title, c.course_code, c.course_id
                FROM learning_activity_submissions s
                JOIN users u ON s.student_id = u.user_id
                JOIN learning_activities a ON s.activity_id = a.activity_id
                JOIN courses c ON a.course_id = c.course_id
                WHERE s.submission_id = %s
            """, (sub_id,))
            sub_info = cur.fetchone()

            # Get Teacher Name
            cur.execute("SELECT full_name FROM users WHERE user_id = %s", (session.get('user_id'),))
            teacher_res = cur.fetchone()
            teacher_name = teacher_res['full_name'] if teacher_res else "Instructor"

            # 3. Threaded Email/Notification
            if sub_info:
                app = current_app._get_current_object()
                
                def notify_grade():
                    with app.app_context():
                        thread_conn = app.get_db_connection()
                        thread_cur = thread_conn.cursor()

                        student_id = sub_info['user_id']
                        student_name = sub_info['full_name']
                        student_email = sub_info['email']
                        act_title = sub_info['activity_title']
                        course_code = sub_info['course_code']
                        course_id = sub_info['course_id']
                        
                        try:
                            try:
                                link_url = url_for('auth.login', _external=True)
                            except:
                                link_url = "#"

                            # Insert DB Notification
                            thread_cur.execute("""
                                INSERT INTO notifications (user_id, course_id, title, message, created_at, read)
                                VALUES (%s, %s, %s, %s, NOW(), FALSE)
                            """, (
                                student_id,
                                course_id,
                                f"Grade Posted: {act_title}",
                                f"Your activity '{act_title}' has been graded. Grade: {grade}"
                            ))

                            # HTML Email Template for Grading
                            email_html = f"""
                            <!DOCTYPE html>
                            <html>
                            <head>
                                <style>
                                    body {{ font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif; background-color: #f4f6f8; margin: 0; padding: 0; }}
                                    .email-container {{ max-width: 600px; margin: 30px auto; background-color: #ffffff; border-radius: 8px; overflow: hidden; box-shadow: 0 4px 6px rgba(0,0,0,0.1); border: 1px solid #e0e0e0; }}
                                    .header {{ background-color: #27ae60; color: #ffffff; padding: 25px; text-align: center; }}
                                    .header h1 {{ margin: 0; font-size: 24px; font-weight: 600; }}
                                    .content {{ padding: 30px; color: #333333; line-height: 1.6; }}
                                    .info-box {{ background-color: #f8f9fa; border-left: 5px solid #27ae60; padding: 15px 20px; margin: 20px 0; border-radius: 4px; }}
                                    .grade-badge {{ background-color: #27ae60; color: white; padding: 4px 10px; border-radius: 12px; font-weight: bold; font-size: 14px; margin-left: 5px; }}
                                    .btn-container {{ text-align: center; margin-top: 30px; margin-bottom: 20px; }}
                                    .btn {{ background-color: #3498db; color: white !important; padding: 12px 25px; text-decoration: none; border-radius: 5px; font-weight: bold; display: inline-block; }}
                                    .footer {{ background-color: #eceff1; padding: 15px; text-align: center; font-size: 12px; color: #7f8c8d; }}
                                </style>
                            </head>
                            <body>
                                <div class="email-container">
                                    <div class="header">
                                        <h1>Activity Graded</h1>
                                    </div>
                                    <div class="content">
                                        <p>Hi <strong>{student_name}</strong>,</p>
                                        <p>Your submission for the activity <strong>{act_title}</strong> in course <strong>{course_code}</strong> has been graded by {teacher_name}.</p>
                                        
                                        <div class="info-box">
                                            <p style="margin: 0; font-size: 16px;"><strong>Grade Received:</strong> <span class="grade-badge">{grade}</span></p>
                                            <p style="margin-top: 10px;"><strong>Feedback:</strong><br><em>{feedback if feedback else 'No feedback provided.'}</em></p>
                                        </div>

                                        <p>You can log in to your portal to view more details.</p>
                                        
                                        <div class="btn-container">
                                            <a href="{link_url}" class="btn">View Grade</a>
                                        </div>
                                    </div>
                                    <div class="footer">
                                        <p>This is an automated notification from your Learning Management System.</p>
                                        <p>&copy; 2024 LMS Team</p>
                                    </div>
                                </div>
                            </body>
                            </html>
                            """

                            # Send Email
                            msg = Message(
                                subject=f"Grade Posted: {act_title} ({course_code})",
                                recipients=[student_email],
                                html=email_html
                            )
                            mail.send(msg)

                            thread_conn.commit()
                        except Exception as e:
                            print(f"Error in grade notification thread: {e}")
                            thread_conn.rollback()
                        finally:
                            thread_cur.close()
                            thread_conn.close()

                Thread(target=notify_grade).start()

            msg = "Grade saved and student notified."

        conn.commit()
        return jsonify({'status': 'success', 'message': msg})
    except Exception as e:
        conn.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        cur.close()
        conn.close()

# ========================================================
# 2. API: GET ACTIVITY DETAILS (For Edit)
# ========================================================
@auth.route('/get_activity_details/<int:activity_id>', methods=['GET'])
def get_activity_details(activity_id):
    conn = current_app.get_db_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            # UPDATED: Fetch publish_at
            cur.execute("""
                SELECT activity_id, title, description, 
                       TO_CHAR(due_date, 'YYYY-MM-DD"T"HH24:MI') as due_date, 
                       points,
                       TO_CHAR(publish_at, 'YYYY-MM-DD"T"HH24:MI') as publish_at
                FROM learning_activities WHERE activity_id = %s
            """, (activity_id,))
            data = cur.fetchone()
            return jsonify({'status': 'success', 'data': data})
    finally:
        conn.close()

@auth.route('/view_activity_submissions/<int:activity_id>', methods=['GET'])
def view_activity_submissions(activity_id):
    conn = current_app.get_db_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT s.submission_id, 
                       u.full_name as student_name, 
                       s.submission_date, 
                       s.file_path, 
                       s.grade, 
                       s.feedback
                FROM learning_activity_submissions s
                JOIN users u ON s.student_id = u.user_id
                WHERE s.activity_id = %s
                ORDER BY s.submission_date DESC
            """, (activity_id,))
            subs = cur.fetchall()
            
            for s in subs:
                if s['submission_date']:
                    s['submission_date'] = s['submission_date'].strftime('%Y-%m-%d %H:%M')
            
            return jsonify({'status': 'success', 'submissions': subs})
    except Exception as e:
        print(f"Error fetching activity submissions: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

from flask import request, jsonify, session, current_app, render_template, flash, redirect, url_for
from flask_mail import Message
from threading import Thread
import psycopg2
import psycopg2.extras
from psycopg2 import sql


# ========================================================
# NEW API: MANAGE STUDENT EXCEPTIONS / EXTENSIONS
# ========================================================
@auth.route('/manage_exception_api', methods=['POST'])
def manage_exception_api():
    # Only Teachers allowed
    if 'user_id' not in session or session.get('role_id') != 2:
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 403

    conn = current_app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        data = request.get_json(force=True)

        action = data.get('action')                 # set_exception / remove_exception
        item_type = data.get('item_type')           # quiz / assignment / activity
        item_id = int(data.get('item_id'))
        student_id_raw = data.get('student_id')     # <-- student_id from frontend

        # 1️⃣ Validate student_id exists in users and FETCH INFO
        cur.execute("""
            SELECT user_id, student_id, full_name, email 
            FROM users 
            WHERE student_id = %s
        """, (student_id_raw,))
        
        student_row = cur.fetchone()

        if not student_row:
            return jsonify({'status': 'error', 'message': 'Invalid Student ID'}), 400

        user_id = student_row["user_id"]  # Real user ID used in exceptions tables
        student_name = student_row["full_name"]
        student_email = student_row["email"]

        # Fields for setting exception
        extended_date = data.get('extended_date')   # YYYY-MM-DD HH:MM
        extra_attempts = int(data.get('extra_attempts', 0))

        if not extended_date:
            extended_date = None

        # Determine Table & Fetch Item Details
        table = ''
        id_col = ''
        item_title = "Unknown Item"
        course_code = "Course"
        course_id = 0

        if item_type == 'quiz':
            table = 'quiz_exceptions'
            id_col = 'quiz_id'
            cur.execute("""
                SELECT q.title, c.course_code, c.course_id 
                FROM quizzes q JOIN courses c ON q.course_id = c.course_id 
                WHERE q.quiz_id = %s
            """, (item_id,))
        elif item_type == 'assignment':
            table = 'assignment_exceptions'
            id_col = 'assignment_id'
            cur.execute("""
                SELECT a.title, c.course_code, c.course_id 
                FROM assignments a JOIN courses c ON a.course_id = c.course_id 
                WHERE a.assignment_id = %s
            """, (item_id,))
        elif item_type == 'activity':
            table = 'activity_exceptions'
            id_col = 'activity_id'
            cur.execute("""
                SELECT a.title, c.course_code, c.course_id 
                FROM learning_activities a JOIN courses c ON a.course_id = c.course_id 
                WHERE a.activity_id = %s
            """, (item_id,))
        else:
            return jsonify({'status': 'error', 'message': 'Invalid type'}), 400

        # Store fetched item info
        item_res = cur.fetchone()
        if item_res:
            item_title = item_res['title']
            course_code = item_res['course_code']
            course_id = item_res['course_id']

        # Get Teacher Name
        cur.execute("SELECT full_name FROM users WHERE user_id = %s", (session['user_id'],))
        teacher_res = cur.fetchone()
        teacher_name = teacher_res['full_name'] if teacher_res else "Instructor"

        # =========================================================
        # EXECUTE DB ACTION
        # =========================================================
        msg = ""

        # REMOVE EXCEPTION
        if action == 'remove_exception':
            query = sql.SQL("""
                DELETE FROM {} 
                WHERE {} = %s AND student_id = %s
            """).format(sql.Identifier(table), sql.Identifier(id_col))

            cur.execute(query, (item_id, user_id))
            msg = "Exception removed."

        # SET EXCEPTION
        elif action == 'set_exception':
            # QUIZ (with extra_attempts)
            if item_type == 'quiz':
                query = sql.SQL("""
                    INSERT INTO {} ({}, student_id, extended_due_date, extra_attempts)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT ({}, student_id)
                    DO UPDATE SET extended_due_date = EXCLUDED.extended_due_date,
                                  extra_attempts = EXCLUDED.extra_attempts
                """).format(sql.Identifier(table), sql.Identifier(id_col), sql.Identifier(id_col))

                cur.execute(query, (item_id, user_id, extended_date, extra_attempts))

            # ASSIGNMENT / ACTIVITY
            else:
                query = sql.SQL("""
                    INSERT INTO {} ({}, student_id, extended_due_date)
                    VALUES (%s, %s, %s)
                    ON CONFLICT ({}, student_id)
                    DO UPDATE SET extended_due_date = EXCLUDED.extended_due_date
                """).format(sql.Identifier(table), sql.Identifier(id_col), sql.Identifier(id_col))

                cur.execute(query, (item_id, user_id, extended_date))

            msg = "Exception granted successfully."

        else:
            return jsonify({'status': 'error', 'message': 'Invalid Action'}), 400

        conn.commit()

        # =========================================================
        # NOTIFICATION LOGIC (Threaded)
        # =========================================================
        app = current_app._get_current_object()

        def notify_exception():
            with app.app_context():
                thread_conn = app.get_db_connection()
                thread_cur = thread_conn.cursor()
                
                try:
                    # Prepare content based on action
                    if action == 'set_exception':
                        notif_title = f"Extension Granted: {item_title}"
                        notif_msg = f"{teacher_name} has granted you an exception/extension for {item_title}."
                        
                        email_subject = f"Update: Exception Granted for {item_title}"
                        email_header_text = "Exception Granted"
                        email_intro = f"Your instructor, <strong>{teacher_name}</strong>, has granted you an exception for <strong>{item_title}</strong> in <strong>{course_code}</strong>."
                        
                        details_html = f"""
                            <p><strong>New Due Date:</strong> {extended_date if extended_date else 'No specific date set (Open)'}</p>
                        """
                        if item_type == 'quiz' and extra_attempts > 0:
                            details_html += f"<p><strong>Extra Attempts Allowed:</strong> {extra_attempts}</p>"
                    
                    else: # remove_exception
                        notif_title = f"Extension Removed: {item_title}"
                        notif_msg = f"The exception for {item_title} has been removed."
                        
                        email_subject = f"Update: Exception Removed for {item_title}"
                        email_header_text = "Exception Removed"
                        email_intro = f"The exception/extension previously granted for <strong>{item_title}</strong> in <strong>{course_code}</strong> has been removed."
                        details_html = "<p>The settings for this item have reverted to the standard course schedule.</p>"

                    try:
                        link_url = url_for('auth.login', _external=True)
                    except:
                        link_url = "#"

                    # 1. Insert DB Notification
                    thread_cur.execute("""
                        INSERT INTO notifications (user_id, course_id, title, message, created_at, read)
                        VALUES (%s, %s, %s, %s, NOW(), FALSE)
                    """, (user_id, course_id, notif_title, notif_msg))

                    # 2. HTML Email Content (Teal/Info Theme)
                    email_html = f"""
                    <!DOCTYPE html>
                    <html>
                    <head>
                        <style>
                            body {{ font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif; background-color: #f4f6f8; margin: 0; padding: 0; }}
                            .email-container {{ max-width: 600px; margin: 30px auto; background-color: #ffffff; border-radius: 8px; overflow: hidden; box-shadow: 0 4px 6px rgba(0,0,0,0.1); border: 1px solid #e0e0e0; }}
                            .header {{ background-color: #17a2b8; color: #ffffff; padding: 25px; text-align: center; }} /* Teal for Info/Exceptions */
                            .header h1 {{ margin: 0; font-size: 24px; font-weight: 600; }}
                            .content {{ padding: 30px; color: #333333; line-height: 1.6; }}
                            .info-box {{ background-color: #e0f7fa; border-left: 5px solid #17a2b8; padding: 15px 20px; margin: 20px 0; border-radius: 4px; }}
                            .btn-container {{ text-align: center; margin-top: 30px; margin-bottom: 20px; }}
                            .btn {{ background-color: #3498db; color: white !important; padding: 12px 25px; text-decoration: none; border-radius: 5px; font-weight: bold; display: inline-block; }}
                            .footer {{ background-color: #eceff1; padding: 15px; text-align: center; font-size: 12px; color: #7f8c8d; }}
                        </style>
                    </head>
                    <body>
                        <div class="email-container">
                            <div class="header">
                                <h1>{email_header_text}</h1>
                            </div>
                            <div class="content">
                                <p>Hi <strong>{student_name}</strong>,</p>
                                <p>{email_intro}</p>
                                
                                <div class="info-box">
                                    {details_html}
                                </div>

                                <p>Please log in to your portal to check the updated details.</p>
                                
                                <div class="btn-container">
                                    <a href="{link_url}" class="btn">View Course</a>
                                </div>
                            </div>
                            <div class="footer">
                                <p>This is an automated notification from your Learning Management System.</p>
                                <p>&copy; 2024 LMS Team</p>
                            </div>
                        </div>
                    </body>
                    </html>
                    """

                    # 3. Send Email
                    msg = Message(
                        subject=email_subject,
                        recipients=[student_email],
                        html=email_html
                    )
                    mail.send(msg)

                    thread_conn.commit()
                except Exception as e:
                    print(f"Error in exception notification thread: {e}")
                    thread_conn.rollback()
                finally:
                    thread_cur.close()
                    thread_conn.close()

        Thread(target=notify_exception).start()

        return jsonify({'status': 'success', 'message': msg})

    except Exception as e:
        conn.rollback()
        print("Exception API Error:", e)
        return jsonify({'status': 'error', 'message': str(e)}), 500

    finally:
        cur.close()
        conn.close()
# ========================================================
# SEARCH STUDENT BY student_id
# ========================================================
@auth.route('/api/get_student_by_pid')
def get_student_by_pid():
    student_id = request.args.get('student_id', '').strip()

    conn = current_app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        cur.execute("""
            SELECT user_id, student_id, full_name
            FROM users
            WHERE student_id ILIKE %s
            LIMIT 10
        """, (student_id + '%',))

        return jsonify({'status': 'success', 'students': cur.fetchall()})

    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

    finally:
        cur.close()
        conn.close()

# ===============================
# CREATE & MANAGE ANNOUNCEMENTS (Teacher Only)
# ===============================
@auth.route('/create_announcement/<int:course_id>', methods=['GET', 'POST'])
def create_announcement(course_id):
    if 'user_id' not in session:
        flash('Please log in first.', 'warning')
        return redirect(url_for('auth.login'))

    if session.get('role_id') != 2:  # Teacher only
        flash('Access denied. Teachers only.', 'danger')
        return redirect(url_for('auth.login'))

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # -------------------------------
    # FETCH TEACHER INFO
    # -------------------------------
    cur.execute("SELECT * FROM users WHERE user_id = %s;", (session['user_id'],))
    user = cur.fetchone()
    teacher_name = user['full_name']

    # -------------------------------
    # FETCH COURSE INFO
    # -------------------------------
    cur.execute("""
        SELECT c.*, p.program_name, a.start_year, a.end_year, a.semester
        FROM courses c
        JOIN programs p ON c.program_id = p.program_id
        JOIN academic_years a ON c.academic_year_id = a.academic_year_id
        WHERE c.course_id = %s;
    """, (course_id,))
    course = cur.fetchone()

    # ===============================
    # HANDLE POST JSON ACTIONS
    # ===============================
    if request.method == 'POST' and request.is_json:
        data = request.get_json()
        action = data.get('action')

        try:
            # -------------------------------
            # CREATE ANNOUNCEMENT
            # -------------------------------
            if action == 'create':
                title = data.get('title')
                content = data.get('content', '')
                if not title:
                    return jsonify({'status': 'error', 'message': 'Missing title.'}), 400

                # Insert announcement
                cur.execute("""
                    INSERT INTO announcements (course_id, title, content, date_posted)
                    VALUES (%s, %s, %s, NOW())
                    RETURNING announcement_id;
                """, (course_id, title, content))
                announcement_id = cur.fetchone()['announcement_id']

                # Fetch students enrolled in this course
                cur.execute("""
                    SELECT u.user_id, u.full_name, u.email
                    FROM users u
                    JOIN courses c ON u.section_id = c.section_id
                    WHERE u.role_id = 3 AND c.course_id = %s;
                """, (course_id,))
                students = cur.fetchall()

                # Function to notify students asynchronously
                def notify_students():
                    with app.app_context():
                        for student in students:
                            # Insert notification in DB
                            cur.execute("""
                                INSERT INTO notifications (user_id, course_id, title, message, created_at, read)
                                VALUES (%s, %s, %s, %s, NOW(), FALSE)
                            """, (
                                student['user_id'],
                                course_id,
                                f"New Announcement in {course['course_code']}",
                                f"{teacher_name} posted a new announcement: '{title}'"
                            ))

                            # Send email
                            try:
                                msg = Message(
                                    subject=f"New Announcement in {course['course_code']}",
                                    recipients=[student['email']],
                                    html=f"""
                                        <p>Hi {student['full_name']},</p>
                                        <p>{teacher_name} posted a new announcement in your course <strong>{course['course_code']}</strong>:</p>
                                        <p><strong>{title}</strong></p>
                                        <p>{content}</p>
                                        <p>Check your LMS portal for more details.</p>
                                    """
                                )
                                mail.send(msg)
                            except Exception as e:
                                print(f"Failed to send email to {student['email']}: {e}")

                    conn.commit()

                Thread(target=notify_students).start()

                return jsonify({
                    'status': 'success',
                    'message': 'Announcement created and students notified!',
                    'announcement_id': announcement_id
                })

            # -------------------------------
            # EDIT ANNOUNCEMENT
            # -------------------------------
            elif action == 'edit':
                announcement_id = data.get('announcement_id')
                title = data.get('title')
                content = data.get('content', '')
                if not announcement_id or not title:
                    return jsonify({'status': 'error', 'message': 'Missing required fields.'}), 400

                cur.execute("""
                    UPDATE announcements
                    SET title = %s, content = %s
                    WHERE announcement_id = %s;
                """, (title, content, announcement_id))
                conn.commit()
                return jsonify({'status': 'success', 'message': 'Announcement updated successfully!'})

            # -------------------------------
            # DELETE ANNOUNCEMENT
            # -------------------------------
            elif action == 'delete':
                announcement_id = data.get('announcement_id')
                if not announcement_id:
                    return jsonify({'status': 'error', 'message': 'Missing announcement ID.'}), 400

                cur.execute("DELETE FROM announcements WHERE announcement_id = %s;", (announcement_id,))
                conn.commit()
                return jsonify({'status': 'success', 'message': 'Announcement deleted successfully!'})

            else:
                return jsonify({'status': 'error', 'message': 'Invalid action.'}), 400

        except Exception as e:
            conn.rollback()
            print("Error managing announcement:", e)
            return jsonify({'status': 'error', 'message': str(e)}), 500

    # -------------------------------
    # FETCH EXISTING ANNOUNCEMENTS
    # -------------------------------
    cur.execute("""
        SELECT announcement_id, title, content, date_posted
        FROM announcements
        WHERE course_id = %s
        ORDER BY date_posted DESC;
    """, (course_id,))
    announcements = cur.fetchall()

    cur.close()
    conn.close()

    # Render template with course and user info
    return render_template(
        'create_announcement.html',
        announcements=announcements,
        course_id=course_id,
        course=course,
        user=user
    )

# ===============================
# STUDENT LIST PAGE
# ===============================
@auth.route('/records', methods=['GET'])
def student_records():
    # 1️⃣ Authentication
    if 'user_id' not in session:
        flash('Please log in first.', 'warning')
        return redirect(url_for('auth.login'))

    # 2️⃣ Authorization (Admin or Teacher only)
    if session.get('role_id') not in [1, 2]:
        flash('Access denied. Admins and Teachers only.', 'danger')
        return redirect(url_for('auth.login'))

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # Get logged-in user info
    cur.execute("SELECT user_id, full_name, email, role_id FROM users WHERE user_id = %s", (session['user_id'],))
    user = cur.fetchone()

    # Get course_id from query parameter
    course_id = request.args.get('course_id', type=int)
    if not course_id:
        flash('No course selected.', 'warning')
        return redirect(url_for('dashboard'))

    # Fetch course info (name, section)
    cur.execute("SELECT * FROM courses WHERE course_id = %s", (course_id,))
    course = cur.fetchone()
    if not course:
        flash('Course not found.', 'danger')
        return redirect(url_for('dashboard'))

    # Fetch students in that course's section
    section_id = course['section_id']
    cur.execute("""
        SELECT u.user_id, u.full_name, u.email, s.section_name, student_id
        FROM users u
        JOIN sections s ON u.section_id = s.section_id
        WHERE u.role_id = 3 AND u.section_id = %s
        ORDER BY u.full_name
    """, (section_id,))
    students = cur.fetchall()

    cur.close()
    conn.close()

    # Pass user, course, course_id explicitly to template
    return render_template('records.html',
                           students=students,
                           course=course,
                           course_id=course_id,
                           user=user)

from openpyxl import Workbook
from flask import send_file
import io

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import io
from flask import send_file, flash, redirect, url_for, request
import psycopg2.extras


# ===============================
# AJAX: GET SINGLE STUDENT DETAILS
# ===============================
@auth.route('/records/<int:student_id>', methods=['GET'])
def get_student_record(student_id):
    # 1. Authentication & Authorization
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401
    
    # Allow Faculty (2) or the Student themselves (1) to view
    if session.get('role_id') == 1 and session.get('user_id') != student_id:
         return jsonify({'status': 'error', 'message': 'Access denied'}), 403

    app = create_app()
    conn = app.get_db_connection()
    # Use RealDictCursor to access columns by name
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        # A. Student Basic Info
        cur.execute("""
            SELECT u.user_id, u.full_name, u.email, s.section_name
            FROM users u
            LEFT JOIN sections s ON u.section_id = s.section_id
            WHERE u.user_id = %s
        """, (student_id,))
        student = cur.fetchone()
        
        if not student:
            return jsonify({'status': 'error', 'message': 'Student not found'}), 404

        # B. Quizzes (Join quizzes and quiz_submissions)
        cur.execute("""
            SELECT q.title, qs.total_score
            FROM quiz_submissions qs
            JOIN quizzes q ON qs.quiz_id = q.quiz_id
            WHERE qs.student_id = %s
        """, (student_id,))
        quizzes = cur.fetchall()

        # C. Assignments (Join assignments and submissions)
        cur.execute("""
            SELECT a.title, s.grade
            FROM submissions s
            JOIN assignments a ON s.assignment_id = a.assignment_id
            WHERE s.student_id = %s
        """, (student_id,))
        assignments = cur.fetchall()

        # D. Activities (FIXED: Removed duplicate fetchall)
        cur.execute("""
            SELECT la.title AS activity_name, las.grade
            FROM learning_activity_submissions las
            JOIN learning_activities la ON las.activity_id = la.activity_id
            WHERE las.student_id = %s
        """, (student_id,))
        activities = cur.fetchall() 

        # E. Calculate Grade
        quiz_scores = [q['total_score'] for q in quizzes if q['total_score'] is not None]
        assign_scores = [a['grade'] for a in assignments if a['grade'] is not None]
        activity_scores = [act['grade'] for act in activities if act['grade'] is not None]
        
        all_scores = quiz_scores + assign_scores + activity_scores
        
        overall_grade = 0
        if len(all_scores) > 0:
            overall_grade = round(sum(all_scores) / len(all_scores), 2)

        # F. Format Response
        response = {
            'name': student['full_name'],
            'section': student['section_name'] if student['section_name'] else 'No Section',
            'quizzes': [{'name': q['title'], 'score': q['total_score'] or 0} for q in quizzes],
            'assignments': [{'name': a['title'], 'score': a['grade'] or 0} for a in assignments],
            'activities': [{'name': act['activity_name'], 'score': act['grade'] or 0} for act in activities],
            'overallGrade': f"{overall_grade}%"
        }

        return jsonify(response)

    except Exception as e:
        print(f"Error fetching record for {student_id}: {e}")
        return jsonify({'status': 'error', 'message': 'Database error'}), 500
    finally:
        cur.close()
        conn.close()


@auth.route('/records/export', methods=['GET'])
def export_all_records():
    # 1. Auth Checks
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))

    course_id = request.args.get('course_id', type=int)
    if not course_id:
        return "Error: No course ID provided", 400

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # REMOVED TRY/EXCEPT BLOCK FOR DEBUGGING PURPOSES
    # If it crashes now, you will see the error on screen.

    # 2. Get Course & Section Info
    cur.execute("SELECT * FROM courses WHERE course_id = %s", (course_id,))
    course = cur.fetchone()
    if not course:
        return "Error: Course not found", 404

    section_id = course['section_id']

    # 3. Get Master Lists (Columns)
    cur.execute("SELECT quiz_id, title FROM quizzes WHERE course_id = %s ORDER BY quiz_id ASC", (course_id,))
    all_quizzes = cur.fetchall()

    cur.execute("SELECT assignment_id, title FROM assignments WHERE course_id = %s ORDER BY assignment_id ASC", (course_id,))
    all_assignments = cur.fetchall()

    cur.execute("SELECT activity_id, title FROM learning_activities WHERE course_id = %s ORDER BY activity_id ASC", (course_id,))
    all_activities = cur.fetchall()

    # 4. Get Students
    cur.execute("""
        SELECT u.user_id, u.student_id, u.full_name, s.section_name
        FROM users u
        JOIN sections s ON u.section_id = s.section_id
        WHERE u.role_id = 3 AND u.section_id = %s
        ORDER BY u.full_name
    """, (section_id,))
    students = cur.fetchall()

    # 5. Build Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Grades"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Header Row
    headers = ["Student ID", "Full Name", "Section"]
    for q in all_quizzes: headers.append(f"Q: {q['title']}")
    for a in all_assignments: headers.append(f"A: {a['title']}")
    for act in all_activities: headers.append(f"Act: {act['title']}")
    headers.append("Overall %")

    ws.append(headers)

    # Style Header
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # 6. Populate Data
    for student in students:
        sid_db = student['user_id']
        row = [student['student_id'], student['full_name'], student['section_name']]
        
        scores = []

        # Get Scores Maps (To avoid N+1 queries ideally, but this is safe for now)
        def get_score(query, item_id):
            cur.execute(query, (sid_db, item_id))
            res = cur.fetchone()
            return res['score'] if res and res['score'] is not None else 0

        for q in all_quizzes:
            s = get_score("SELECT total_score as score FROM quiz_submissions WHERE student_id=%s AND quiz_id=%s", q['quiz_id'])
            row.append(s)
            scores.append(s)

        for a in all_assignments:
            s = get_score("SELECT grade as score FROM submissions WHERE student_id=%s AND assignment_id=%s", a['assignment_id'])
            row.append(s)
            scores.append(s)

        for act in all_activities:
            s = get_score("SELECT grade as score FROM learning_activity_submissions WHERE student_id=%s AND activity_id=%s", act['activity_id'])
            row.append(s)
            scores.append(s)

        # Average
        avg = round(sum(scores) / len(scores), 2) if len(scores) > 0 else 0
        row.append(avg)

        ws.append(row)

    # Cleanup DB
    cur.close()
    conn.close()

    # 7. Prepare Download
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{course['course_code']}_Grades.xlsx"

    # FLASK COMPATIBILITY FIX
    try:
        return send_file(
            output,
            as_attachment=True,
            download_name=filename, # Flask 2.0+
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except TypeError:
        return send_file(
            output,
            as_attachment=True,
            attachment_filename=filename, # Flask < 2.0
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

@auth.route('/course_forum/<int:course_id>')
def course_forum(course_id):
    if 'user_id' not in session:
        flash('Please log in.', 'warning')
        return redirect(url_for('auth.login'))

    user_id = session['user_id']  # <-- logged-in user
    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        # 1️⃣ Fetch logged-in user info
        cur.execute("SELECT user_id, full_name, email, role_id FROM users WHERE user_id = %s", (user_id,))
        user = cur.fetchone()

        # 2️⃣ Get Course Info
        cur.execute("SELECT course_code, course_title FROM courses WHERE course_id = %s", (course_id,))
        course = cur.fetchone()

        if not course:
            flash('Course not found.', 'danger')
            return redirect(url_for('auth.materials', course_id=course_id))

        # 3️⃣ Get Topics
        cur.execute("""
            SELECT t.topic_id, t.title, t.content, t.created_at, t.is_pinned,
                   u.full_name, u.role_id,
                   (SELECT COUNT(*) FROM forum_replies r WHERE r.topic_id = t.topic_id) as reply_count
            FROM forum_topics t
            JOIN users u ON t.user_id = u.user_id
            WHERE t.course_id = %s
            ORDER BY t.is_pinned DESC, t.created_at DESC
        """, (course_id,))
        topics = cur.fetchall()

    except Exception as e:
        print(f"FORUM ERROR: {e}") 
        flash(f'Error loading forum: {str(e)}', 'danger')
        topics = []
    finally:
        cur.close()
        conn.close()

    return render_template(
        'forum_index.html',
        user=user,          # <-- pass logged-in user to template
        course=course,
        topics=topics,
        course_id=course_id
    )

# ========================================================
# 2. VIEW SINGLE TOPIC (Read thread + replies)
# ========================================================
@auth.route('/forum_topic/<int:course_id>/<int:topic_id>')
def view_topic(course_id, topic_id):
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        # 1. Fetch Main Topic
        # REMOVED 'u.profile_picture'
        cur.execute("""
            SELECT t.topic_id, t.title, t.content, t.created_at, t.course_id,
                   u.full_name, u.role_id
            FROM forum_topics t
            JOIN users u ON t.user_id = u.user_id
            WHERE t.topic_id = %s
        """, (topic_id,))
        topic = cur.fetchone()

        if not topic:
            flash('Topic not found.', 'danger')
            return redirect(url_for('auth.course_forum', course_id=course_id))

        # 2. Fetch Replies
        # REMOVED 'u.profile_picture'
        cur.execute("""
            SELECT r.reply_id, r.content, r.created_at,
                   u.full_name, u.role_id, r.user_id
            FROM forum_replies r
            JOIN users u ON r.user_id = u.user_id
            WHERE r.topic_id = %s
            ORDER BY r.created_at ASC
        """, (topic_id,))
        replies = cur.fetchall()

    except Exception as e:
        print(f"TOPIC ERROR: {e}")
        flash(f"Error loading topic: {str(e)}", 'danger')
        return redirect(url_for('auth.course_forum', course_id=course_id))
    finally:
        cur.close()
        conn.close()

    return render_template('forum_topic.html', topic=topic, replies=replies, course_id=course_id, current_user_id=session['user_id'])


# ========================================================
# 3. CREATE NEW TOPIC (POST)
# ========================================================
@auth.route('/create_topic/<int:course_id>', methods=['POST'])
def create_topic(course_id):
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))

    title = request.form.get('title')
    content = request.form.get('content')
    user_id = session['user_id']
    
    # Only teachers (Role 2) can pin
    is_pinned = False
    if session.get('role_id') == 2: 
        is_pinned = (request.form.get('is_pinned') == 'on')

    if not title or not content:
        flash('Title and Content are required.', 'warning')
        return redirect(url_for('auth.course_forum', course_id=course_id))

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            INSERT INTO forum_topics (course_id, user_id, title, content, created_at, is_pinned)
            VALUES (%s, %s, %s, %s, NOW(), %s)
        """, (course_id, user_id, title, content, is_pinned))
        conn.commit()
        flash('Topic created successfully!', 'success')
    except Exception as e:
        conn.rollback()
        print(f"CREATE ERROR: {e}")
        flash(f'Error creating topic: {e}', 'danger')
    finally:
        cur.close()
        conn.close()

    return redirect(url_for('auth.course_forum', course_id=course_id))

@auth.route('/edit_topic/<int:course_id>/<int:topic_id>', methods=['POST'])
def edit_topic(course_id, topic_id):
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))

    # Only title, content, and pinned status are needed
    title = request.form.get('edit_title')
    content = request.form.get('edit_content')
    is_pinned = (request.form.get('edit_is_pinned') == 'on') if session.get('role_id') == 2 else False

    if not title or not content:
        flash('Title and Content are required.', 'warning')
        return redirect(url_for('auth.course_forum', course_id=course_id))

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("""
            UPDATE forum_topics
            SET title = %s, content = %s, is_pinned = %s
            WHERE topic_id = %s
        """, (title, content, is_pinned, topic_id))
        conn.commit()
        flash('Topic updated successfully!', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Error updating topic: {e}', 'danger')
    finally:
        cur.close()
        conn.close()

    return redirect(url_for('auth.course_forum', course_id=course_id))


# ========================================================
# DELETE TOPIC
# ========================================================
@auth.route('/delete_topic/<int:topic_id>', methods=['POST'])
def delete_topic(topic_id):
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))

    user_id = session['user_id']
    role_id = session.get('role_id')
    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    course_id = None
    try:
        # Fetch topic
        cur.execute("SELECT course_id, user_id FROM forum_topics WHERE topic_id = %s", (topic_id,))
        topic = cur.fetchone()
        if not topic:
            flash('Topic not found.', 'danger')
            return redirect(request.referrer or url_for('auth.course_forum', course_id=1))

        course_id = topic['course_id']

        # Only owner or teacher
        if user_id != topic['user_id'] and role_id != 2:
            flash('You do not have permission to delete this topic.', 'danger')
        else:
            cur.execute("DELETE FROM forum_topics WHERE topic_id = %s", (topic_id,))
            conn.commit()
            flash('Topic deleted successfully!', 'success')

        # Re-fetch forum topics
        cur.execute("""
            SELECT t.topic_id, t.title, t.content, t.created_at, t.is_pinned,
                   u.role_id,
                   (SELECT COUNT(*) FROM forum_replies r WHERE r.topic_id = t.topic_id) as reply_count
            FROM forum_topics t
            JOIN users u ON t.user_id = u.user_id
            WHERE t.course_id = %s
            ORDER BY t.is_pinned DESC, t.created_at DESC
        """, (course_id,))
        topics = cur.fetchall()

        # Fetch course info
        cur.execute("SELECT course_code, course_title FROM courses WHERE course_id = %s", (course_id,))
        course = cur.fetchone()

    finally:
        cur.close()
        conn.close()

    return render_template('forum_index.html', user={'user_id': user_id, 'role_id': role_id}, 
                           course=course, topics=topics, course_id=course_id)


# ========================================================
# 4. REPLY TO TOPIC (POST)
# ========================================================
@auth.route('/reply_topic/<int:course_id>/<int:topic_id>', methods=['POST'])
def reply_topic(course_id, topic_id):
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))

    content = request.form.get('content')
    user_id = session['user_id']

    if not content:
        flash('Reply cannot be empty.', 'warning')
        return redirect(url_for('auth.view_topic', course_id=course_id, topic_id=topic_id))

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            INSERT INTO forum_replies (topic_id, user_id, content, created_at)
            VALUES (%s, %s, %s, NOW())
        """, (topic_id, user_id, content))
        conn.commit()
        flash('Reply posted.', 'success')
    except Exception as e:
        conn.rollback()
        print(f"REPLY ERROR: {e}")
        flash(f'Error posting reply: {e}', 'danger')
    finally:
        cur.close()
        conn.close()

    return redirect(url_for('auth.view_topic', course_id=course_id, topic_id=topic_id))


# ========================================================
# 5. DELETE TOPIC/REPLY
# ========================================================
@auth.route('/delete_forum_item/<int:course_id>/<string:item_type>/<int:item_id>')
def delete_forum_item(course_id, item_type, item_id):
    if session.get('role_id') != 2: 
        flash('Access denied.', 'danger')
        return redirect(url_for('auth.course_forum', course_id=course_id))

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor()

    try:
        if item_type == 'topic':
            cur.execute("DELETE FROM forum_topics WHERE topic_id = %s", (item_id,))
            msg = "Topic deleted."
            next_url = url_for('auth.course_forum', course_id=course_id)
        elif item_type == 'reply':
            cur.execute("DELETE FROM forum_replies WHERE reply_id = %s RETURNING topic_id", (item_id,))
            res = cur.fetchone()
            topic_id = res[0] if res else None
            msg = "Reply deleted."
            next_url = url_for('auth.view_topic', course_id=course_id, topic_id=topic_id) if topic_id else url_for('auth.course_forum', course_id=course_id)

        conn.commit()
        flash(msg, 'success')
    except Exception as e:
        conn.rollback()
        print(f"DELETE ERROR: {e}")
        flash('Error deleting item.', 'danger')
        next_url = url_for('auth.course_forum', course_id=course_id)
    finally:
        cur.close()
        conn.close()

    return redirect(next_url)

# ========================================================
# VIEW FORUM (Student)
# ========================================================
@auth.route('/course_forum_student/<int:course_id>')
def course_forum_student(course_id):
    # 1️⃣ Authentication
    if 'user_id' not in session:
        flash('Please log in.', 'warning')
        return redirect(url_for('auth.login'))

    user_id = session['user_id']  # logged-in user

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        # 2️⃣ Fetch logged-in user info
        cur.execute("""
            SELECT user_id, full_name, email, role_id
            FROM users
            WHERE user_id = %s
        """, (user_id,))
        user = cur.fetchone()

        # 3️⃣ Fetch course info (include course_id and section_id)
        cur.execute("""
            SELECT course_id, course_code, course_title, section_id
            FROM courses
            WHERE course_id = %s
        """, (course_id,))
        course = cur.fetchone()

        # 4️⃣ Convert section_id to section_name
        if course and course.get('section_id'):
            cur.execute("""
                SELECT section_name
                FROM sections
                WHERE section_id = %s
            """, (course['section_id'],))
            section_row = cur.fetchone()
            course['section_name'] = section_row['section_name'] if section_row else "No Section"
        else:
            course['section_name'] = "No Section"


        if not course:
            flash('Course not found.', 'danger')
            return redirect(url_for('auth.materials', course_id=course_id))

        # 4️⃣ Fetch all forum topics for this course
        cur.execute("""
            SELECT 
                t.topic_id, t.title, t.content, t.created_at, t.is_pinned,
                u.full_name, u.role_id,
                (SELECT COUNT(*) 
                 FROM forum_replies r 
                 WHERE r.topic_id = t.topic_id) AS reply_count
            FROM forum_topics t
            JOIN users u ON t.user_id = u.user_id
            WHERE t.course_id = %s
            ORDER BY t.is_pinned DESC, t.created_at DESC
        """, (course_id,))
        topics = cur.fetchall()

    except Exception as e:
        print(f"FORUM ERROR: {e}")
        flash(f'Error loading forum: {str(e)}', 'danger')
        topics = []
    finally:
        cur.close()
        conn.close()

    # 5️⃣ Pass all necessary variables to template
    return render_template(
        'forum_index_student.html',
        user=user,          # logged-in user info
        course=course,      # includes course_id, course_code, course_title
        topics=topics,
        course_id=course_id # in case template uses it directly
    )

# ========================================================
# 2. VIEW SINGLE TOPIC (Read thread + replies)
# ========================================================
@auth.route('/forum_topic_student/<int:course_id>/<int:topic_id>')
def view_topic_student(course_id, topic_id):
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        # 1. Fetch Main Topic
        # REMOVED 'u.profile_picture'
        cur.execute("""
            SELECT t.topic_id, t.title, t.content, t.created_at, t.course_id,
                   u.full_name, u.role_id
            FROM forum_topics t
            JOIN users u ON t.user_id = u.user_id
            WHERE t.topic_id = %s
        """, (topic_id,))
        topic = cur.fetchone()

        if not topic:
            flash('Topic not found.', 'danger')
            return redirect(url_for('auth.course_forum_student', course_id=course_id))

        # 2. Fetch Replies
        # REMOVED 'u.profile_picture'
        cur.execute("""
            SELECT r.reply_id, r.content, r.created_at,
                   u.full_name, u.role_id, r.user_id
            FROM forum_replies r
            JOIN users u ON r.user_id = u.user_id
            WHERE r.topic_id = %s
            ORDER BY r.created_at ASC
        """, (topic_id,))
        replies = cur.fetchall()

    except Exception as e:
        print(f"TOPIC ERROR: {e}")
        flash(f"Error loading topic: {str(e)}", 'danger')
        return redirect(url_for('auth.course_forum_student', course_id=course_id))
    finally:
        cur.close()
        conn.close()

    return render_template('forum_topic_student.html', topic=topic, replies=replies, course_id=course_id, current_user_id=session['user_id'])

# ========================================================
# 4. REPLY TO TOPIC (POST)
# ========================================================
@auth.route('/reply_topic_student/<int:course_id>/<int:topic_id>', methods=['POST'])
def reply_topic_student(course_id, topic_id):
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))

    content = request.form.get('content')
    user_id = session['user_id']

    if not content:
        flash('Reply cannot be empty.', 'warning')
        return redirect(url_for('auth.view_topic_student', course_id=course_id, topic_id=topic_id))

    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            INSERT INTO forum_replies (topic_id, user_id, content, created_at)
            VALUES (%s, %s, %s, NOW())
        """, (topic_id, user_id, content))
        conn.commit()
        flash('Reply posted.', 'success')
    except Exception as e:
        conn.rollback()
        print(f"REPLY ERROR: {e}")
        flash(f'Error posting reply: {e}', 'danger')
    finally:
        cur.close()
        conn.close()

    return redirect(url_for('auth.view_topic_student', course_id=course_id, topic_id=topic_id))



@auth.route('/meeting')
def meeting():
    # ===============================
    # Authentication
    # ===============================
    if 'user_id' not in session:
        flash('Please log in first.', 'warning')
        return redirect(url_for('auth.login'))

    allowed_roles = [1, 2, 3]  # example: 1=admin, 2=teacher, 3=student
    if session.get('role_id') not in allowed_roles:
        flash('Access denied. You do not have permission.', 'danger')
        return redirect(url_for('auth.login'))
    # ===============================
    # Database connection
    # ===============================
    app = create_app()
    conn = app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        # ===============================
        # Fetch user info
        # ===============================
        cur.execute("""
            SELECT user_id, full_name, email, role_id, program_id
            FROM users
            WHERE user_id = %s
        """, (session['user_id'],))
        user_info = cur.fetchone()

    finally:
        cur.close()
        conn.close()

    return render_template('meeting.html', user=user_info)

from flask import render_template, request, redirect, url_for, flash, session, current_app
from werkzeug.security import generate_password_hash
import psycopg2.extras

import os
from werkzeug.utils import secure_filename

# CONFIGURATION (Add this near the top of your file or in config.py)
UPLOAD_FOLDER = 'static/uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@auth.route('/edit_profile', methods=['GET', 'POST'])
def edit_profile():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))

    user_id = session['user_id']
    conn = current_app.get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        # GET: Fetch user data
        if request.method == 'GET':
            # Fetched profile_image column here
            cur.execute("SELECT full_name, email, profile_image FROM users WHERE user_id = %s", (user_id,))
            user = cur.fetchone()
            return render_template('edit_profile.html', user=user)

        # POST: Update Data
        if request.method == 'POST':
            full_name = request.form.get('full_name')
            email = request.form.get('email')
            password = request.form.get('password')
            confirm_password = request.form.get('confirm_password')

            # 1. HANDLE IMAGE UPLOAD
            filename = None
            if 'profile_image' in request.files:
                file = request.files['profile_image']
                if file and file.filename != '' and allowed_file(file.filename):
                    # Create directory if not exists
                    if not os.path.exists(UPLOAD_FOLDER):
                        os.makedirs(UPLOAD_FOLDER)
                    
                    # Secure filename and prepend user_id to prevent overwrite/conflicts
                    ext = file.filename.rsplit('.', 1)[1].lower()
                    filename = f"user_{user_id}.{ext}"
                    file.save(os.path.join(UPLOAD_FOLDER, filename))

            # 2. UPDATE DATABASE
            # We build the query dynamically based on whether password/image changed
            update_fields = ["full_name = %s", "email = %s"]
            params = [full_name, email]

            if password:
                if password != confirm_password:
                    flash('Passwords do not match!', 'danger')
                    return redirect(url_for('auth.edit_profile'))
                update_fields.append("password_hash = %s")
                params.append(generate_password_hash(password))

            if filename:
                update_fields.append("profile_image = %s")
                params.append(filename)

            params.append(user_id) # For WHERE clause

            query = f"UPDATE users SET {', '.join(update_fields)} WHERE user_id = %s"
            
            cur.execute(query, tuple(params))
            conn.commit()

            session['full_name'] = full_name
            # Update session image if changed so navbar updates immediately
            if filename:
                session['profile_image'] = filename 

            flash('Profile updated successfully.', 'success')
            return redirect(url_for('auth.edit_profile'))

    except Exception as e:
        conn.rollback()
        print(f"Error: {e}")
        flash('An error occurred.', 'danger')
        return redirect(url_for('auth.edit_profile'))
    finally:
        cur.close()
        conn.close()

    

@auth.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('auth.login'))

@auth.route('/')
def home():
    return render_template('home.html')

@auth.route('/error')
def error():
    return render_template('test.html')

if __name__ == '__main__':
    app.run(debug=True)